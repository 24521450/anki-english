"""Bilingual semantic-audit contracts and XLSX round-trip helpers."""
from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.deck_builder.card_identity import primary_list_from_tags
from src.deck_builder.example_policy import main_example_pos_shortfall
from src.deck_builder.simplify_senses import _flatten_senses
from src.deck_builder.source_sense_identity import source_sense_id
from src.deck_builder.text_integrity import has_suspected_lossy_unicode
from src.deck_builder.word_lookup import get_word_candidates


AUDIT_SCHEMA_VERSION = 1
REVIEWER = "chatgpt-5.6-sol-ultra"
REVIEW_BATCH_SIZE = 100
REGISTER_RE = re.compile(r"^(?:\[[^]]+\])+")
CHECK_VALUES = ("pending", "pass", "repair", "uncertain", "not_applicable")
CHECK_FIELDS = (
    "english_semantics",
    "vietnamese_semantics",
    "simplicity",
    "example_pos_alignment",
)
DECISION_VALUES = ("pending", "pass", "repair_proposed", "uncertain", "not_applicable")
APPROVAL_VALUES = ("", "approved", "rejected")
CAMBRIDGE_MATCH_VALUES = ("pending", "exact", "partial", "missing", "conflict")

IMMUTABLE_COLUMNS = (
    "row_id", "batch_id", "guid", "word", "card_pos", "cefr", "list", "variant",
    "sense_index", "semantic_sense_id", "candidate_source_sense_ids", "candidate_source_definitions",
    "candidate_source_examples", "current_en", "current_vi", "current_examples",
    "cambridge_url", "row_fingerprint",
)
EDITABLE_COLUMNS = (
    "cambridge_match", "cambridge_summary", "translation_provenance",
    "english_check", "vietnamese_check", "simplicity_check", "example_pos_check",
    "decision", "proposed_en", "proposed_vi", "proposed_examples", "confidence",
    "review_reason", "reviewer", "reviewed_at", "approval",
)
REVIEW_COLUMNS = IMMUTABLE_COLUMNS + EDITABLE_COLUMNS
COVERAGE_COLUMNS = (
    "coverage_row_id", "batch_id", "guid", "word", "card_pos", "card_cefr", "variant",
    "source_sense_id", "source", "source_pos", "source_cefr", "source_definition",
    "source_examples", "semantic_sense_ids_available", "disposition",
    "target_semantic_sense_ids", "reason", "row_fingerprint",
)
COVERAGE_DISPOSITIONS = ("pending", "mapped", "excluded")
EXPLICIT_DUPLICATE_EXCLUSION_RE = re.compile(
    r"\b(?:pos|part[- ]of[- ]speech|homonym|variant|clon(?:e|ed)|pollution|"
    r"mislabell?ed|idiom|content|wrong card|related (?:headword|lemma))\b",
    re.IGNORECASE,
)


def _digest(value: object, length: int = 24) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length]


def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def serialize_jsonl(rows: Iterable[dict]) -> str:
    return "".join(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n" for row in rows)


def split_definition_chunk(chunk: str) -> tuple[str, str]:
    """Split the deck convention ``English (Vietnamese)`` at the final pair."""
    value = (chunk or "").strip()
    if not value.endswith(")"):
        return value, ""
    depth = 0
    for index in range(len(value) - 1, -1, -1):
        if value[index] == ")":
            depth += 1
        elif value[index] == "(":
            depth -= 1
            if depth == 0:
                english = value[:index].strip()
                vietnamese = value[index + 1 : -1].strip()
                if english and vietnamese:
                    return english, vietnamese
                break
            if depth < 0:
                break
    return value, ""


def _normalized_source_definition(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def _source_senses(records: list[dict], word: str) -> list[dict]:
    candidates = {candidate.casefold() for candidate in get_word_candidates(word.casefold())}
    senses: list[dict] = []
    for record in records:
        if (record.get("word") or "").casefold() not in candidates:
            continue
        for flat in _flatten_senses(record):
            definition = record["pos_data"][flat.pd_idx]["definitions"][flat.def_idx]
            examples = [
                item.get("text", "") if isinstance(item, dict) else str(item)
                for item in definition.get("examples") or []
            ]
            senses.append({
                "source_sense_id": source_sense_id(record, flat),
                "source": "Cambridge" if (record.get("source") or "").casefold() == "cambridge" else "Oxford",
                "source_url": record.get("source_url") or "",
                "source_files": record.get("source_files") or [],
                "homonym_index": record.get("homonym_index"),
                "pos": flat.pos,
                "cefr_original": flat.cefr_original,
                "cefr_resolved": flat.cefr_resolved or "UNCLASSIFIED",
                "cefr_source": flat.cefr_source,
                "sensenum_local": definition.get("sensenum_local"),
                "definition": definition.get("text") or "",
                "examples": [text for text in examples if text],
                "register_tags": definition.get("register_tags") or [],
                "domain": definition.get("domain"),
            })
    return senses


def _semantic_senses(card: dict) -> list[dict]:
    definitions = (card.get("definition") or "").split("|") if card.get("definition") else []
    definitions_vi = (
        (card.get("definition_vi") or "").split("|")
        if card.get("definition_vi")
        else []
    )
    examples = (card.get("example") or "").split("|") if card.get("example") else []
    senses: list[dict] = []
    for index, chunk in enumerate(definitions, 1):
        english, legacy_vietnamese = split_definition_chunk(chunk)
        vietnamese = (
            definitions_vi[index - 1].strip()
            if index <= len(definitions_vi) and definitions_vi[index - 1].strip()
            else legacy_vietnamese
        )
        current_examples = [
            part.strip()
            for part in re.split(r"<br\s*/?><br\s*/?>", examples[index - 1] if index <= len(examples) else "", flags=re.I)
            if part.strip()
        ]
        senses.append({
            "semantic_sense_id": "sem_" + _digest({"guid": card.get("guid"), "index": index, "en": english}),
            "order": index,
            "source_sense_ids": [],
            "current": {"definition_en": english, "definition_vi": vietnamese, "examples": current_examples},
            "checks": {
                "english_semantics": "pending",
                "vietnamese_semantics": "pending",
                "simplicity": "pending",
                "example_pos_alignment": "pending",
            },
            "decision": "pending",
            "proposed": {"definition_en": "", "definition_vi": "", "examples": []},
            "cambridge": {
                "url": f"https://dictionary.cambridge.org/dictionary/english-vietnamese/{word_slug(card.get('word') or '')}",
                "match": "pending",
                "summary": "",
                "translation_provenance": "",
                "accessed_at": "",
            },
            "confidence": "",
            "review_reason": "",
            "reviewer": "",
            "reviewed_at": "",
            "approval": "",
        })
    return senses


def semantic_sense_id(guid: str, order: int, definition_en: str) -> str:
    """Return the stable ID used when a reviewer adds a missing semantic sense."""
    return "sem_" + _digest({"guid": guid, "index": order, "en": definition_en})


def _added_semantic_sense(guid: str, update: dict) -> dict:
    proposed = update.get("proposed") or {}
    definition_en = proposed.get("definition_en") or ""
    order = int(update.get("order") or 0)
    if not definition_en or order < 1:
        raise ValueError(f"Added semantic sense requires order and proposed English: {guid}")
    semantic_id = update.get("semantic_sense_id") or semantic_sense_id(guid, order, definition_en)
    return {
        "semantic_sense_id": semantic_id,
        "order": order,
        "source_sense_ids": [],
        "current": {"definition_en": "", "definition_vi": "", "examples": []},
        "checks": dict(update.get("checks") or {}),
        "decision": update.get("decision") or "repair_proposed",
        "proposed": {
            "definition_en": definition_en,
            "definition_vi": proposed.get("definition_vi") or "",
            "examples": proposed.get("examples") or [],
        },
        "cambridge": {
            "url": f"https://dictionary.cambridge.org/dictionary/english-vietnamese/{word_slug(update.get('word') or '')}",
            "match": "pending",
            "summary": "",
            "translation_provenance": "",
            "accessed_at": "",
            **dict(update.get("cambridge") or {}),
        },
        "confidence": update.get("confidence") or "",
        "review_reason": update.get("review_reason") or "",
        "reviewer": update.get("reviewer") or REVIEWER,
        "reviewed_at": update.get("reviewed_at") or date.today().isoformat(),
        "approval": update.get("approval") or "",
    }


def word_slug(word: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", word.casefold()).strip("-")


def build_audit_rows(
    cards: list[dict],
    registry_rows: list[dict],
    oxford_records: list[dict],
    cambridge_records: list[dict] | None = None,
) -> list[dict]:
    registry_by_guid = {row.get("guid"): row for row in registry_rows if row.get("status") == "active"}
    rows: list[dict] = []
    for card in cards:
        registry = registry_by_guid.get(card.get("guid"))
        if registry is None:
            continue
        word = registry.get("word") or card.get("word") or ""
        raw_senses = _source_senses(oxford_records, word)
        raw_senses.extend(_source_senses(cambridge_records or [], word))
        pos_parts = {part.strip().casefold() for part in (registry.get("pos") or card.get("pos") or "").split(",") if part.strip()}
        candidates = [sense for sense in raw_senses if sense["pos"].casefold() in pos_parts]
        exact_cefr = [
            sense for sense in raw_senses
            if sense["pos"].casefold() in pos_parts and sense["cefr_resolved"] == (registry.get("cefr") or "UNCLASSIFIED")
        ]
        semantic = _semantic_senses(card)
        idiom_only = not semantic and bool(card.get("idioms"))
        source_fingerprint = _digest(raw_senses, 64)
        rows.append({
            "schema_version": AUDIT_SCHEMA_VERSION,
            "guid": card.get("guid") or "",
            "word": registry.get("word") or card.get("word") or "",
            "cefr": registry.get("cefr") or card.get("cefr") or "UNCLASSIFIED",
            "list": registry.get("list") or primary_list_from_tags(card.get("tags"), canonical=True),
            "variant": registry.get("variant") or "",
            "pos": registry.get("pos") or card.get("pos") or "",
            "current": {
                "definition": card.get("definition") or "",
                "example": card.get("example") or "",
                "idioms": card.get("idioms") or "",
            },
            "source_fingerprint": source_fingerprint,
            "source_senses": raw_senses,
            "coverage": {
                "status": "not_applicable" if idiom_only else "pending",
                "reason": "idiom_only" if idiom_only else "",
                "candidate_source_sense_ids": [sense["source_sense_id"] for sense in candidates],
                "expected_same_cefr_source_sense_ids": [sense["source_sense_id"] for sense in exact_cefr],
            },
            "source_coverage": [
                {"source_sense_id": sense["source_sense_id"], "disposition": "pending", "target_semantic_sense_ids": [], "reason": ""}
                for sense in candidates
            ],
            "semantic_senses": semantic,
        })
    return rows


def flatten_review_rows(audit_rows: list[dict]) -> list[dict]:
    flat: list[dict] = []
    ordinal = 0
    for card in audit_rows:
        senses = card.get("semantic_senses") or []
        if not senses:
            senses = [{
                "semantic_sense_id": "",
                "order": 0,
                "source_sense_ids": [],
                "current": {"definition_en": "", "definition_vi": "", "examples": []},
                "checks": {"english_semantics": "not_applicable", "vietnamese_semantics": "not_applicable", "simplicity": "not_applicable", "example_pos_alignment": "not_applicable"},
                "decision": "not_applicable",
                "proposed": {"definition_en": "", "definition_vi": "", "examples": []},
                "cambridge": {"url": "", "match": "missing", "summary": "", "translation_provenance": "", "accessed_at": ""},
                "confidence": "", "review_reason": card.get("coverage", {}).get("reason", ""), "reviewer": "", "reviewed_at": "", "approval": "",
            }]
        source_by_id = {item["source_sense_id"]: item for item in card.get("source_senses") or []}
        eligible_ids = card.get("coverage", {}).get("candidate_source_sense_ids") or []
        for sense in senses:
            ordinal += 1
            candidate_items = [source_by_id[item] for item in eligible_ids if item in source_by_id]
            row_id = f"{card['guid']}::{sense.get('semantic_sense_id') or 'card'}"
            immutable = {
                "row_id": row_id,
                "batch_id": f"B{((ordinal - 1) // REVIEW_BATCH_SIZE) + 1:03d}",
                "guid": card["guid"], "word": card["word"], "card_pos": card["pos"], "cefr": card["cefr"],
                "list": card["list"], "variant": card["variant"], "sense_index": sense.get("order", 0),
                "semantic_sense_id": sense.get("semantic_sense_id", ""),
                "candidate_source_sense_ids": "|".join(eligible_ids),
                "candidate_source_definitions": " || ".join(item.get("definition") or "" for item in candidate_items),
                "candidate_source_examples": " || ".join(" / ".join(item.get("examples") or []) for item in candidate_items),
                "current_en": sense["current"]["definition_en"], "current_vi": sense["current"]["definition_vi"],
                "current_examples": " || ".join(sense["current"].get("examples") or []),
                "cambridge_url": sense.get("cambridge", {}).get("url", ""),
            }
            fingerprint = _digest(immutable, 64)
            flat.append({
                **immutable,
                "row_fingerprint": fingerprint,
                "cambridge_match": sense.get("cambridge", {}).get("match", "pending"),
                "cambridge_summary": sense.get("cambridge", {}).get("summary", ""),
                "translation_provenance": sense.get("cambridge", {}).get("translation_provenance", ""),
                "english_check": sense["checks"].get("english_semantics", "pending"),
                "vietnamese_check": sense["checks"].get("vietnamese_semantics", "pending"),
                "simplicity_check": sense["checks"].get("simplicity", "pending"),
                "example_pos_check": sense["checks"].get("example_pos_alignment", "pending"),
                "decision": sense.get("decision", "pending"),
                "proposed_en": sense.get("proposed", {}).get("definition_en", ""),
                "proposed_vi": sense.get("proposed", {}).get("definition_vi", ""),
                "proposed_examples": " || ".join(sense.get("proposed", {}).get("examples") or []),
                "confidence": sense.get("confidence", ""), "review_reason": sense.get("review_reason", ""),
                "reviewer": sense.get("reviewer", ""), "reviewed_at": sense.get("reviewed_at", ""),
                "approval": sense.get("approval", ""),
            })
    return flat


def flatten_coverage_rows(audit_rows: list[dict]) -> list[dict]:
    flat: list[dict] = []
    ordinal = 0
    for card in audit_rows:
        source_by_id = {item["source_sense_id"]: item for item in card.get("source_senses") or []}
        semantic_ids = [sense.get("semantic_sense_id") for sense in card.get("semantic_senses") or []]
        for coverage in card.get("source_coverage") or []:
            ordinal += 1
            source = source_by_id[coverage["source_sense_id"]]
            immutable = {
                "coverage_row_id": f"{card['guid']}::{source['source_sense_id']}",
                "batch_id": f"B{((ordinal - 1) // REVIEW_BATCH_SIZE) + 1:03d}",
                "guid": card["guid"], "word": card["word"], "card_pos": card["pos"], "card_cefr": card["cefr"],
                "variant": card["variant"], "source_sense_id": source["source_sense_id"], "source": source["source"],
                "source_pos": source["pos"], "source_cefr": source["cefr_resolved"],
                "source_definition": source["definition"], "source_examples": " || ".join(source.get("examples") or []),
                "semantic_sense_ids_available": "|".join(semantic_ids),
            }
            flat.append({
                **immutable,
                "disposition": coverage.get("disposition", "pending"),
                "target_semantic_sense_ids": "|".join(coverage.get("target_semantic_sense_ids") or []),
                "reason": coverage.get("reason", ""),
                "row_fingerprint": _digest(immutable, 64),
            })
    return flat


def validate_audit_rows(rows: list[dict], registry_rows: list[dict] | None = None, *, require_complete: bool = False) -> list[str]:
    errors: list[str] = []
    seen_guid: set[str] = set()
    seen_semantic_ids: set[str] = set()
    mapped_source_owners: dict[str, str] = {}
    active_by_guid = {
        row.get("guid"): row
        for row in registry_rows or []
        if row.get("status") == "active"
    }
    active_guids = set(active_by_guid)
    for card in rows:
        guid = card.get("guid") or ""
        if not guid or guid in seen_guid:
            errors.append(f"duplicate_or_empty_guid:{guid}")
        seen_guid.add(guid)
        expected_identity = active_by_guid.get(guid)
        if expected_identity is not None:
            for field in ("word", "cefr", "list", "variant", "pos"):
                if (card.get(field) or "") != (expected_identity.get(field) or ""):
                    errors.append(f"identity_mismatch:{guid}:{field}")
        source_ids = [item.get("source_sense_id") for item in card.get("source_senses") or []]
        if len(source_ids) != len(set(source_ids)):
            errors.append(f"duplicate_source_sense_id:{guid}")
        known = set(source_ids)
        semantic_senses = card.get("semantic_senses") or []
        current = card.get("current") or {}
        idiom_only_shape = (
            card.get("coverage", {}).get("reason") == "idiom_only"
            and not current.get("definition")
            and not current.get("example")
            and bool(current.get("idioms"))
        )
        if idiom_only_shape and semantic_senses:
            errors.append(f"idiom_only_has_semantic_senses:{guid}")
        semantic_id_list = [sense.get("semantic_sense_id") for sense in semantic_senses]
        if not all(semantic_id_list) or len(semantic_id_list) != len(set(semantic_id_list)):
            errors.append(f"duplicate_or_empty_semantic_sense_id:{guid}")
        for semantic_id in semantic_id_list:
            if not semantic_id:
                continue
            if semantic_id in seen_semantic_ids:
                errors.append(f"duplicate_semantic_sense_id:{semantic_id}")
            seen_semantic_ids.add(semantic_id)
        orders = [sense.get("order") for sense in semantic_senses]
        if any(not isinstance(order, int) or order < 1 for order in orders) or len(orders) != len(set(orders)):
            errors.append(f"invalid_semantic_sense_order:{guid}")
        mapped: list[str] = []
        for sense in semantic_senses:
            mapped.extend(sense.get("source_sense_ids") or [])
            decision = sense.get("decision")
            if decision not in DECISION_VALUES:
                errors.append(f"invalid_decision:{guid}:{decision}")
            checks = sense.get("checks") or {}
            if set(checks) != set(CHECK_FIELDS):
                errors.append(f"invalid_check_set:{guid}:{sense.get('semantic_sense_id')}")
            if any(value not in CHECK_VALUES for value in checks.values()):
                errors.append(f"invalid_check:{guid}")
            if decision == "repair_proposed" and not any((sense.get("proposed") or {}).values()):
                errors.append(f"empty_repair:{guid}:{sense.get('semantic_sense_id')}")
            check_values = set(checks.values())
            if decision == "repair_proposed" and "repair" not in check_values:
                errors.append(f"repair_without_repair_check:{guid}:{sense.get('semantic_sense_id')}")
            if decision == "pass" and not check_values.issubset({"pass", "not_applicable"}):
                errors.append(f"pass_with_open_check:{guid}:{sense.get('semantic_sense_id')}")
            if decision == "uncertain" and "uncertain" not in check_values:
                errors.append(f"uncertain_without_uncertain_check:{guid}:{sense.get('semantic_sense_id')}")
            for location in ("current", "proposed"):
                vietnamese = str((sense.get(location) or {}).get("definition_vi") or "")
                if has_suspected_lossy_unicode(vietnamese):
                    errors.append(
                        f"corrupt_vietnamese_text:{guid}:{sense.get('semantic_sense_id')}:{location}"
                    )
            cambridge = sense.get("cambridge") or {}
            cambridge_match = cambridge.get("match")
            if cambridge_match not in CAMBRIDGE_MATCH_VALUES:
                errors.append(f"invalid_cambridge_match:{guid}:{sense.get('semantic_sense_id')}")
            if sense.get("approval", "") not in APPROVAL_VALUES:
                errors.append(f"invalid_approval:{guid}:{sense.get('semantic_sense_id')}")
            if require_complete and (decision in {"pending", "uncertain"} or sense.get("approval") == "rejected"):
                errors.append(f"open_review:{guid}:{sense.get('semantic_sense_id')}")
            if require_complete and decision == "repair_proposed" and sense.get("approval") != "approved":
                errors.append(f"unapproved_repair:{guid}:{sense.get('semantic_sense_id')}")
            if require_complete and any(value in {"pending", "uncertain"} for value in checks.values()):
                errors.append(f"open_check:{guid}:{sense.get('semantic_sense_id')}")
            if require_complete and not sense.get("reviewer"):
                errors.append(f"missing_reviewer:{guid}:{sense.get('semantic_sense_id')}")
            if require_complete and cambridge_match == "pending":
                errors.append(f"open_cambridge_match:{guid}:{sense.get('semantic_sense_id')}")
            if require_complete and not cambridge.get("translation_provenance"):
                errors.append(f"missing_translation_provenance:{guid}:{sense.get('semantic_sense_id')}")
        if semantic_senses:
            effective_examples = [
                example
                for sense in semantic_senses
                for example in (
                    (
                        sense.get("proposed")
                        if sense.get("decision") == "repair_proposed"
                        else sense.get("current")
                    )
                    or {}
                ).get("examples") or []
            ]
            shortfall = main_example_pos_shortfall(card.get("pos"), effective_examples)
            if shortfall is not None:
                actual, required = shortfall
                errors.append(f"main_example_pos_shortfall:{guid}:{actual}<{required}")
        unknown = set(mapped) - known
        if unknown:
            errors.append(f"unknown_source_sense_id:{guid}:{','.join(sorted(unknown))}")
        if require_complete and card.get("coverage", {}).get("status") in {"pending", "uncertain"}:
            errors.append(f"open_coverage:{guid}")
        semantic_ids = {sense.get("semantic_sense_id") for sense in card.get("semantic_senses") or []}
        candidates = set(card.get("coverage", {}).get("candidate_source_sense_ids") or [])
        coverage_items = card.get("source_coverage") or []
        coverage_ids = [item.get("source_sense_id") for item in coverage_items]
        if set(coverage_ids) != candidates or len(coverage_ids) != len(set(coverage_ids)):
            errors.append(f"invalid_source_coverage_set:{guid}")
        coverage_mapped: dict[str, list[str]] = {}
        source_by_id = {
            str(item.get("source_sense_id") or ""): item
            for item in card.get("source_senses") or []
        }
        duplicate_groups: dict[str, list[dict]] = {}
        for item in coverage_items:
            disposition = item.get("disposition")
            targets = item.get("target_semantic_sense_ids") or []
            if disposition not in COVERAGE_DISPOSITIONS:
                errors.append(f"invalid_source_disposition:{guid}:{disposition}")
            if set(targets) - semantic_ids:
                errors.append(f"unknown_target_semantic_sense:{guid}:{item.get('source_sense_id')}")
            if disposition == "mapped" and not targets:
                errors.append(f"mapped_source_without_target:{guid}:{item.get('source_sense_id')}")
            if disposition == "excluded" and not (item.get("reason") or "").strip():
                errors.append(f"excluded_source_without_reason:{guid}:{item.get('source_sense_id')}")
            if require_complete and disposition == "pending":
                errors.append(f"unaccounted_source_sense:{guid}:{item.get('source_sense_id')}")
            if disposition == "mapped":
                coverage_mapped[item.get("source_sense_id")] = targets
                source_id = str(item.get("source_sense_id") or "")
                previous_owner = mapped_source_owners.setdefault(source_id, guid)
                if previous_owner != guid:
                    errors.append(f"source_mapped_to_multiple_cards:{source_id}:{previous_owner}:{guid}")
            source = source_by_id.get(str(item.get("source_sense_id") or ""), {})
            normalized = _normalized_source_definition(source.get("definition"))
            if normalized:
                duplicate_groups.setdefault(normalized, []).append(item)
        for group in duplicate_groups.values():
            dispositions = {item.get("disposition") for item in group}
            unexplained = [
                item for item in group
                if item.get("disposition") == "excluded"
                and not EXPLICIT_DUPLICATE_EXCLUSION_RE.search(str(item.get("reason") or ""))
            ]
            if "mapped" in dispositions and unexplained:
                source_ids_text = ",".join(sorted(str(item.get("source_sense_id") or "") for item in group))
                errors.append(f"contradictory_duplicate_source_disposition:{guid}:{source_ids_text}")
        for sense in card.get("semantic_senses") or []:
            expected = sorted(source_id for source_id, targets in coverage_mapped.items() if sense.get("semantic_sense_id") in targets)
            if sorted(sense.get("source_sense_ids") or []) != expected:
                errors.append(f"source_mapping_mismatch:{guid}:{sense.get('semantic_sense_id')}")
        if not card.get("semantic_senses"):
            valid_idiom_only = (
                card.get("coverage", {}).get("status") == "not_applicable"
                and card.get("coverage", {}).get("reason") == "idiom_only"
                and not current.get("definition")
                and not current.get("example")
                and bool(current.get("idioms"))
            )
            if not valid_idiom_only:
                errors.append(f"invalid_empty_card:{guid}")
    if active_guids:
        for guid in sorted(active_guids - seen_guid):
            errors.append(f"missing_active_guid:{guid}")
        for guid in sorted(seen_guid - active_guids):
            errors.append(f"unknown_audit_guid:{guid}")
    return errors


def _style_sheet(sheet, widths: dict[str, int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for name, width in widths.items():
        sheet.column_dimensions[get_column_letter(REVIEW_COLUMNS.index(name) + 1)].width = width


def _append_literal_row(sheet, values: list[object]) -> None:
    sheet.append(values)
    row_number = sheet.max_row
    for column, value in enumerate(values, 1):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            sheet.cell(row_number, column).data_type = "s"


def export_workbook(audit_rows: list[dict], path: Path) -> None:
    flat = flatten_review_rows(audit_rows)
    coverage_flat = flatten_coverage_rows(audit_rows)
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    review = wb.create_sheet("Review")
    source_coverage = wb.create_sheet("Source Coverage")
    open_review = wb.create_sheet("Open Review")
    instructions = wb.create_sheet("Instructions")

    overview.append(["Metric", "Value"])
    overview.append(["Review rows", "=COUNTA(Review!A:A)-1"])
    decision_letter = get_column_letter(REVIEW_COLUMNS.index("decision") + 1)
    approval_letter = get_column_letter(REVIEW_COLUMNS.index("approval") + 1)
    overview.append(["Pending", f'=COUNTIF(Review!{decision_letter}:{decision_letter},"pending")'])
    overview.append(["Repair proposed", f'=COUNTIF(Review!{decision_letter}:{decision_letter},"repair_proposed")'])
    overview.append(["Uncertain", f'=COUNTIF(Review!{decision_letter}:{decision_letter},"uncertain")'])
    overview.append(["Approved", f'=COUNTIF(Review!{approval_letter}:{approval_letter},"approved")'])
    for row in overview.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.row == 1)
    overview.column_dimensions["A"].width = 24
    overview.column_dimensions["B"].width = 18

    review.append(list(REVIEW_COLUMNS))
    for item in flat:
        _append_literal_row(review, [item.get(column, "") for column in REVIEW_COLUMNS])
    _style_sheet(review, {
        "row_id": 28, "batch_id": 10, "guid": 14, "word": 20, "card_pos": 18,
        "candidate_source_definitions": 48, "candidate_source_examples": 48, "current_en": 34, "current_vi": 28,
        "current_examples": 48, "cambridge_url": 38, "cambridge_summary": 34,
        "proposed_en": 34, "proposed_vi": 28, "proposed_examples": 48, "review_reason": 42,
    })
    review.sheet_view.showGridLines = False
    for column in IMMUTABLE_COLUMNS:
        review.column_dimensions[get_column_letter(REVIEW_COLUMNS.index(column) + 1)].hidden = column in {"row_id", "semantic_sense_id", "row_fingerprint"}

    validations = {
        "cambridge_match": CAMBRIDGE_MATCH_VALUES,
        "english_check": CHECK_VALUES,
        "vietnamese_check": CHECK_VALUES,
        "simplicity_check": CHECK_VALUES,
        "example_pos_check": CHECK_VALUES,
        "decision": DECISION_VALUES,
        "approval": APPROVAL_VALUES[1:],
    }
    for column, values in validations.items():
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=column == "approval")
        review.add_data_validation(dv)
        letter = get_column_letter(REVIEW_COLUMNS.index(column) + 1)
        dv.add(f"{letter}2:{letter}{max(2, len(flat) + 1)}")
    review.conditional_formatting.add(
        f"A2:{get_column_letter(len(REVIEW_COLUMNS))}{max(2, len(flat) + 1)}",
        FormulaRule(formula=[f'${decision_letter}2="uncertain"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )

    source_coverage.append(list(COVERAGE_COLUMNS))
    for item in coverage_flat:
        _append_literal_row(source_coverage, [item.get(column, "") for column in COVERAGE_COLUMNS])
    source_coverage.freeze_panes = "A2"
    source_coverage.auto_filter.ref = source_coverage.dimensions
    source_coverage.sheet_view.showGridLines = False
    for cell in source_coverage[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in source_coverage.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    coverage_widths = {"word": 20, "source_definition": 52, "source_examples": 52, "reason": 42, "target_semantic_sense_ids": 42}
    for name, width in coverage_widths.items():
        source_coverage.column_dimensions[get_column_letter(COVERAGE_COLUMNS.index(name) + 1)].width = width
    for column in ("coverage_row_id", "row_fingerprint"):
        source_coverage.column_dimensions[get_column_letter(COVERAGE_COLUMNS.index(column) + 1)].hidden = True
    disposition_validation = DataValidation(type="list", formula1='"' + ",".join(COVERAGE_DISPOSITIONS) + '"')
    source_coverage.add_data_validation(disposition_validation)
    disposition_letter = get_column_letter(COVERAGE_COLUMNS.index("disposition") + 1)
    disposition_validation.add(f"{disposition_letter}2:{disposition_letter}{max(2, len(coverage_flat) + 1)}")

    open_review.append(list(REVIEW_COLUMNS))
    for item in flat:
        if item["decision"] not in {"pass", "not_applicable"} or (item["decision"] == "repair_proposed" and item["approval"] != "approved"):
            _append_literal_row(open_review, [item.get(column, "") for column in REVIEW_COLUMNS])
    _style_sheet(open_review, {"word": 20, "current_en": 34, "current_vi": 28, "review_reason": 42})
    open_review.protection.sheet = True

    instructions.append(["Bilingual Semantic Audit"])
    instructions.append(["Edit only the review fields in the Review sheet. JSONL remains canonical."])
    instructions.append(["Use friendly English without a hard word cap; preserve every distinct learner-relevant Oxford sense at the card CEFR."])
    instructions.append(["Write English and Vietnamese as independent Lexical Glosses: prefer a familiar concise dictionary equivalent over a clause-by-clause copy of the source definition."])
    instructions.append(["Source definitions are semantic evidence, not learner-facing display text. Do not keep source objects, examples, or enumerations when one established lexical term preserves the same sense."])
    instructions.append(["If a longer gloss is necessary, the review reason must name the exact condition, restriction, or contrast a shorter wording would lose; a generic claim that it 'preserves nuance' is not sufficient."])
    instructions.append(["Do not close a long-gloss finding by changing only punctuation or word order when a concise lexical equivalent exists."])
    instructions.append(["Domain labels are review signals, not automatic deletions. A niche sense may be removed only by an explicit review bundle that remaps or excludes every affected source."])
    instructions.append(["Vietnamese may use natural Hán-Việt or slang when it matches the sense."])
    instructions.column_dimensions["A"].width = 110
    for row in instructions.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.row == 1)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def import_workbook(audit_rows: list[dict], path: Path) -> list[dict]:
    expected = {row["row_id"]: row for row in flatten_review_rows(audit_rows)}
    wb = load_workbook(path, data_only=False)
    sheet = wb["Review"]
    headers = [cell.value for cell in sheet[1]]
    if headers != list(REVIEW_COLUMNS):
        raise ValueError("Review sheet columns do not match the semantic-audit contract")
    updates: dict[str, dict] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {header: (value if value is not None else "") for header, value in zip(headers, values)}
        row_id = str(row["row_id"])
        original = expected.get(row_id)
        if original is None:
            raise ValueError(f"Unknown review row: {row_id}")
        if row_id in updates:
            raise ValueError(f"Duplicate review row: {row_id}")
        if str(row["row_fingerprint"]) != original["row_fingerprint"]:
            raise ValueError(f"Stale or modified immutable review row: {row_id}")
        for column in IMMUTABLE_COLUMNS:
            if str(row[column]) != str(original[column]):
                raise ValueError(f"Immutable column {column!r} changed for {row_id}")
        updates[row_id] = row
    if set(updates) != set(expected):
        missing = sorted(set(expected) - set(updates))
        raise ValueError(f"Workbook is missing review rows: {missing[:5]}")

    expected_coverage = {row["coverage_row_id"]: row for row in flatten_coverage_rows(audit_rows)}
    coverage_sheet = wb["Source Coverage"]
    coverage_headers = [cell.value for cell in coverage_sheet[1]]
    if coverage_headers != list(COVERAGE_COLUMNS):
        raise ValueError("Source Coverage sheet columns do not match the semantic-audit contract")
    coverage_updates: dict[str, dict] = {}
    for values in coverage_sheet.iter_rows(min_row=2, values_only=True):
        row = {header: (value if value is not None else "") for header, value in zip(coverage_headers, values)}
        row_id = str(row["coverage_row_id"])
        original = expected_coverage.get(row_id)
        if original is None or row_id in coverage_updates:
            raise ValueError(f"Unknown or duplicate source-coverage row: {row_id}")
        if str(row["row_fingerprint"]) != original["row_fingerprint"]:
            raise ValueError(f"Stale source-coverage row: {row_id}")
        immutable_names = set(COVERAGE_COLUMNS) - {"disposition", "target_semantic_sense_ids", "reason", "row_fingerprint"}
        for column in immutable_names:
            if str(row[column]) != str(original[column]):
                raise ValueError(f"Immutable source-coverage column {column!r} changed for {row_id}")
        coverage_updates[row_id] = row
    if set(coverage_updates) != set(expected_coverage):
        missing = sorted(set(expected_coverage) - set(coverage_updates))
        raise ValueError(f"Workbook is missing source-coverage rows: {missing[:5]}")

    for card in audit_rows:
        semantic_by_id = {sense["semantic_sense_id"]: sense for sense in card.get("semantic_senses") or []}
        for sense in semantic_by_id.values():
            sense["source_sense_ids"] = []
        for coverage in card.get("source_coverage") or []:
            row_id = f"{card['guid']}::{coverage['source_sense_id']}"
            row = coverage_updates[row_id]
            targets = [part.strip() for part in str(row["target_semantic_sense_ids"]).split("|") if part.strip()]
            coverage.update({"disposition": str(row["disposition"]), "target_semantic_sense_ids": targets, "reason": str(row["reason"])})
            if coverage["disposition"] == "mapped":
                for target in targets:
                    if target in semantic_by_id:
                        semantic_by_id[target]["source_sense_ids"].append(coverage["source_sense_id"])
        for sense in card.get("semantic_senses") or []:
            row_id = f"{card['guid']}::{sense['semantic_sense_id']}"
            row = updates[row_id]
            sense["cambridge"].update({
                "match": str(row["cambridge_match"]), "summary": str(row["cambridge_summary"]),
                "translation_provenance": str(row["translation_provenance"]),
                "accessed_at": str(row["reviewed_at"] or ""),
            })
            sense["checks"] = {
                "english_semantics": str(row["english_check"]), "vietnamese_semantics": str(row["vietnamese_check"]),
                "simplicity": str(row["simplicity_check"]), "example_pos_alignment": str(row["example_pos_check"]),
            }
            sense["decision"] = str(row["decision"])
            sense["proposed"] = {
                "definition_en": str(row["proposed_en"]), "definition_vi": str(row["proposed_vi"]),
                "examples": [part.strip() for part in str(row["proposed_examples"]).split("||") if part.strip()],
            }
            sense["confidence"] = str(row["confidence"])
            sense["review_reason"] = str(row["review_reason"])
            sense["reviewer"] = str(row["reviewer"])
            sense["reviewed_at"] = str(row["reviewed_at"])
            sense["approval"] = str(row["approval"])
        decisions = {sense.get("decision") for sense in card.get("semantic_senses") or []}
        if decisions:
            if any(item.get("disposition") == "pending" for item in card.get("source_coverage") or []):
                card["coverage"]["status"] = "pending"
            elif "uncertain" in decisions:
                card["coverage"]["status"] = "uncertain"
            elif "pending" in decisions:
                card["coverage"]["status"] = "pending"
            elif "repair_proposed" in decisions:
                card["coverage"]["status"] = "repair_proposed"
            else:
                card["coverage"]["status"] = "pass"
    return audit_rows


def apply_review_bundle(audit_rows: list[dict], decisions: list[dict]) -> list[dict]:
    """Apply explicit card decisions; scaffold heuristics never call this.

    ``remove_senses`` is deliberately review-only: every source that targeted a
    removed sense must be explicitly remapped or excluded in the same bundle.
    """
    cards = {card["guid"]: card for card in audit_rows}
    seen: set[str] = set()
    for decision_card in decisions:
        guid = decision_card.get("guid") or ""
        if guid not in cards or guid in seen:
            raise ValueError(f"Unknown or duplicate review GUID: {guid}")
        seen.add(guid)
        card = cards[guid]
        current = card.get("current") or {}
        idiom_only_shape = (
            card.get("coverage", {}).get("reason") == "idiom_only"
            and not current.get("definition")
            and not current.get("example")
            and bool(current.get("idioms"))
        )
        if idiom_only_shape and decision_card.get("add_senses"):
            raise ValueError(f"Cannot add semantic senses to idiom-only card: {guid}")
        for update in decision_card.get("add_senses") or []:
            update = {"word": card.get("word") or "", **update}
            added = _added_semantic_sense(guid, update)
            existing_ids = {sense.get("semantic_sense_id") for sense in card.get("semantic_senses") or []}
            if added["semantic_sense_id"] in existing_ids:
                raise ValueError(f"Duplicate added semantic sense for {guid}: {added['semantic_sense_id']}")
            card.setdefault("semantic_senses", []).append(added)
        semantic_by_id = {sense["semantic_sense_id"]: sense for sense in card.get("semantic_senses") or []}
        coverage_by_id = {item["source_sense_id"]: item for item in card.get("source_coverage") or []}
        raw_remove_ids = decision_card.get("remove_senses", [])
        if not isinstance(raw_remove_ids, list) or any(
            not isinstance(semantic_id, str) or not semantic_id
            for semantic_id in raw_remove_ids
        ):
            raise ValueError(f"remove_senses must be a list of semantic IDs for {guid}")
        remove_ids = list(raw_remove_ids)
        if len(remove_ids) != len(set(remove_ids)):
            raise ValueError(f"Duplicate semantic sense removal for {guid}")
        unknown_remove_ids = set(remove_ids) - set(semantic_by_id)
        if unknown_remove_ids:
            raise ValueError(
                f"Unknown semantic sense for {guid}: {','.join(sorted(unknown_remove_ids))}"
            )
        if remove_ids and len(remove_ids) == len(semantic_by_id):
            raise ValueError(f"Cannot remove every semantic sense for {guid}")

        for update in decision_card.get("source_coverage") or []:
            source_id = update.get("source_sense_id") or ""
            if source_id not in coverage_by_id:
                raise ValueError(f"Unknown source sense for {guid}: {source_id}")
            targets = update.get("target_semantic_sense_ids") or []
            if set(targets) - set(semantic_by_id):
                raise ValueError(f"Unknown semantic target for {guid}/{source_id}")
            coverage_by_id[source_id].update({
                "disposition": update.get("disposition") or "pending",
                "target_semantic_sense_ids": targets,
                "reason": update.get("reason") or "",
            })

        removal_targets = [
            source_id
            for source_id, coverage in coverage_by_id.items()
            if set(coverage.get("target_semantic_sense_ids") or []) & set(remove_ids)
        ]
        if removal_targets:
            raise ValueError(
                "Source coverage still targets removed semantic sense for "
                f"{guid}: {','.join(sorted(removal_targets))}"
            )
        for semantic_id in remove_ids:
            del semantic_by_id[semantic_id]

        for update in decision_card.get("senses") or []:
            semantic_id = update.get("semantic_sense_id") or ""
            if semantic_id not in semantic_by_id:
                raise ValueError(f"Unknown semantic sense for {guid}: {semantic_id}")
            sense = semantic_by_id[semantic_id]
            if "order" in update:
                sense["order"] = int(update["order"])
            if "checks" in update:
                sense["checks"] = dict(update["checks"])
            sense["decision"] = update.get("decision", sense.get("decision", "pending"))
            if "proposed" in update:
                sense["proposed"] = {
                    "definition_en": update["proposed"].get("definition_en") or "",
                    "definition_vi": update["proposed"].get("definition_vi") or "",
                    "examples": update["proposed"].get("examples") or [],
                }
            if "cambridge" in update:
                sense["cambridge"].update(update["cambridge"])
            sense["confidence"] = update.get("confidence", sense.get("confidence", ""))
            sense["review_reason"] = update.get("review_reason", sense.get("review_reason", ""))
            sense["reviewer"] = update.get("reviewer") or REVIEWER
            sense["reviewed_at"] = update.get("reviewed_at") or date.today().isoformat()
            sense["approval"] = update.get("approval", sense.get("approval", ""))

        orders = [sense.get("order") for sense in semantic_by_id.values()]
        if any(not isinstance(order, int) or order < 1 for order in orders) or len(orders) != len(set(orders)):
            raise ValueError(f"Duplicate or invalid semantic sense order for {guid}")
        card["semantic_senses"] = sorted(semantic_by_id.values(), key=lambda sense: sense["order"])
        for order, sense in enumerate(card["semantic_senses"], 1):
            sense["order"] = order

        for sense in semantic_by_id.values():
            sense["source_sense_ids"] = sorted(
                source_id
                for source_id, coverage in coverage_by_id.items()
                if coverage.get("disposition") == "mapped"
                and sense["semantic_sense_id"] in (coverage.get("target_semantic_sense_ids") or [])
            )
        if any(item.get("disposition") == "pending" for item in coverage_by_id.values()):
            card["coverage"]["status"] = "pending"
        else:
            sense_decisions = {sense.get("decision") for sense in semantic_by_id.values()}
            if "uncertain" in sense_decisions:
                card["coverage"]["status"] = "uncertain"
            elif "pending" in sense_decisions:
                card["coverage"]["status"] = "pending"
            elif "repair_proposed" in sense_decisions:
                card["coverage"]["status"] = "repair_proposed"
            elif sense_decisions:
                card["coverage"]["status"] = "pass"
    return audit_rows


def audit_summary(rows: list[dict]) -> dict[str, int]:
    decisions = Counter(
        sense.get("decision", "pending")
        for card in rows
        for sense in card.get("semantic_senses") or []
    )
    return {"cards": len(rows), "senses": sum(decisions.values()), **dict(sorted(decisions.items()))}


def mark_reviewed_pass(sense: dict, *, reason: str) -> None:
    """Small helper for the ChatGPT reviewer; never called by scaffold heuristics."""
    sense["checks"] = {key: "pass" for key in sense["checks"]}
    sense["decision"] = "pass"
    sense["confidence"] = "high"
    sense["review_reason"] = reason
    sense["reviewer"] = REVIEWER
    sense["reviewed_at"] = date.today().isoformat()
