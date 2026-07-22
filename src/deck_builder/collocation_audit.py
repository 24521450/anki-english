"""Reviewed collocation evidence, audit, promotion, and registry contracts."""
from __future__ import annotations

import copy
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.deck_builder.simplify_senses import _flatten_senses
from src.deck_builder.source_sense_identity import source_sense_id


AUDIT_SCHEMA_VERSION = 2
REGISTRY_SCHEMA_VERSION = 2
MAX_FINAL_ITEMS = 5

SOURCE_ORDER = {"oxford": 0, "cambridge": 1}
FINAL_SOURCE_ORDER = {
    "oxford": 0,
    "oxford+cambridge": 0,
    "cambridge": 1,
    "curated": 2,
}
FINAL_SOURCES = tuple(FINAL_SOURCE_ORDER)
EVIDENCE_KINDS = {"example_linked", "supporting"}
EVIDENCE_ORIGINS = {
    "oxford_example_cf",
    "cambridge_example_lu",
    "oxford_collocations_snippet",
    "cambridge_bare_lu",
    "cambridge_grammar_cl",
}
ORIGIN_CONTRACT = {
    "oxford_example_cf": ("oxford", "example_linked"),
    "cambridge_example_lu": ("cambridge", "example_linked"),
    "oxford_collocations_snippet": ("oxford", "supporting"),
    "cambridge_bare_lu": ("cambridge", "supporting"),
    "cambridge_grammar_cl": ("cambridge", "supporting"),
}
CURRENT_DECISIONS = (
    "pending",
    "keep_source",
    "keep_curated",
    "rewrite_or_split",
    "remove",
    "uncertain",
)
CANDIDATE_DECISIONS = (
    "pending",
    "included",
    "covered",
    "excluded",
    "uncertain",
)
APPROVAL_VALUES = ("", "approved", "rejected")

ROW_FIELDS = {
    "schema_version",
    "guid",
    "word",
    "cefr",
    "list",
    "variant",
    "pos",
    "source_mappings",
    "source_evidence",
    "mandatory_candidates",
    "current_items",
    "final_items",
    "idiom_phrases",
    "idiom_fingerprint",
    "current_fingerprint",
    "source_fingerprint",
    "input_fingerprint",
    "empty_reason",
    "empty_reviewer",
    "empty_reviewed_at",
    "empty_approval",
}
SOURCE_MAPPING_FIELDS = {"semantic_sense_id", "order", "source_sense_ids"}
EVIDENCE_FIELDS = {
    "evidence_id",
    "source_sense_id",
    "source_headword",
    "semantic_sense_ids",
    "text",
    "source",
    "origin",
    "evidence_kind",
    "example_index",
    "example_text",
    "container_index",
    "item_index",
    "category",
    "truncated",
    "full_entry_url",
}
CURRENT_ITEM_FIELDS = {
    "current_item_id",
    "text",
    "order",
    "evidence_ids",
    "decision",
    "target_final_item_ids",
    "reason",
    "reviewer",
    "reviewed_at",
    "approval",
}
CANDIDATE_FIELDS = {
    "candidate_id",
    "text",
    "order",
    "sources",
    "evidence_ids",
    "decision",
    "target_final_item_ids",
    "reason",
    "reviewer",
    "reviewed_at",
    "approval",
}
FINAL_ITEM_FIELDS = {
    "final_item_id",
    "text",
    "order",
    "source",
    "evidence_ids",
    "current_item_ids",
}
REGISTRY_ROW_FIELDS = {
    "schema_version",
    "guid",
    "word",
    "cefr",
    "list",
    "variant",
    "audit_sha256",
    "audit_row_sha256",
    "idiom_fingerprint",
    "current_fingerprint",
    "source_fingerprint",
    "items",
    "empty_reason",
}
REGISTRY_ITEM_FIELDS = {"text", "order", "source", "evidence_ids"}

CARD_COLUMNS = (
    "schema_version",
    "guid",
    "word",
    "cefr",
    "list",
    "variant",
    "pos",
    "current_fingerprint",
    "source_fingerprint",
    "input_fingerprint",
    "idiom_phrases",
    "idiom_fingerprint",
    "empty_reason",
    "empty_reviewer",
    "empty_reviewed_at",
    "empty_approval",
)
CARD_IMMUTABLE_COLUMNS = CARD_COLUMNS[:12]
CARD_EDITABLE_COLUMNS = CARD_COLUMNS[12:]
CURRENT_REVIEW_COLUMNS = (
    "guid",
    "current_item_id",
    "text",
    "order",
    "evidence_ids",
    "input_fingerprint",
    "decision",
    "target_final_item_ids",
    "reason",
    "reviewer",
    "reviewed_at",
    "approval",
)
CURRENT_IMMUTABLE_COLUMNS = CURRENT_REVIEW_COLUMNS[:6]
CURRENT_EDITABLE_COLUMNS = CURRENT_REVIEW_COLUMNS[6:]
CANDIDATE_REVIEW_COLUMNS = (
    "guid",
    "candidate_id",
    "text",
    "order",
    "sources",
    "evidence_ids",
    "input_fingerprint",
    "decision",
    "target_final_item_ids",
    "reason",
    "reviewer",
    "reviewed_at",
    "approval",
)
CANDIDATE_IMMUTABLE_COLUMNS = CANDIDATE_REVIEW_COLUMNS[:7]
CANDIDATE_EDITABLE_COLUMNS = CANDIDATE_REVIEW_COLUMNS[7:]
FINAL_REVIEW_COLUMNS = (
    "guid",
    "final_item_id",
    "text",
    "order",
    "source",
    "evidence_ids",
    "current_item_ids",
)
EVIDENCE_COLUMNS = (
    "guid",
    "evidence_id",
    "source_sense_id",
    "source_headword",
    "semantic_sense_ids",
    "text",
    "source",
    "origin",
    "evidence_kind",
    "example_index",
    "example_text",
    "container_index",
    "item_index",
    "category",
    "truncated",
    "full_entry_url",
    "input_fingerprint",
)

_HTML_RE = re.compile(r"<\s*/?\s*[a-zA-Z][^>]*>")
_REASON_SPACE_RE = re.compile(r"\s+")


def _canonical_json(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _sha256(value: object) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _trim(value: object) -> str:
    return str(value or "").strip()


def _display_text(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", _trim(value))
    normalized = "".join(
        char for char in normalized if unicodedata.category(char) != "Cf"
    )
    return " ".join(normalized.split())


def normalize_collocation(value: object) -> str:
    """Normalize matching text without deleting punctuation or learning slots."""
    return _display_text(value).casefold()


def collocation_text_matches_evidence(
    final_text: object,
    evidence_text: object,
    *,
    headword: object,
) -> bool:
    """Allow only a reviewed singular/plural change of the card headword."""
    final = normalize_collocation(final_text)
    evidence = normalize_collocation(evidence_text)
    if final == evidence:
        return True

    word = normalize_collocation(headword)
    if not word or " " in word:
        return False
    if word.endswith("y") and len(word) > 1 and word[-2] not in "aeiou":
        plural = word[:-1] + "ies"
    elif word.endswith(("s", "x", "z", "ch", "sh")):
        plural = word + "es"
    else:
        plural = word + "s"

    pattern = re.compile(rf"(?<!\w)(?:{re.escape(word)}|{re.escape(plural)})(?!\w)")
    return pattern.sub(word, final) == pattern.sub(word, evidence)


def _normalize_reason(value: object) -> str:
    return _REASON_SPACE_RE.sub(" ", _trim(value).casefold())


def _optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = _trim(value)
    return text or None


def _canonical_sources(values: Iterable[str]) -> list[str]:
    unique = {str(value).casefold() for value in values if str(value).strip()}
    return sorted(unique, key=lambda source: (SOURCE_ORDER.get(source, 99), source))


def _active_registry_index(registry_rows: Sequence[dict]) -> dict[str, dict]:
    indexed: dict[str, dict] = {}
    for row in registry_rows:
        if not isinstance(row, dict):
            raise ValueError("Card Registry rows must be objects")
        if row.get("status") != "active":
            continue
        guid = _trim(row.get("guid"))
        if not guid or guid in indexed:
            raise ValueError(f"duplicate or empty active Card Registry GUID: {guid}")
        indexed[guid] = row
    return indexed


def load_jsonl(path: Path) -> list[dict]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def serialize_audit_rows(rows: Iterable[dict]) -> str:
    return "".join(
        _canonical_json(row) + "\n"
        for row in sorted(rows, key=lambda row: _trim(row.get("guid")))
    )


def serialize_registry_rows(rows: Iterable[dict]) -> str:
    return "".join(
        _canonical_json(row) + "\n"
        for row in sorted(rows, key=lambda row: _trim(row.get("guid")))
    )


def parse_serialized_collocations(value: object) -> list[str]:
    raw = _trim(value)
    if not raw:
        return []
    items = [item.strip() for item in raw.split("|")]
    if any(not item for item in items):
        raise ValueError("Collocations contains an empty pipe-delimited item")
    return items


def parse_serialized_idiom_phrases(value: object) -> list[str]:
    """Return the normalized displayed phrase from each ``$$`` Idiom cell."""
    raw = _trim(value)
    if not raw:
        return []
    phrases: list[str] = []
    for cell in raw.split("$$"):
        phrase = normalize_collocation(cell.split("::", 1)[0])
        if not phrase:
            raise ValueError("Idioms contains an empty displayed phrase")
        phrases.append(phrase)
    if len(phrases) != len(set(phrases)):
        raise ValueError("Idioms contains a duplicate displayed phrase")
    return phrases


def collocation_current_item_id(guid: str, order: int, text: str) -> str:
    return "cur_" + _sha256({"guid": guid, "order": order, "text": text})[:24]


def collocation_candidate_id(guid: str, text: str) -> str:
    return "cand_" + _sha256({"guid": guid, "text": normalize_collocation(text)})[:24]


def collocation_final_item_id(guid: str, text: str) -> str:
    return "col_" + _sha256({"guid": guid, "text": normalize_collocation(text)})[:24]


def _evidence_id(source_sense: str, evidence: Mapping[str, object]) -> str:
    return "ev_" + _sha256({"source_sense_id": source_sense, **dict(evidence)})[:24]


def _validate_index(value: object, label: str) -> int | None:
    if value is None:
        return None
    if type(value) is not int or value < 1:
        raise ValueError(f"{label} must be a positive integer or null")
    return value


def _normalize_source_evidence(
    raw: object,
    *,
    record_source: str,
    source_headword: object,
    source_sense: str,
    semantic_ids: list[str],
    definition_examples: Sequence[object] | None = None,
) -> dict:
    if not isinstance(raw, dict):
        raise ValueError(f"Collocation evidence must be an object: {source_sense}")
    required = {
        "text",
        "source",
        "origin",
        "evidence_kind",
        "example_index",
        "example_text",
        "container_index",
        "item_index",
        "category",
        "truncated",
        "full_entry_url",
    }
    missing = required - set(raw)
    if missing:
        raise ValueError(
            f"Collocation evidence is missing fields for {source_sense}: {sorted(missing)}"
        )
    text = _trim(raw.get("text"))
    source = _trim(raw.get("source")).casefold()
    headword = _display_text(source_headword)
    origin = _trim(raw.get("origin"))
    evidence_kind = _trim(raw.get("evidence_kind"))
    if not text:
        raise ValueError(f"Collocation evidence text is empty: {source_sense}")
    if not headword:
        raise ValueError(f"Collocation evidence source headword is empty: {source_sense}")
    if source != record_source or source not in SOURCE_ORDER:
        raise ValueError(f"Collocation evidence source mismatch: {source_sense}:{source}")
    if origin not in EVIDENCE_ORIGINS:
        raise ValueError(f"Unknown collocation evidence origin: {source_sense}:{origin}")
    if evidence_kind not in EVIDENCE_KINDS:
        raise ValueError(
            f"Unknown collocation evidence kind: {source_sense}:{evidence_kind}"
        )
    if ORIGIN_CONTRACT.get(origin) != (source, evidence_kind):
        raise ValueError(
            f"Collocation evidence origin contract mismatch: {source_sense}:{origin}"
        )
    if type(raw.get("truncated")) is not bool:
        raise ValueError(f"Collocation evidence truncated must be bool: {source_sense}")
    normalized = {
        "source_headword": headword,
        "text": text,
        "source": source,
        "origin": origin,
        "evidence_kind": evidence_kind,
        "example_index": _validate_index(raw.get("example_index"), "example_index"),
        "example_text": _optional_text(raw.get("example_text")),
        "container_index": _validate_index(raw.get("container_index"), "container_index"),
        "item_index": _validate_index(raw.get("item_index"), "item_index"),
        "category": _optional_text(raw.get("category")),
        "truncated": raw["truncated"],
        "full_entry_url": _optional_text(raw.get("full_entry_url")),
    }
    if evidence_kind == "example_linked" and (
        normalized["example_index"] is None or not normalized["example_text"]
    ):
        raise ValueError(
            f"Example-linked collocation evidence requires an example pair: {source_sense}"
        )
    if evidence_kind == "example_linked" and definition_examples is not None:
        example_index = normalized["example_index"]
        if example_index is None or example_index > len(definition_examples):
            raise ValueError(
                f"Collocation evidence example index is out of range: {source_sense}"
            )
        raw_example = definition_examples[example_index - 1]
        expected_text = _trim(
            raw_example.get("text") if isinstance(raw_example, dict) else raw_example
        )
        if normalized["example_text"] != (expected_text or None):
            raise ValueError(
                f"Collocation evidence example text mismatch: {source_sense}"
            )
    return {
        "evidence_id": _evidence_id(source_sense, normalized),
        "source_sense_id": source_sense,
        "semantic_sense_ids": list(semantic_ids),
        **normalized,
    }


def _source_sense_index(records: Sequence[dict]) -> dict[str, dict]:
    indexed: dict[str, dict] = {}
    for record in records:
        source = _trim(record.get("source")).casefold()
        if source not in SOURCE_ORDER:
            raise ValueError(f"Unsupported dictionary source: {source}")
        flattened = _flatten_senses(record)
        if not flattened:
            continue
        source_headword = _display_text(record.get("word"))
        if not source_headword:
            raise ValueError(f"Dictionary source record has no headword: {source}")
        source_files = tuple(str(item) for item in record.get("source_files") or [])
        for flat in flattened:
            source_id = source_sense_id(record, flat)
            if source_id in indexed:
                raise ValueError(f"Duplicate source sense ID: {source_id}")
            definition = record["pos_data"][flat.pd_idx]["definitions"][flat.def_idx]
            evidence = definition.get("collocation_evidence") or []
            if not isinstance(evidence, list):
                raise ValueError(f"collocation_evidence must be a list: {source_id}")
            indexed[source_id] = {
                "source": source,
                "source_headword": source_headword,
                "definition": definition,
                "raw_evidence": evidence,
                "sort_key": (
                    SOURCE_ORDER[source],
                    _trim(record.get("word")).casefold(),
                    source_files,
                    flat.pd_idx,
                    flat.def_idx,
                    source_id,
                ),
            }
    return indexed


def _source_mappings(semantic_row: dict, source_index: Mapping[str, dict]) -> list[dict]:
    mappings: list[dict] = []
    for sense in sorted(
        semantic_row.get("senses") or [],
        key=lambda item: (item.get("order", 10**9), _trim(item.get("semantic_sense_id"))),
    ):
        semantic_id = _trim(sense.get("semantic_sense_id"))
        order = sense.get("order")
        source_ids = list(sense.get("source_sense_ids") or [])
        if not semantic_id or type(order) is not int or order < 1:
            raise ValueError(
                f"Invalid Semantic Registry sense: {semantic_row.get('guid')}:{semantic_id}"
            )
        unknown = [source_id for source_id in source_ids if source_id not in source_index]
        if unknown:
            raise ValueError(
                f"Semantic Registry references missing source senses for {semantic_row.get('guid')}: {unknown[:5]}"
            )
        source_ids = sorted(
            set(source_ids), key=lambda source_id: source_index[source_id]["sort_key"]
        )
        mappings.append({
            "semantic_sense_id": semantic_id,
            "order": order,
            "source_sense_ids": source_ids,
        })
    return mappings


def _build_evidence(
    mappings: Sequence[dict], source_index: Mapping[str, dict]
) -> list[dict]:
    semantic_by_source: dict[str, list[tuple[int, str]]] = defaultdict(list)
    for mapping in mappings:
        for source_id in mapping["source_sense_ids"]:
            semantic_by_source[source_id].append(
                (mapping["order"], mapping["semantic_sense_id"])
            )
    source_ids = sorted(
        semantic_by_source,
        key=lambda source_id: (
            SOURCE_ORDER[source_index[source_id]["source"]],
            min(order for order, _ in semantic_by_source[source_id]),
            source_index[source_id]["sort_key"],
        ),
    )
    evidence: list[dict] = []
    seen_ids: set[str] = set()
    for source_id in source_ids:
        semantic_ids = [
            semantic_id
            for _, semantic_id in sorted(set(semantic_by_source[source_id]))
        ]
        meta = source_index[source_id]
        for raw in meta["raw_evidence"]:
            item = _normalize_source_evidence(
                raw,
                record_source=meta["source"],
                source_headword=meta["source_headword"],
                source_sense=source_id,
                semantic_ids=semantic_ids,
                definition_examples=meta["definition"].get("examples") or [],
            )
            if item["evidence_id"] in seen_ids:
                continue
            seen_ids.add(item["evidence_id"])
            evidence.append(item)
    return evidence


def _candidate_immutable(item: dict) -> dict:
    return {
        "candidate_id": item["candidate_id"],
        "text": item["text"],
        "order": item["order"],
        "sources": item["sources"],
        "evidence_ids": item["evidence_ids"],
    }


def _build_candidates(guid: str, evidence: Sequence[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    for item in evidence:
        if item.get("evidence_kind") != "example_linked":
            continue
        normalized = normalize_collocation(item.get("text"))
        if not normalized:
            continue
        if normalized not in grouped:
            text = _display_text(item.get("text"))
            grouped[normalized] = {
                "candidate_id": collocation_candidate_id(guid, text),
                "text": text,
                "order": len(grouped) + 1,
                "sources": [],
                "evidence_ids": [],
                "decision": "pending",
                "target_final_item_ids": [],
                "reason": "",
                "reviewer": "",
                "reviewed_at": "",
                "approval": "",
            }
        candidate = grouped[normalized]
        if item["source"] not in candidate["sources"]:
            candidate["sources"].append(item["source"])
            candidate["sources"] = _canonical_sources(candidate["sources"])
        candidate["evidence_ids"].append(item["evidence_id"])
    return list(grouped.values())


def _current_immutable(item: dict) -> dict:
    return {
        "current_item_id": item["current_item_id"],
        "text": item["text"],
        "order": item["order"],
        "evidence_ids": item["evidence_ids"],
    }


def _build_current_items(
    guid: str, serialized: object, evidence: Sequence[dict]
) -> list[dict]:
    evidence_by_text: dict[str, list[str]] = defaultdict(list)
    for item in evidence:
        evidence_by_text[normalize_collocation(item.get("text"))].append(
            item["evidence_id"]
        )
    current: list[dict] = []
    for order, text in enumerate(parse_serialized_collocations(serialized), 1):
        current.append({
            "current_item_id": collocation_current_item_id(guid, order, text),
            "text": text,
            "order": order,
            "evidence_ids": list(evidence_by_text.get(normalize_collocation(text), [])),
            "decision": "pending",
            "target_final_item_ids": [],
            "reason": "",
            "reviewer": "",
            "reviewed_at": "",
            "approval": "",
        })
    return current


def _current_fingerprint(current: Sequence[dict]) -> str:
    return _sha256([_current_immutable(item) for item in current])


def _idiom_fingerprint(idiom_phrases: Sequence[str]) -> str:
    return _sha256(list(idiom_phrases))


def _source_fingerprint(
    mappings: Sequence[dict], evidence: Sequence[dict], candidates: Sequence[dict]
) -> str:
    return _sha256({
        "source_mappings": list(mappings),
        "source_evidence": list(evidence),
        "mandatory_candidates": [_candidate_immutable(item) for item in candidates],
    })


def _input_fingerprint(row: Mapping[str, object]) -> str:
    return _sha256({
        "identity": {
            field: row.get(field) or ""
            for field in ("guid", "word", "cefr", "list", "variant", "pos")
        },
        "idiom_fingerprint": row.get("idiom_fingerprint"),
        "current_fingerprint": row.get("current_fingerprint"),
        "source_fingerprint": row.get("source_fingerprint"),
    })


def _reuse_review_state(fresh: dict, existing: dict) -> None:
    identity_fields = ("guid", "word", "cefr", "list", "variant", "pos")
    if any(_trim(fresh.get(field)) != _trim(existing.get(field)) for field in identity_fields):
        return
    if fresh["current_fingerprint"] != existing.get("current_fingerprint"):
        return
    if fresh["idiom_fingerprint"] != existing.get("idiom_fingerprint"):
        return
    if fresh["source_fingerprint"] != existing.get("source_fingerprint"):
        return

    old_current = {
        item.get("current_item_id"): item for item in existing.get("current_items") or []
    }
    old_candidates = {
        item.get("candidate_id"): item
        for item in existing.get("mandatory_candidates") or []
    }
    if set(old_current) != {item["current_item_id"] for item in fresh["current_items"]}:
        return
    if set(old_candidates) != {
        item["candidate_id"] for item in fresh["mandatory_candidates"]
    }:
        return
    for item in fresh["current_items"]:
        old = old_current[item["current_item_id"]]
        for field in CURRENT_EDITABLE_COLUMNS:
            item[field] = copy.deepcopy(old.get(field, [] if field == "target_final_item_ids" else ""))
    for item in fresh["mandatory_candidates"]:
        old = old_candidates[item["candidate_id"]]
        for field in CANDIDATE_EDITABLE_COLUMNS:
            item[field] = copy.deepcopy(old.get(field, [] if field == "target_final_item_ids" else ""))
    fresh["final_items"] = copy.deepcopy(existing.get("final_items") or [])
    for field in CARD_EDITABLE_COLUMNS:
        fresh[field] = _trim(existing.get(field))


def build_audit_rows(
    cards: Sequence[dict],
    registry_rows: Sequence[dict],
    semantic_registry_rows: Sequence[dict],
    oxford_records: Sequence[dict],
    cambridge_records: Sequence[dict],
    *,
    existing_rows: Sequence[dict] | None = None,
) -> list[dict]:
    """Scaffold one pending, fingerprint-bound Collocation Audit row per active GUID."""
    active = _active_registry_index(registry_rows)
    cards_by_guid: dict[str, dict] = {}
    for card in cards:
        guid = _trim(card.get("guid") or card.get("GUID"))
        if not guid or guid in cards_by_guid:
            raise ValueError(f"duplicate or empty built-card GUID: {guid}")
        cards_by_guid[guid] = card
    semantic_by_guid: dict[str, dict] = {}
    for row in semantic_registry_rows:
        guid = _trim(row.get("guid"))
        if not guid or guid in semantic_by_guid:
            raise ValueError(f"duplicate or empty Semantic Registry GUID: {guid}")
        semantic_by_guid[guid] = row
    for label, actual in (("built notes", set(cards_by_guid)), ("Semantic Registry", set(semantic_by_guid))):
        if actual != set(active):
            missing = sorted(set(active) - actual)
            extra = sorted(actual - set(active))
            raise ValueError(f"{label} GUID coverage mismatch: missing={missing[:5]} extra={extra[:5]}")

    source_index = _source_sense_index([*oxford_records, *cambridge_records])
    old_by_guid = {
        _trim(row.get("guid")): row
        for row in existing_rows or []
        if isinstance(row, dict) and _trim(row.get("guid"))
    }
    rows: list[dict] = []
    source_owner: dict[str, str] = {}
    for guid in sorted(active):
        registry = active[guid]
        card = cards_by_guid[guid]
        semantic = semantic_by_guid[guid]
        for field in ("word", "cefr", "list", "variant", "pos"):
            if _trim(semantic.get(field)) != _trim(registry.get(field)):
                raise ValueError(f"Semantic Registry identity mismatch: {guid}:{field}")
        mappings = _source_mappings(semantic, source_index)
        for mapping in mappings:
            for source_id in mapping["source_sense_ids"]:
                previous = source_owner.setdefault(source_id, guid)
                if previous != guid:
                    raise ValueError(
                        f"Source sense mapped to multiple cards: {source_id}:{previous}:{guid}"
                    )
        evidence = _build_evidence(mappings, source_index)
        candidates = _build_candidates(guid, evidence)
        current_value = card.get("collocations")
        if current_value is None:
            current_value = card.get("Collocations", "")
        current = _build_current_items(guid, current_value, evidence)
        idiom_value = card.get("idioms")
        if idiom_value is None:
            idiom_value = card.get("Idioms", "")
        idiom_phrases = parse_serialized_idiom_phrases(idiom_value)
        row = {
            "schema_version": AUDIT_SCHEMA_VERSION,
            "guid": guid,
            "word": _trim(registry.get("word")),
            "cefr": _trim(registry.get("cefr")),
            "list": _trim(registry.get("list")),
            "variant": _trim(registry.get("variant")),
            "pos": _trim(registry.get("pos")),
            "source_mappings": mappings,
            "source_evidence": evidence,
            "mandatory_candidates": candidates,
            "current_items": current,
            "final_items": [],
            "idiom_phrases": idiom_phrases,
            "idiom_fingerprint": _idiom_fingerprint(idiom_phrases),
            "current_fingerprint": _current_fingerprint(current),
            "source_fingerprint": _source_fingerprint(mappings, evidence, candidates),
            "input_fingerprint": "",
            "empty_reason": "",
            "empty_reviewer": "",
            "empty_reviewed_at": "",
            "empty_approval": "",
        }
        row["input_fingerprint"] = _input_fingerprint(row)
        existing = old_by_guid.get(guid)
        if existing is not None:
            _reuse_review_state(row, existing)
        rows.append(row)
    return rows


def refresh_audit_rows(
    audit_rows: Sequence[dict],
    cards: Sequence[dict],
    registry_rows: Sequence[dict],
    semantic_registry_rows: Sequence[dict],
    oxford_records: Sequence[dict],
    cambridge_records: Sequence[dict],
) -> list[dict]:
    """Rebuild immutable inputs while retaining only still-bound review state."""
    return build_audit_rows(
        cards,
        registry_rows,
        semantic_registry_rows,
        oxford_records,
        cambridge_records,
        existing_rows=audit_rows,
    )


def _card_field(card: Mapping[str, object], field: str, legacy_field: str) -> str:
    value = card.get(field)
    if value is None:
        value = card.get(legacy_field, "")
    return str(value or "")


def _matches_promoted_projection(
    audit_rows: Sequence[dict],
    refreshed_rows: Sequence[dict],
    cards: Sequence[dict],
) -> bool:
    """Return whether live notes are the exact projection of reviewed finals.

    A completed audit is initially fingerprinted against the legacy/current
    chips.  After promotion, the build quite correctly replaces those chips
    with ``final_items``; treating that expected projection as a new audit
    baseline would make every legitimate include/remove/rewrite stale.  This
    helper accepts that one post-promotion state while still requiring the
    live source/idiom inputs to match the reviewed row.
    """
    audit_by_guid = {str(row.get("guid") or ""): row for row in audit_rows}
    fresh_by_guid = {str(row.get("guid") or ""): row for row in refreshed_rows}
    cards_by_guid = {
        str(card.get("guid") or card.get("GUID") or ""): card
        for card in cards
    }
    if set(audit_by_guid) != set(fresh_by_guid) or set(audit_by_guid) != set(cards_by_guid):
        return False
    identity_fields = ("guid", "word", "cefr", "list", "variant", "pos")
    for guid, row in audit_by_guid.items():
        card = cards_by_guid[guid]
        fresh = fresh_by_guid[guid]
        finals = row.get("final_items") or []
        expected_collocations = "|".join(
            str(item.get("text") or "") for item in finals
        )
        expected_sources = "|".join(
            str(item.get("source") or "") for item in finals
        )
        if (
            _card_field(card, "collocations", "Collocations") != expected_collocations
            or _card_field(card, "collocation_sources", "CollocationSources")
            != expected_sources
        ):
            return False
        for field in identity_fields:
            if _trim(fresh.get(field)) != _trim(row.get(field)):
                return False
        if (
            _trim(fresh.get("idiom_fingerprint"))
            != _trim(row.get("idiom_fingerprint"))
            or _trim(fresh.get("source_fingerprint"))
            != _trim(row.get("source_fingerprint"))
        ):
            return False
    return True


def validate_current_audit(
    audit_rows: Sequence[dict],
    cards: Sequence[dict],
    registry_rows: Sequence[dict],
    semantic_registry_rows: Sequence[dict],
    oxford_records: Sequence[dict],
    cambridge_records: Sequence[dict],
    *,
    require_complete: bool = False,
) -> list[str]:
    """Validate the ledger against a fresh projection of every live input."""
    errors = validate_audit_rows(
        audit_rows,
        registry_rows,
        require_complete=require_complete,
    )
    if errors:
        return errors
    try:
        refreshed = refresh_audit_rows(
            audit_rows,
            cards,
            registry_rows,
            semantic_registry_rows,
            oxford_records,
            cambridge_records,
        )
    except (IndexError, KeyError, TypeError, ValueError) as exc:
        return [f"collocation_audit_refresh_failed:{exc}"]
    refreshed_errors = validate_audit_rows(
        refreshed,
        registry_rows,
        # The refreshed rows are only an immutable-input probe here.  Their
        # current items are pending when live notes are the promoted final
        # projection, so completeness belongs to the canonical audit above.
        require_complete=False,
    )
    if refreshed_errors:
        errors.extend(f"refreshed:{error}" for error in refreshed_errors)
    elif serialize_audit_rows(refreshed) != serialize_audit_rows(audit_rows):
        if not _matches_promoted_projection(audit_rows, refreshed, cards):
            errors.append("stale_collocation_audit_projection")
    return errors


def _is_valid_string_list(value: object) -> bool:
    return (
        isinstance(value, list)
        and all(isinstance(item, str) and item for item in value)
        and len(value) == len(set(value))
    )


def _has_control(value: str) -> bool:
    return any(unicodedata.category(char) in {"Cc", "Cf"} for char in value)


def _final_text_errors(guid: str, item: dict) -> list[str]:
    item_id = _trim(item.get("final_item_id"))
    text = item.get("text")
    errors: list[str] = []
    if not isinstance(text, str) or not text.strip() or text != text.strip():
        errors.append(f"invalid_final_text:{guid}:{item_id}")
        return errors
    if (
        _has_control(text)
        or _HTML_RE.search(text)
        or "|" in text
        or ";" in text
        or "::" in text
        or "$$" in text
    ):
        errors.append(f"forbidden_final_text:{guid}:{item_id}")
    if item.get("source") != "curated" and "/" in text:
        errors.append(f"source_item_contains_slash:{guid}:{item_id}")
    return errors


def _validate_review_date(value: object, prefix: str, errors: list[str]) -> None:
    reviewed_at = _trim(value)
    if not reviewed_at:
        return
    try:
        date.fromisoformat(reviewed_at)
    except ValueError:
        errors.append(f"invalid_reviewed_at:{prefix}")


def _validate_resolved_review(
    item: dict,
    prefix: str,
    *,
    require_complete: bool,
    resolved: set[str],
    decision_field: str = "decision",
) -> list[str]:
    errors: list[str] = []
    decision = item.get(decision_field)
    approval = item.get("approval")
    _validate_review_date(item.get("reviewed_at"), prefix, errors)
    if approval not in APPROVAL_VALUES:
        errors.append(f"invalid_approval:{prefix}:{approval}")
    if require_complete:
        if decision not in resolved:
            errors.append(f"unresolved_{prefix}:{decision}")
        if decision in resolved:
            if approval != "approved":
                errors.append(f"unapproved_{prefix}:{approval}")
            if not _trim(item.get("reviewer")):
                errors.append(f"missing_reviewer:{prefix}")
            if not _trim(item.get("reviewed_at")):
                errors.append(f"missing_reviewed_at:{prefix}")
    return errors


def _final_source_set(token: str) -> set[str]:
    if token == "oxford+cambridge":
        return {"oxford", "cambridge"}
    if token in {"oxford", "cambridge"}:
        return {token}
    return set()


def _validate_final_items(row: dict, *, require_complete: bool) -> list[str]:
    guid = _trim(row.get("guid"))
    headword = row.get("word")
    errors: list[str] = []
    evidence = row.get("source_evidence") or []
    evidence_by_id = {item.get("evidence_id"): item for item in evidence if isinstance(item, dict)}
    current = row.get("current_items") or []
    current_by_id = {item.get("current_item_id"): item for item in current if isinstance(item, dict)}
    finals = row.get("final_items")
    if not isinstance(finals, list):
        return [f"invalid_final_items:{guid}"]
    if len(finals) > MAX_FINAL_ITEMS:
        errors.append(f"too_many_final_items:{guid}:{len(finals)}")
    final_by_id: dict[str, dict] = {}
    normalized_seen: set[str] = set()
    orders: list[int] = []
    order_keys: list[tuple[int, int]] = []
    evidence_position = {
        item.get("evidence_id"): index for index, item in enumerate(evidence)
        if isinstance(item, dict)
    }
    for item in finals:
        if not isinstance(item, dict):
            errors.append(f"invalid_final_item_type:{guid}")
            continue
        item_id = _trim(item.get("final_item_id"))
        if set(item) != FINAL_ITEM_FIELDS:
            errors.append(f"invalid_final_item_fields:{guid}:{item_id}")
        text = _trim(item.get("text"))
        order = item.get("order")
        source = item.get("source")
        if item_id != collocation_final_item_id(guid, text) or item_id in final_by_id:
            errors.append(f"invalid_or_duplicate_final_item_id:{guid}:{item_id}")
        final_by_id[item_id] = item
        if type(order) is not int or order < 1:
            errors.append(f"invalid_final_order:{guid}:{item_id}")
        else:
            orders.append(order)
        normalized = normalize_collocation(text)
        if not normalized or normalized in normalized_seen:
            errors.append(f"normalized_duplicate_final_item:{guid}:{item_id}")
        normalized_seen.add(normalized)
        if source not in FINAL_SOURCES:
            errors.append(f"invalid_final_source:{guid}:{item_id}:{source}")
        errors.extend(_final_text_errors(guid, item))
        evidence_ids = item.get("evidence_ids")
        current_ids = item.get("current_item_ids")
        if not isinstance(evidence_ids, list) or any(
            not isinstance(value, str) or not value for value in evidence_ids
        ) or len(evidence_ids or []) != len(set(evidence_ids or [])):
            errors.append(f"invalid_final_evidence_ids:{guid}:{item_id}")
            evidence_ids = []
        if not isinstance(current_ids, list) or any(
            not isinstance(value, str) or not value for value in current_ids
        ) or len(current_ids or []) != len(set(current_ids or [])):
            errors.append(f"invalid_final_current_item_ids:{guid}:{item_id}")
            current_ids = []
        if set(evidence_ids) - set(evidence_by_id):
            errors.append(f"unknown_final_evidence:{guid}:{item_id}")
        if set(current_ids) - set(current_by_id):
            errors.append(f"unknown_final_current_item:{guid}:{item_id}")
        allowed_sources = _final_source_set(source)
        source_evidence = [
            evidence_item
            for evidence_item in evidence
            if isinstance(evidence_item, dict)
            and evidence_item.get("source") in allowed_sources
        ]
        # Prefer exact surface evidence whenever it exists.  The reviewed
        # singular/plural fallback is only for a surface that has no exact
        # dictionary row (for example ``generous portion`` backed by
        # ``generous portions``); otherwise two separately reviewed chips
        # such as ``loyalty to``/``loyalties to`` would claim the same
        # evidence through morphology.
        exact_surface = normalize_collocation(text)
        exact_surface_evidence = [
            evidence_item
            for evidence_item in source_evidence
            if normalize_collocation(evidence_item.get("text")) == exact_surface
        ]
        matching_evidence = exact_surface_evidence or [
            evidence_item
            for evidence_item in source_evidence
            if collocation_text_matches_evidence(
                text,
                evidence_item.get("text"),
                headword=headword,
            )
        ]
        exact_evidence = [
            evidence_item["evidence_id"] for evidence_item in matching_evidence
        ]
        if source == "curated":
            if evidence_ids:
                errors.append(f"curated_item_claims_evidence:{guid}:{item_id}")
            current_positions = [
                current_by_id[current_id].get("order", 10**9)
                for current_id in current_ids
                if current_id in current_by_id
            ]
            order_keys.append((FINAL_SOURCE_ORDER[source], min(current_positions, default=10**9)))
        elif source in FINAL_SOURCE_ORDER:
            if not evidence_ids:
                errors.append(f"source_item_without_evidence:{guid}:{item_id}")
            if set(evidence_ids) != set(exact_evidence):
                errors.append(f"source_evidence_text_mismatch:{guid}:{item_id}")
            actual_sources = {
                evidence_by_id[evidence_id]["source"]
                for evidence_id in evidence_ids
                if evidence_id in evidence_by_id
            }
            if actual_sources != _final_source_set(source):
                errors.append(f"source_claim_mismatch:{guid}:{item_id}:{source}")
            order_keys.append((
                FINAL_SOURCE_ORDER[source],
                min((evidence_position.get(value, 10**9) for value in evidence_ids), default=10**9),
            ))
    expected_orders = list(range(1, len(finals) + 1))
    if orders != expected_orders:
        errors.append(f"non_contiguous_final_order:{guid}")
    if order_keys != sorted(order_keys):
        errors.append(f"invalid_final_source_order:{guid}")

    expected_current_by_final: dict[str, set[str]] = defaultdict(set)
    for item in current:
        if not isinstance(item, dict):
            continue
        if item.get("decision") in {"keep_source", "keep_curated", "rewrite_or_split"}:
            for final_id in item.get("target_final_item_ids") or []:
                expected_current_by_final[final_id].add(item.get("current_item_id"))
    if require_complete:
        candidate_targets = {
            final_id
            for candidate in row.get("mandatory_candidates") or []
            if isinstance(candidate, dict)
            for final_id in candidate.get("target_final_item_ids") or []
            if candidate.get("decision") in {"included", "covered"}
        }
        for final_id, item in final_by_id.items():
            if set(item.get("current_item_ids") or []) != expected_current_by_final.get(final_id, set()):
                errors.append(f"final_current_reverse_mapping_mismatch:{guid}:{final_id}")
            if final_id not in expected_current_by_final and final_id not in candidate_targets:
                errors.append(f"orphan_final_item:{guid}:{final_id}")
    return errors


def validate_audit_rows(
    rows: Sequence[dict],
    registry_rows: Sequence[dict] | None = None,
    *,
    require_complete: bool = False,
) -> list[str]:
    """Validate exact card/evidence coverage and explicit review completeness."""
    errors: list[str] = []
    active: dict[str, dict] = {}
    if registry_rows is not None:
        try:
            active = _active_registry_index(registry_rows)
        except ValueError as exc:
            errors.append(f"invalid_card_registry:{exc}")
    row_guids = [
        _trim(row.get("guid")) for row in rows if isinstance(row, dict)
    ]
    if row_guids != sorted(row_guids):
        errors.append("non_deterministic_audit_row_order")
    seen_guids: set[str] = set()
    reason_owners: dict[str, str] = {}
    for row in rows:
        if not isinstance(row, dict):
            errors.append("invalid_audit_row_type")
            continue
        guid = _trim(row.get("guid"))
        if set(row) != ROW_FIELDS:
            errors.append(f"invalid_audit_row_fields:{guid}")
        if row.get("schema_version") != AUDIT_SCHEMA_VERSION or type(row.get("schema_version")) is not int:
            errors.append(f"invalid_audit_schema_version:{guid}")
        if not guid or guid in seen_guids:
            errors.append(f"duplicate_or_empty_audit_guid:{guid}")
        seen_guids.add(guid)
        if active:
            registry = active.get(guid)
            if registry is None:
                errors.append(f"unknown_audit_guid:{guid}")
            else:
                for field in ("word", "cefr", "list", "variant", "pos"):
                    if _trim(row.get(field)) != _trim(registry.get(field)):
                        errors.append(f"audit_identity_mismatch:{guid}:{field}")

        idiom_phrases = row.get("idiom_phrases")
        if not _is_valid_string_list(idiom_phrases):
            errors.append(f"invalid_idiom_phrases:{guid}")
            idiom_phrases = []
        else:
            canonical_idioms = [normalize_collocation(phrase) for phrase in idiom_phrases]
            if canonical_idioms != idiom_phrases or any(not phrase for phrase in canonical_idioms):
                errors.append(f"noncanonical_idiom_phrases:{guid}")
                idiom_phrases = canonical_idioms
        if row.get("idiom_fingerprint") != _idiom_fingerprint(idiom_phrases):
            errors.append(f"idiom_fingerprint_mismatch:{guid}")

        mappings = row.get("source_mappings")
        if not isinstance(mappings, list):
            errors.append(f"invalid_source_mappings:{guid}")
            mappings = []
        semantic_ids: set[str] = set()
        mapped_source_ids: set[str] = set()
        mapping_orders: list[int] = []
        for mapping in mappings:
            if not isinstance(mapping, dict):
                errors.append(f"invalid_source_mapping_type:{guid}")
                continue
            if set(mapping) != SOURCE_MAPPING_FIELDS:
                errors.append(f"invalid_source_mapping_fields:{guid}")
            semantic_id = _trim(mapping.get("semantic_sense_id"))
            order = mapping.get("order")
            source_ids = mapping.get("source_sense_ids")
            if not semantic_id or semantic_id in semantic_ids:
                errors.append(f"duplicate_or_empty_semantic_mapping:{guid}:{semantic_id}")
            semantic_ids.add(semantic_id)
            if type(order) is not int or order < 1:
                errors.append(f"invalid_semantic_mapping_order:{guid}:{semantic_id}")
            else:
                mapping_orders.append(order)
            if not isinstance(source_ids, list) or any(
                not isinstance(value, str) or not value for value in source_ids
            ) or len(source_ids or []) != len(set(source_ids or [])):
                errors.append(f"invalid_mapping_source_ids:{guid}:{semantic_id}")
                source_ids = []
            mapped_source_ids.update(source_ids)
        if mapping_orders != sorted(mapping_orders) or len(mapping_orders) != len(set(mapping_orders)):
            errors.append(f"non_deterministic_semantic_mapping_order:{guid}")

        evidence = row.get("source_evidence")
        if not isinstance(evidence, list):
            errors.append(f"invalid_source_evidence:{guid}")
            evidence = []
        evidence_ids: set[str] = set()
        evidence_source_ids: set[str] = set()
        for item in evidence:
            if not isinstance(item, dict):
                errors.append(f"invalid_evidence_type:{guid}")
                continue
            evidence_id = _trim(item.get("evidence_id"))
            if set(item) != EVIDENCE_FIELDS:
                errors.append(f"invalid_evidence_fields:{guid}:{evidence_id}")
            if not evidence_id or evidence_id in evidence_ids:
                errors.append(f"duplicate_or_empty_evidence_id:{guid}:{evidence_id}")
            evidence_ids.add(evidence_id)
            source_id = _trim(item.get("source_sense_id"))
            evidence_source_ids.add(source_id)
            semantic_for_evidence = item.get("semantic_sense_ids")
            if not _is_valid_string_list(semantic_for_evidence) or set(semantic_for_evidence) - semantic_ids:
                errors.append(f"invalid_evidence_semantic_ids:{guid}:{evidence_id}")
                semantic_for_evidence = []
            raw = {field: item.get(field) for field in (
                "text", "source", "origin", "evidence_kind", "example_index",
                "example_text", "container_index", "item_index", "category",
                "truncated", "full_entry_url",
            )}
            try:
                normalized = _normalize_source_evidence(
                    raw,
                    record_source=_trim(item.get("source")).casefold(),
                    source_headword=item.get("source_headword"),
                    source_sense=source_id,
                    semantic_ids=list(semantic_for_evidence),
                )
                if normalized != item:
                    errors.append(f"evidence_contract_mismatch:{guid}:{evidence_id}")
            except ValueError:
                errors.append(f"invalid_evidence_contract:{guid}:{evidence_id}")
        if evidence_source_ids - mapped_source_ids:
            errors.append(f"evidence_for_unmapped_source_sense:{guid}")

        expected_candidates = _build_candidates(guid, evidence)
        candidates = row.get("mandatory_candidates")
        if not isinstance(candidates, list):
            errors.append(f"invalid_mandatory_candidates:{guid}")
            candidates = []
        candidate_ids: set[str] = set()
        final_ids = {
            item.get("final_item_id") for item in row.get("final_items") or []
            if isinstance(item, dict)
        }
        current = row.get("current_items")
        if not isinstance(current, list):
            errors.append(f"invalid_current_items:{guid}")
            current = []
        current_ids: set[str] = set()
        current_by_id: dict[str, dict] = {}
        exact_evidence_by_text: dict[str, list[str]] = defaultdict(list)
        for item in evidence:
            if isinstance(item, dict):
                exact_evidence_by_text[normalize_collocation(item.get("text"))].append(item.get("evidence_id"))
        current_orders: list[int] = []
        for item in current:
            if not isinstance(item, dict):
                errors.append(f"invalid_current_item_type:{guid}")
                continue
            current_id = _trim(item.get("current_item_id"))
            if set(item) != CURRENT_ITEM_FIELDS:
                errors.append(f"invalid_current_item_fields:{guid}:{current_id}")
            text = item.get("text")
            order = item.get("order")
            if not isinstance(text, str) or not text.strip():
                errors.append(f"invalid_current_text:{guid}:{current_id}")
            if type(order) is not int or order < 1:
                errors.append(f"invalid_current_order:{guid}:{current_id}")
            else:
                current_orders.append(order)
            if current_id != collocation_current_item_id(guid, order, text) or current_id in current_ids:
                errors.append(f"invalid_or_duplicate_current_item_id:{guid}:{current_id}")
            current_ids.add(current_id)
            current_by_id[current_id] = item
            if item.get("evidence_ids") != exact_evidence_by_text.get(normalize_collocation(text), []):
                errors.append(f"current_evidence_coverage_mismatch:{guid}:{current_id}")
            decision = item.get("decision")
            if decision not in CURRENT_DECISIONS:
                errors.append(f"invalid_current_decision:{guid}:{current_id}:{decision}")
            targets = item.get("target_final_item_ids")
            if not isinstance(targets, list) or any(
                not isinstance(value, str) or not value for value in targets
            ) or len(targets or []) != len(set(targets or [])):
                errors.append(f"invalid_current_targets:{guid}:{current_id}")
                targets = []
            if set(targets) - final_ids:
                errors.append(f"unknown_current_target:{guid}:{current_id}")
            if decision in {"keep_source", "keep_curated"} and len(targets) != 1:
                errors.append(f"keep_current_requires_one_target:{guid}:{current_id}")
            if decision == "keep_source" and not item.get("evidence_ids"):
                errors.append(f"keep_source_without_evidence:{guid}:{current_id}")
            if decision == "rewrite_or_split" and not targets:
                errors.append(f"rewrite_without_target:{guid}:{current_id}")
            if decision in {"remove", "pending", "uncertain"} and targets:
                errors.append(f"nonkeeping_current_has_target:{guid}:{current_id}")
            if decision in {"keep_curated", "rewrite_or_split", "remove"} and not _trim(item.get("reason")):
                errors.append(f"current_reason_required:{guid}:{current_id}")
            prefix = f"current_item:{guid}:{current_id}"
            errors.extend(_validate_resolved_review(
                item,
                prefix,
                require_complete=require_complete,
                resolved={"keep_source", "keep_curated", "rewrite_or_split", "remove"},
            ))
            if require_complete and decision in {
                "keep_source",
                "keep_curated",
                "rewrite_or_split",
                "remove",
            }:
                reason = normalize_collocation(item.get("reason"))
                surface = normalize_collocation(text)
                if not reason or surface not in reason:
                    errors.append(f"review_reason_missing_surface:{guid}:{current_id}")
                if decision == "keep_source":
                    missing_evidence = [
                        evidence_id
                        for evidence_id in item.get("evidence_ids") or []
                        if evidence_id not in str(item.get("reason") or "")
                    ]
                    if missing_evidence:
                        errors.append(
                            f"review_reason_missing_evidence:{guid}:{current_id}"
                        )
                normalized_reason = _normalize_reason(item.get("reason"))
                owner = reason_owners.setdefault(normalized_reason, prefix)
                if owner != prefix:
                    errors.append(f"duplicate_bulk_review_reason:{owner}:{prefix}")
        if sorted(current_orders) != list(range(1, len(current) + 1)) or len(current_orders) != len(current):
            errors.append(f"non_contiguous_current_order:{guid}")
        if row.get("current_fingerprint") != _current_fingerprint(current):
            errors.append(f"current_fingerprint_mismatch:{guid}")

        expected_candidate_by_id = {
            item["candidate_id"]: _candidate_immutable(item)
            for item in expected_candidates
        }
        for candidate in candidates:
            if not isinstance(candidate, dict):
                errors.append(f"invalid_candidate_type:{guid}")
                continue
            candidate_id = _trim(candidate.get("candidate_id"))
            if set(candidate) != CANDIDATE_FIELDS:
                errors.append(f"invalid_candidate_fields:{guid}:{candidate_id}")
            if candidate_id in candidate_ids:
                errors.append(f"duplicate_candidate_id:{guid}:{candidate_id}")
            candidate_ids.add(candidate_id)
            if _candidate_immutable(candidate) != expected_candidate_by_id.get(candidate_id):
                errors.append(f"candidate_evidence_coverage_mismatch:{guid}:{candidate_id}")
            decision = candidate.get("decision")
            if decision not in CANDIDATE_DECISIONS:
                errors.append(f"invalid_candidate_decision:{guid}:{candidate_id}:{decision}")
            targets = candidate.get("target_final_item_ids")
            if not isinstance(targets, list) or any(
                not isinstance(value, str) or not value for value in targets
            ) or len(targets or []) != len(set(targets or [])):
                errors.append(f"invalid_candidate_targets:{guid}:{candidate_id}")
                targets = []
            if set(targets) - final_ids:
                errors.append(f"unknown_candidate_target:{guid}:{candidate_id}")
            if decision in {"included", "covered"} and len(targets) != 1:
                errors.append(f"included_candidate_requires_one_target:{guid}:{candidate_id}")
            if decision in {"excluded", "pending", "uncertain"} and targets:
                errors.append(f"nonincluded_candidate_has_target:{guid}:{candidate_id}")
            if decision == "excluded" and not _trim(candidate.get("reason")):
                errors.append(f"candidate_exclusion_reason_required:{guid}:{candidate_id}")
            prefix = f"candidate:{guid}:{candidate_id}"
            errors.extend(_validate_resolved_review(
                candidate,
                prefix,
                require_complete=require_complete,
                resolved={"included", "covered", "excluded"},
            ))
            if require_complete and decision in {"included", "covered", "excluded"}:
                reason = normalize_collocation(candidate.get("reason"))
                surface = normalize_collocation(candidate.get("text"))
                if not reason or surface not in reason:
                    errors.append(f"review_reason_missing_surface:{guid}:{candidate_id}")
                if decision in {"included", "covered"}:
                    missing_evidence = [
                        evidence_id
                        for evidence_id in candidate.get("evidence_ids") or []
                        if evidence_id not in str(candidate.get("reason") or "")
                    ]
                    if missing_evidence:
                        errors.append(
                            f"review_reason_missing_evidence:{guid}:{candidate_id}"
                        )
            if require_complete and decision == "excluded":
                normalized_reason = _normalize_reason(candidate.get("reason"))
                owner = reason_owners.setdefault(normalized_reason, prefix)
                if owner != prefix:
                    errors.append(f"duplicate_bulk_review_reason:{owner}:{prefix}")
        if set(candidate_ids) != set(expected_candidate_by_id):
            errors.append(f"mandatory_candidate_set_mismatch:{guid}")
        if row.get("source_fingerprint") != _source_fingerprint(mappings, evidence, candidates):
            errors.append(f"source_fingerprint_mismatch:{guid}")
        if row.get("input_fingerprint") != _input_fingerprint(row):
            errors.append(f"input_fingerprint_mismatch:{guid}")

        errors.extend(_validate_final_items(row, require_complete=require_complete))
        final_by_id = {
            item.get("final_item_id"): item for item in row.get("final_items") or []
            if isinstance(item, dict)
        }
        for item in current:
            if not isinstance(item, dict) or item.get("decision") not in {"keep_source", "keep_curated"}:
                continue
            target_ids = item.get("target_final_item_ids") or []
            if len(target_ids) != 1 or target_ids[0] not in final_by_id:
                continue
            final = final_by_id[target_ids[0]]
            if normalize_collocation(final.get("text")) != normalize_collocation(item.get("text")):
                errors.append(f"kept_current_text_mismatch:{guid}:{item.get('current_item_id')}")
            if item.get("decision") == "keep_source":
                if final.get("source") == "curated" or not set(item.get("evidence_ids") or []).issubset(set(final.get("evidence_ids") or [])):
                    errors.append(f"keep_source_target_mismatch:{guid}:{item.get('current_item_id')}")
            elif final.get("source") != "curated":
                errors.append(f"keep_curated_target_mismatch:{guid}:{item.get('current_item_id')}")
        for candidate in candidates:
            if not isinstance(candidate, dict) or candidate.get("decision") not in {"included", "covered"}:
                continue
            targets = candidate.get("target_final_item_ids") or []
            if len(targets) != 1 or targets[0] not in final_by_id:
                continue
            final = final_by_id[targets[0]]
            if (
                final.get("source") == "curated"
                or normalize_collocation(final.get("text")) != normalize_collocation(candidate.get("text"))
                or not set(candidate.get("evidence_ids") or []).issubset(set(final.get("evidence_ids") or []))
            ):
                errors.append(f"candidate_target_mismatch:{guid}:{candidate.get('candidate_id')}")
            exact_current = any(
                normalize_collocation(current_by_id[current_id].get("text"))
                == normalize_collocation(candidate.get("text"))
                for current_id in final.get("current_item_ids") or []
                if current_id in current_by_id
            )
            if candidate.get("decision") == "covered" and not exact_current:
                errors.append(f"covered_candidate_not_current:{guid}:{candidate.get('candidate_id')}")
            if candidate.get("decision") == "included" and exact_current:
                errors.append(f"included_candidate_already_current:{guid}:{candidate.get('candidate_id')}")

        finals = row.get("final_items") or []
        if require_complete:
            idiom_set = set(idiom_phrases)
            for final in finals:
                if (
                    isinstance(final, dict)
                    and normalize_collocation(final.get("text")) in idiom_set
                ):
                    errors.append(
                        f"collocation_duplicates_idiom:{guid}:"
                        f"{_trim(final.get('final_item_id'))}"
                    )
        empty_fields = (
            _trim(row.get("empty_reason")),
            _trim(row.get("empty_reviewer")),
            _trim(row.get("empty_reviewed_at")),
            _trim(row.get("empty_approval")),
        )
        _validate_review_date(row.get("empty_reviewed_at"), f"empty_card:{guid}", errors)
        if row.get("empty_approval") not in APPROVAL_VALUES:
            errors.append(f"invalid_empty_approval:{guid}")
        if finals and any(empty_fields):
            errors.append(f"nonempty_card_has_empty_review:{guid}")
        if require_complete and not finals:
            if not empty_fields[0]:
                errors.append(f"missing_empty_reason:{guid}")
            if not empty_fields[1]:
                errors.append(f"missing_empty_reviewer:{guid}")
            if not empty_fields[2]:
                errors.append(f"missing_empty_reviewed_at:{guid}")
            if empty_fields[3] != "approved":
                errors.append(f"unapproved_empty_card:{guid}")
            normalized_reason = _normalize_reason(empty_fields[0])
            owner = reason_owners.setdefault(normalized_reason, f"empty_card:{guid}")
            if normalized_reason and owner != f"empty_card:{guid}":
                errors.append(f"duplicate_bulk_review_reason:{owner}:empty_card:{guid}")

    if active:
        for guid in sorted(set(active) - seen_guids):
            errors.append(f"missing_active_guid:{guid}")
        for guid in sorted(seen_guids - set(active)):
            errors.append(f"unknown_audit_guid:{guid}")
    return errors


def _audit_sha256(rows: Sequence[dict]) -> str:
    return hashlib.sha256(serialize_audit_rows(rows).encode("utf-8")).hexdigest()


def promote_audit_rows(
    audit_rows: Sequence[dict], registry_rows: Sequence[dict]
) -> list[dict]:
    """Deterministically promote a complete Collocation Audit."""
    errors = validate_audit_rows(audit_rows, registry_rows, require_complete=True)
    if errors:
        raise ValueError("Collocation Audit is not promotion-ready:\n" + "\n".join(errors))
    audit_hash = _audit_sha256(audit_rows)
    promoted: list[dict] = []
    for row in sorted(audit_rows, key=lambda item: item["guid"]):
        promoted.append({
            "schema_version": REGISTRY_SCHEMA_VERSION,
            "guid": row["guid"],
            "word": row["word"],
            "cefr": row["cefr"],
            "list": row["list"],
            "variant": row["variant"],
            "audit_sha256": audit_hash,
            "audit_row_sha256": _sha256(row),
            "idiom_fingerprint": row["idiom_fingerprint"],
            "current_fingerprint": row["current_fingerprint"],
            "source_fingerprint": row["source_fingerprint"],
            "items": [
                {
                    "text": item["text"],
                    "order": item["order"],
                    "source": item["source"],
                    "evidence_ids": list(item["evidence_ids"]),
                }
                for item in row["final_items"]
            ],
            "empty_reason": row["empty_reason"],
        })
    return promoted


def validate_registry_rows(
    rows: Sequence[dict],
    card_registry_rows: Sequence[dict],
    *,
    audit_rows: Sequence[dict] | None = None,
    current_source_fingerprints: Mapping[str, object] | None = None,
) -> list[str]:
    """Validate exact active GUID coverage and promoted item/source contracts."""
    errors: list[str] = []
    try:
        active = _active_registry_index(card_registry_rows)
    except ValueError as exc:
        return [f"invalid_card_registry:{exc}"]
    guids = [_trim(row.get("guid")) for row in rows if isinstance(row, dict)]
    if guids != sorted(guids):
        errors.append("non_deterministic_collocation_registry_order")
    seen: set[str] = set()
    audit_hashes: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            errors.append("invalid_collocation_registry_row_type")
            continue
        guid = _trim(row.get("guid"))
        if set(row) != REGISTRY_ROW_FIELDS:
            errors.append(f"invalid_collocation_registry_fields:{guid}")
        if row.get("schema_version") != REGISTRY_SCHEMA_VERSION or type(row.get("schema_version")) is not int:
            errors.append(f"invalid_collocation_registry_schema:{guid}")
        if not guid or guid in seen:
            errors.append(f"duplicate_or_empty_collocation_registry_guid:{guid}")
        seen.add(guid)
        registry = active.get(guid)
        if registry is None:
            errors.append(f"unknown_collocation_registry_guid:{guid}")
        else:
            for field in ("word", "cefr", "list", "variant"):
                if _trim(row.get(field)) != _trim(registry.get(field)):
                    errors.append(f"collocation_registry_identity_mismatch:{guid}:{field}")
        audit_hash = _trim(row.get("audit_sha256"))
        if len(audit_hash) != 64:
            errors.append(f"invalid_collocation_audit_sha256:{guid}")
        audit_hashes.add(audit_hash)
        if len(_trim(row.get("audit_row_sha256"))) != 64:
            errors.append(f"invalid_collocation_audit_row_sha256:{guid}")
        if len(_trim(row.get("idiom_fingerprint"))) != 64:
            errors.append(f"invalid_collocation_idiom_fingerprint:{guid}")
        items = row.get("items")
        if not isinstance(items, list):
            errors.append(f"invalid_collocation_registry_items:{guid}")
            items = []
        if len(items) > MAX_FINAL_ITEMS:
            errors.append(f"too_many_collocation_registry_items:{guid}:{len(items)}")
        normalized: set[str] = set()
        orders: list[int] = []
        source_ranks: list[int] = []
        for item in items:
            if not isinstance(item, dict):
                errors.append(f"invalid_collocation_registry_item_type:{guid}")
                continue
            if set(item) != REGISTRY_ITEM_FIELDS:
                errors.append(f"invalid_collocation_registry_item_fields:{guid}")
            text = _trim(item.get("text"))
            order = item.get("order")
            source = item.get("source")
            if type(order) is not int or order < 1:
                errors.append(f"invalid_collocation_registry_order:{guid}:{order}")
            else:
                orders.append(order)
            normalized_text = normalize_collocation(text)
            if not normalized_text or normalized_text in normalized:
                errors.append(f"duplicate_collocation_registry_text:{guid}:{text}")
            normalized.add(normalized_text)
            if source not in FINAL_SOURCES:
                errors.append(f"invalid_collocation_registry_source:{guid}:{source}")
            else:
                source_ranks.append(FINAL_SOURCE_ORDER[source])
            surrogate = {
                "final_item_id": collocation_final_item_id(guid, text),
                "text": item.get("text"),
                "source": source,
            }
            errors.extend(_final_text_errors(guid, surrogate))
            evidence_ids = item.get("evidence_ids")
            if not isinstance(evidence_ids, list) or any(
                not isinstance(value, str) or not value for value in evidence_ids
            ) or len(evidence_ids or []) != len(set(evidence_ids or [])):
                errors.append(f"invalid_collocation_registry_evidence:{guid}:{text}")
                evidence_ids = []
            if source == "curated" and evidence_ids:
                errors.append(f"curated_collocation_registry_evidence:{guid}:{text}")
            if source in {"oxford", "cambridge", "oxford+cambridge"} and not evidence_ids:
                errors.append(f"source_collocation_registry_without_evidence:{guid}:{text}")
        if orders != list(range(1, len(items) + 1)):
            errors.append(f"non_contiguous_collocation_registry_order:{guid}")
        if source_ranks != sorted(source_ranks):
            errors.append(f"invalid_collocation_registry_source_order:{guid}")
        if items and _trim(row.get("empty_reason")):
            errors.append(f"nonempty_collocation_registry_has_empty_reason:{guid}")
        if not items and not _trim(row.get("empty_reason")):
            errors.append(f"empty_collocation_registry_without_reason:{guid}")
        if current_source_fingerprints is not None and guid in current_source_fingerprints:
            expected = current_source_fingerprints[guid]
            if isinstance(expected, Mapping):
                expected_source = expected.get("source_fingerprint")
                expected_idiom = expected.get("idiom_fingerprint")
            else:
                expected_source = expected
                expected_idiom = None
            if _trim(row.get("source_fingerprint")) != _trim(expected_source):
                errors.append(f"stale_collocation_source_fingerprint:{guid}")
            if (
                expected_idiom is not None
                and _trim(row.get("idiom_fingerprint")) != _trim(expected_idiom)
            ):
                errors.append(f"stale_collocation_idiom_fingerprint:{guid}")
    if len(audit_hashes) > 1:
        errors.append("mixed_collocation_audit_sha256")
    for guid in sorted(set(active) - seen):
        errors.append(f"missing_collocation_registry_guid:{guid}")
    for guid in sorted(seen - set(active)):
        errors.append(f"unknown_collocation_registry_guid:{guid}")
    if audit_rows is not None:
        audit_errors = validate_audit_rows(
            audit_rows, card_registry_rows, require_complete=True
        )
        errors.extend(f"audit:{error}" for error in audit_errors)
        if not audit_errors:
            expected = promote_audit_rows(audit_rows, card_registry_rows)
            if list(rows) != expected:
                errors.append("collocation_registry_audit_projection_mismatch")
    return errors


def load_collocation_registry(
    path: Path,
    card_registry_rows: Sequence[dict],
    *,
    audit_rows: Sequence[dict] | None = None,
    current_source_fingerprints: Mapping[str, object] | None = None,
) -> list[dict]:
    rows = load_jsonl(path)
    errors = validate_registry_rows(
        rows,
        card_registry_rows,
        audit_rows=audit_rows,
        current_source_fingerprints=current_source_fingerprints,
    )
    if errors:
        raise ValueError("Invalid Collocation Registry:\n" + "\n".join(errors))
    return rows


def registry_payload_by_guid(rows: Sequence[dict]) -> dict[str, dict[str, str]]:
    payload: dict[str, dict[str, str]] = {}
    for row in rows:
        guid = row["guid"]
        if guid in payload:
            raise ValueError(f"Duplicate Collocation Registry GUID: {guid}")
        raw_items = row.get("items") or []
        raw_orders = [
            item.get("order") for item in raw_items
            if isinstance(item, dict)
        ]
        if raw_orders != list(range(1, len(raw_items) + 1)):
            raise ValueError(f"Non-contiguous Collocation Registry order: {guid}")
        items = sorted(
            raw_items,
            key=lambda item: item.get("order", 10**9) if isinstance(item, dict) else 10**9,
        )
        if len(items) > MAX_FINAL_ITEMS:
            raise ValueError(f"Collocation Registry exceeds five items: {guid}")
        normalized_seen: set[str] = set()
        orders: list[int] = []
        source_ranks: list[int] = []
        for item in items:
            if not isinstance(item, dict) or set(item) != REGISTRY_ITEM_FIELDS:
                raise ValueError(f"Malformed Collocation Registry item: {guid}")
            text = _trim(item.get("text"))
            item_id = collocation_final_item_id(guid, text)
            errors = _final_text_errors(
                guid,
                {"final_item_id": item_id, "text": item.get("text"), "source": item.get("source")},
            )
            if errors:
                raise ValueError("; ".join(errors))
            normalized = normalize_collocation(text)
            if normalized in normalized_seen:
                raise ValueError(f"Duplicate Collocation Registry text: {guid}:{text}")
            normalized_seen.add(normalized)
            if item.get("source") not in FINAL_SOURCES:
                raise ValueError(f"Invalid Collocation Registry source: {guid}:{item.get('source')}")
            if item.get("source") == "curated":
                if item.get("evidence_ids"):
                    raise ValueError(f"Curated Collocation Registry item claims evidence: {guid}:{text}")
            elif not item.get("evidence_ids"):
                raise ValueError(f"Source Collocation Registry item lacks evidence: {guid}:{text}")
            order = item.get("order")
            if type(order) is not int or order < 1:
                raise ValueError(f"Invalid Collocation Registry order: {guid}:{text}")
            orders.append(order)
            source_ranks.append(FINAL_SOURCE_ORDER[item["source"]])
        if orders != list(range(1, len(items) + 1)):
            raise ValueError(f"Non-contiguous Collocation Registry order: {guid}")
        if source_ranks != sorted(source_ranks):
            raise ValueError(f"Invalid Collocation Registry source order: {guid}")
        if not items and not _trim(row.get("empty_reason")):
            raise ValueError(f"Empty Collocation Registry row lacks reason: {guid}")
        if items and _trim(row.get("empty_reason")):
            raise ValueError(f"Non-empty Collocation Registry row has empty reason: {guid}")
        payload[guid] = {
            "collocations": "|".join(item["text"] for item in items),
            "collocation_sources": "|".join(item["source"] for item in items),
        }
    return payload


def apply_collocation_registry(cards: Sequence[object], rows: Sequence[dict]) -> list[object]:
    """Apply promoted text/provenance to dict rows or ``BuiltCard`` values."""
    payload = registry_payload_by_guid(rows)
    registry_by_guid = {row["guid"]: row for row in rows}
    seen: set[str] = set()
    updated: list[object] = []
    for card in cards:
        if isinstance(card, dict):
            guid = _trim(card.get("guid") or card.get("GUID"))
            card_word = _trim(card.get("word") or card.get("Word"))
            card_cefr = _trim(card.get("cefr") or card.get("CEFRLevel"))
        else:
            guid = _trim(getattr(card, "guid", ""))
            card_word = _trim(getattr(card, "word", ""))
            card_cefr = _trim(getattr(card, "cefr", ""))
        if not guid or guid in seen:
            raise ValueError(f"Duplicate or empty built-card GUID: {guid}")
        seen.add(guid)
        values = payload.get(guid)
        if values is None:
            raise ValueError(f"Collocation Registry is missing built-card GUID: {guid}")
        registry_row = registry_by_guid[guid]
        if card_word and card_word != _trim(registry_row.get("word")):
            raise ValueError(f"Collocation Registry word mismatch for built-card GUID: {guid}")
        if card_cefr and card_cefr != _trim(registry_row.get("cefr")):
            raise ValueError(f"Collocation Registry CEFR mismatch for built-card GUID: {guid}")
        if isinstance(card, dict):
            copy_card = dict(card)
            copy_card.update(values)
            if "Collocations" in copy_card:
                copy_card["Collocations"] = values["collocations"]
            if "CollocationSources" in copy_card:
                copy_card["CollocationSources"] = values["collocation_sources"]
        elif hasattr(card, "_replace"):
            copy_card = card._replace(**values)
        else:
            raise TypeError(f"Unsupported built-card type: {type(card).__name__}")
        updated.append(copy_card)
    extras = sorted(set(payload) - seen)
    if extras:
        raise ValueError(f"Collocation Registry has unknown built-card GUIDs: {extras[:5]}")
    return updated


def audit_summary(rows: Sequence[dict]) -> dict[str, int]:
    current = Counter(
        item.get("decision") or "missing"
        for row in rows
        for item in row.get("current_items") or []
    )
    candidates = Counter(
        item.get("decision") or "missing"
        for row in rows
        for item in row.get("mandatory_candidates") or []
    )
    return {
        "cards": len(rows),
        "current_items": sum(current.values()),
        "current_pending": current["pending"],
        "mandatory_candidates": sum(candidates.values()),
        "candidate_pending": candidates["pending"],
        "source_evidence": sum(len(row.get("source_evidence") or []) for row in rows),
        "final_items": sum(len(row.get("final_items") or []) for row in rows),
        "empty_cards": sum(not (row.get("final_items") or []) for row in rows),
    }


def _excel_value(value: object) -> object:
    if isinstance(value, (list, dict)):
        return _canonical_json(value)
    if value is None:
        return ""
    return value


def _append_literal_row(sheet, values: list[object]) -> None:
    sheet.append(values)
    row_number = sheet.max_row
    for column, value in enumerate(values, 1):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            sheet.cell(row_number, column).data_type = "s"


def _style_sheet(sheet, color: str, widths: Mapping[str, int], columns: Sequence[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for name, width in widths.items():
        sheet.column_dimensions[get_column_letter(columns.index(name) + 1)].width = width


def export_workbook(audit_rows: Sequence[dict], path: Path) -> None:
    errors = validate_audit_rows(audit_rows)
    if errors:
        raise ValueError("Cannot export invalid Collocation Audit:\n" + "\n".join(errors))
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    cards = workbook.create_sheet("Cards")
    current_sheet = workbook.create_sheet("Current Items")
    candidate_sheet = workbook.create_sheet("Source Candidates")
    final_sheet = workbook.create_sheet("Final Items")
    evidence_sheet = workbook.create_sheet("Evidence")
    instructions = workbook.create_sheet("Instructions")

    overview.append(["Metric", "Value"])
    for key, value in audit_summary(audit_rows).items():
        overview.append([key, value])
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 18

    cards.append(list(CARD_COLUMNS))
    current_sheet.append(list(CURRENT_REVIEW_COLUMNS))
    candidate_sheet.append(list(CANDIDATE_REVIEW_COLUMNS))
    final_sheet.append(list(FINAL_REVIEW_COLUMNS))
    evidence_sheet.append(list(EVIDENCE_COLUMNS))
    for row in audit_rows:
        _append_literal_row(cards, [_excel_value(row.get(column)) for column in CARD_COLUMNS])
        for item in row["current_items"]:
            values = {**item, "guid": row["guid"], "input_fingerprint": row["input_fingerprint"]}
            _append_literal_row(current_sheet, [_excel_value(values.get(column)) for column in CURRENT_REVIEW_COLUMNS])
        for item in row["mandatory_candidates"]:
            values = {**item, "guid": row["guid"], "input_fingerprint": row["input_fingerprint"]}
            _append_literal_row(candidate_sheet, [_excel_value(values.get(column)) for column in CANDIDATE_REVIEW_COLUMNS])
        for item in row["final_items"]:
            values = {**item, "guid": row["guid"]}
            _append_literal_row(final_sheet, [_excel_value(values.get(column)) for column in FINAL_REVIEW_COLUMNS])
        for item in row["source_evidence"]:
            values = {**item, "guid": row["guid"], "input_fingerprint": row["input_fingerprint"]}
            _append_literal_row(evidence_sheet, [_excel_value(values.get(column)) for column in EVIDENCE_COLUMNS])

    _style_sheet(
        cards,
        "1F4E78",
        {"word": 22, "idiom_phrases": 48, "empty_reason": 48},
        CARD_COLUMNS,
    )
    _style_sheet(current_sheet, "806000", {"text": 38, "evidence_ids": 45, "reason": 48}, CURRENT_REVIEW_COLUMNS)
    _style_sheet(candidate_sheet, "548235", {"text": 38, "evidence_ids": 45, "reason": 48}, CANDIDATE_REVIEW_COLUMNS)
    _style_sheet(final_sheet, "7030A0", {"text": 38, "evidence_ids": 45, "current_item_ids": 45}, FINAL_REVIEW_COLUMNS)
    _style_sheet(
        evidence_sheet,
        "5B9BD5",
        {
            "source_headword": 22,
            "text": 38,
            "example_text": 55,
            "full_entry_url": 45,
        },
        EVIDENCE_COLUMNS,
    )

    for sheet, columns, decision_name, decisions in (
        (current_sheet, CURRENT_REVIEW_COLUMNS, "decision", CURRENT_DECISIONS),
        (candidate_sheet, CANDIDATE_REVIEW_COLUMNS, "decision", CANDIDATE_DECISIONS),
    ):
        last_row = max(2, sheet.max_row)
        decision = DataValidation(type="list", formula1='"' + ",".join(decisions) + '"')
        approval = DataValidation(type="list", formula1='"approved,rejected"', allow_blank=True)
        sheet.add_data_validation(decision)
        sheet.add_data_validation(approval)
        decision_letter = get_column_letter(columns.index(decision_name) + 1)
        approval_letter = get_column_letter(columns.index("approval") + 1)
        decision.add(f"{decision_letter}2:{decision_letter}{last_row}")
        approval.add(f"{approval_letter}2:{approval_letter}{last_row}")
        sheet.conditional_formatting.add(
            f"A2:{get_column_letter(len(columns))}{last_row}",
            FormulaRule(
                formula=[f'${decision_letter}2="uncertain"'],
                fill=PatternFill("solid", fgColor="F4CCCC"),
            ),
        )
    final_validation = DataValidation(
        type="list", formula1='"' + ",".join(FINAL_SOURCES) + '"'
    )
    final_sheet.add_data_validation(final_validation)
    final_source_letter = get_column_letter(FINAL_REVIEW_COLUMNS.index("source") + 1)
    final_validation.add(f"{final_source_letter}2:{final_source_letter}10000")
    empty_approval = DataValidation(type="list", formula1='"approved,rejected"', allow_blank=True)
    cards.add_data_validation(empty_approval)
    empty_letter = get_column_letter(CARD_COLUMNS.index("empty_approval") + 1)
    empty_approval.add(f"{empty_letter}2:{empty_letter}{max(2, cards.max_row)}")

    instructions.append(["Collocation Audit"])
    instructions.append(["Review every row in Current Items and Source Candidates; no item is approved automatically."])
    instructions.append(["Only evidence_kind=example_linked creates a mandatory source candidate. Supporting evidence may substantiate a source-backed current item but does not require inclusion."])
    instructions.append(["Add the reviewed output in Final Items. Use exact separate source phrases; source-backed rows must not compress alternatives with '/'."])
    instructions.append(["Order Oxford and Oxford+Cambridge items first, Cambridge items next, then curated items in their existing order. Maximum five final items per card."])
    instructions.append(["Use covered when an exact current item already accounts for a mandatory candidate; use included when adding the missing exact phrase; exclusions need a row-specific reason."])
    instructions.append(["Every resolved item needs reviewer, ISO review date, and approval=approved. keep_curated, rewrite_or_split, remove, and excluded also require a unique row-specific reason."])
    instructions.append(["Idioms owns every phrase listed in Cards.idiom_phrases. Do not retain an exact normalized duplicate in Final Items; any Idioms phrase change invalidates reused review state."])
    instructions.append(["Evidence.source_headword records the raw dictionary entry headword. It is provenance for morphology review, not an automatic rejection rule."])
    instructions.append(["Edit only review fields plus Final Items. Cards, Evidence, and immutable item columns are fingerprint-protected; JSONL remains canonical."])
    instructions.column_dimensions["A"].width = 118
    for row in instructions.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.row == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _parse_json_cell(value: object, label: str) -> list:
    if value in (None, ""):
        return []
    if not isinstance(value, str):
        raise ValueError(f"{label} must be JSON text")
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{label} contains malformed JSON") from exc
    if not isinstance(parsed, list):
        raise ValueError(f"{label} must contain a JSON list")
    return parsed


def _sheet_rows(workbook, name: str, columns: Sequence[str]) -> list[dict]:
    if name not in workbook.sheetnames:
        raise ValueError(f"Workbook is missing the {name} sheet")
    sheet = workbook[name]
    headers = [cell.value for cell in sheet[1]]
    if headers != list(columns):
        raise ValueError(f"{name} sheet columns do not match the Collocation Audit contract")
    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append({column: value for column, value in zip(columns, values)})
    return rows


def _workbook_value(column: str, value: object) -> object:
    if column in {
        "evidence_ids",
        "sources",
        "target_final_item_ids",
        "current_item_ids",
        "semantic_sense_ids",
        "idiom_phrases",
    }:
        return _parse_json_cell(value, column)
    if column in {"schema_version", "order", "example_index", "container_index", "item_index"}:
        if value in (None, "") and column in {"example_index", "container_index", "item_index"}:
            return None
        return int(value)
    if column == "truncated":
        if type(value) is bool:
            return value
        if value in (0, 1):
            return bool(value)
        raise ValueError("truncated must be boolean")
    if column in {"category", "example_text", "full_entry_url"}:
        return None if value in (None, "") else str(value)
    return "" if value is None else str(value)


def import_workbook(audit_rows: Sequence[dict], path: Path) -> list[dict]:
    """Import editable item decisions/final rows after immutable verification."""
    workbook = load_workbook(path, data_only=False)
    result = copy.deepcopy(list(audit_rows))
    by_guid = {row["guid"]: row for row in result}

    card_rows = _sheet_rows(workbook, "Cards", CARD_COLUMNS)
    if len(card_rows) != len(by_guid):
        raise ValueError("Cards sheet does not have exact audit GUID coverage")
    seen_cards: set[str] = set()
    for raw in card_rows:
        guid = _trim(raw.get("guid"))
        row = by_guid.get(guid)
        if row is None or guid in seen_cards:
            raise ValueError(f"Unknown or duplicate Cards row: {guid}")
        seen_cards.add(guid)
        for column in CARD_IMMUTABLE_COLUMNS:
            actual = _workbook_value(column, raw.get(column))
            if actual != row[column]:
                raise ValueError(f"Immutable Cards column {column!r} changed for {guid}")
        for column in CARD_EDITABLE_COLUMNS:
            row[column] = _workbook_value(column, raw.get(column))

    def import_items(
        sheet_name: str,
        columns: Sequence[str],
        collection_name: str,
        id_name: str,
        immutable: Sequence[str],
        editable: Sequence[str],
    ) -> None:
        expected: dict[tuple[str, str], dict] = {}
        for row in result:
            for item in row[collection_name]:
                expected[(row["guid"], item[id_name])] = item
        updates = _sheet_rows(workbook, sheet_name, columns)
        seen: set[tuple[str, str]] = set()
        for raw in updates:
            key = (_trim(raw.get("guid")), _trim(raw.get(id_name)))
            original = expected.get(key)
            if original is None or key in seen:
                raise ValueError(f"Unknown or duplicate {sheet_name} row: {key}")
            seen.add(key)
            row = by_guid[key[0]]
            expanded = {**original, "guid": key[0], "input_fingerprint": row["input_fingerprint"]}
            for column in immutable:
                actual = _workbook_value(column, raw.get(column))
                if actual != expanded[column]:
                    raise ValueError(
                        f"Immutable {sheet_name} column {column!r} changed for {key[1]}"
                    )
            for column in editable:
                original[column] = _workbook_value(column, raw.get(column))
        if seen != set(expected):
            missing = sorted(set(expected) - seen)
            raise ValueError(f"{sheet_name} is missing review rows: {missing[:5]}")

    import_items(
        "Current Items",
        CURRENT_REVIEW_COLUMNS,
        "current_items",
        "current_item_id",
        CURRENT_IMMUTABLE_COLUMNS,
        CURRENT_EDITABLE_COLUMNS,
    )
    import_items(
        "Source Candidates",
        CANDIDATE_REVIEW_COLUMNS,
        "mandatory_candidates",
        "candidate_id",
        CANDIDATE_IMMUTABLE_COLUMNS,
        CANDIDATE_EDITABLE_COLUMNS,
    )

    expected_evidence: dict[tuple[str, str], dict] = {}
    for row in result:
        for item in row["source_evidence"]:
            expected_evidence[(row["guid"], item["evidence_id"])] = {
                **item,
                "guid": row["guid"],
                "input_fingerprint": row["input_fingerprint"],
            }
    evidence_rows = _sheet_rows(workbook, "Evidence", EVIDENCE_COLUMNS)
    seen_evidence: set[tuple[str, str]] = set()
    for raw in evidence_rows:
        key = (_trim(raw.get("guid")), _trim(raw.get("evidence_id")))
        expected = expected_evidence.get(key)
        if expected is None or key in seen_evidence:
            raise ValueError(f"Unknown or duplicate Evidence row: {key}")
        seen_evidence.add(key)
        for column in EVIDENCE_COLUMNS:
            if _workbook_value(column, raw.get(column)) != expected[column]:
                raise ValueError(f"Immutable Evidence column {column!r} changed for {key[1]}")
    if seen_evidence != set(expected_evidence):
        raise ValueError("Evidence sheet is missing immutable evidence rows")

    for row in result:
        row["final_items"] = []
    final_rows = _sheet_rows(workbook, "Final Items", FINAL_REVIEW_COLUMNS)
    seen_final: set[tuple[str, str]] = set()
    for raw in final_rows:
        guid = _trim(raw.get("guid"))
        if guid not in by_guid:
            raise ValueError(f"Unknown Final Items GUID: {guid}")
        final = {
            "final_item_id": _workbook_value("final_item_id", raw.get("final_item_id")),
            "text": _workbook_value("text", raw.get("text")),
            "order": _workbook_value("order", raw.get("order")),
            "source": _workbook_value("source", raw.get("source")),
            "evidence_ids": _workbook_value("evidence_ids", raw.get("evidence_ids")),
            "current_item_ids": _workbook_value("current_item_ids", raw.get("current_item_ids")),
        }
        key = (guid, final["final_item_id"])
        if key in seen_final:
            raise ValueError(f"Duplicate Final Items row: {key}")
        seen_final.add(key)
        by_guid[guid]["final_items"].append(final)
    for row in result:
        row["final_items"].sort(key=lambda item: (item.get("order", 10**9), item.get("final_item_id", "")))

    errors = validate_audit_rows(result)
    if errors:
        raise ValueError("Workbook review is invalid:\n" + "\n".join(errors))
    return result


# Naming aliases mirror the semantic/idiom audit modules and keep integration
# callers independent from the collocation-specific serializer name.
serialize_jsonl = serialize_audit_rows
promoted_collocations_by_guid = registry_payload_by_guid
