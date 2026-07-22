"""Phrase-level bilingual Idiom Audit contracts and XLSX helpers."""
from __future__ import annotations

import copy
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.deck_builder.card_identity import primary_list_from_tags
from src.deck_builder.text_integrity import has_suspected_lossy_unicode


AUDIT_SCHEMA_VERSION = 1
IDIOM_ID_RE = re.compile(r"^idm_[0-9a-f]{24}$")
DISPLAY_MODES = ("", "vi_equivalent", "bilingual_gloss")
EQUIVALENCE_KINDS = ("", "proverb", "idiom", "saying", "none")
DECISIONS = ("pending", "pass", "uncertain", "rejected")
CONFIDENCE_VALUES = ("", "high", "medium", "low")
APPROVAL_VALUES = ("", "approved", "rejected")

OCCURRENCE_FIELDS = {
    "guid",
    "word",
    "cefr",
    "list",
    "variant",
    "order",
    "example",
    "provenance",
    "phrase_en",
    "source_explanation_en",
    "source_fingerprint",
}
ROW_FIELDS = {
    "schema_version",
    "idiom_id",
    "phrase_en",
    "source_explanation_en",
    "source_examples",
    "occurrences",
    "content_fingerprint",
    "coverage_fingerprint",
    "display_mode",
    "equivalence_kind",
    "explanation_en_simple",
    "explanation_vi",
    "decision",
    "confidence",
    "review_reason",
    "reviewer",
    "reviewed_at",
    "approval",
    "translation_provenance",
}
IMMUTABLE_COLUMNS = (
    "schema_version",
    "idiom_id",
    "phrase_en",
    "source_explanation_en",
    "source_examples",
    "occurrences",
    "content_fingerprint",
    "coverage_fingerprint",
)
EDITABLE_COLUMNS = (
    "display_mode",
    "equivalence_kind",
    "explanation_en_simple",
    "explanation_vi",
    "decision",
    "confidence",
    "review_reason",
    "reviewer",
    "reviewed_at",
    "approval",
    "translation_provenance",
)
REVIEW_COLUMNS = IMMUTABLE_COLUMNS + EDITABLE_COLUMNS

_LOSSY_UNICODE_REVIEW_FIELDS = (
    "explanation_en_simple",
    "explanation_vi",
    "review_reason",
    "translation_provenance",
)


def _canonical_json(value: object) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _sha256(value: object) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _trim(value: object) -> str:
    return str(value or "").strip()


def _semantic_text(value: object) -> str:
    """Normalize identity text without deleting punctuation or learning slots."""
    return " ".join(unicodedata.normalize("NFKC", _trim(value)).split()).casefold()


def _semantic_key(phrase_en: object, source_explanation_en: object) -> tuple[str, str]:
    return _semantic_text(phrase_en), _semantic_text(source_explanation_en)


def _idiom_id(phrase_en: object, source_explanation_en: object) -> str:
    digest = _sha256({
        "phrase_en": _semantic_text(phrase_en),
        "source_explanation_en": _semantic_text(source_explanation_en),
    })[:24]
    return f"idm_{digest}"


def idiom_source_fingerprint(
    phrase_en: str,
    source_explanation_en: str,
    examples: Iterable[str],
) -> str:
    """Hash exact trimmed source content using canonical sorted-key JSON."""
    return _sha256({
        "phrase_en": _trim(phrase_en),
        "source_explanation_en": _trim(source_explanation_en),
        "examples": [_trim(example) for example in examples],
    })


def parse_serialized_idioms(value: object) -> list[dict]:
    """Parse the production ``phrase :: EN :: example`` / ``$$`` grammar."""
    raw = _trim(value)
    if not raw:
        return []

    parsed: list[dict] = []
    for ordinal, raw_entry in enumerate(raw.split("$$"), 1):
        if not raw_entry.strip():
            raise ValueError(f"empty idiom entry at ordinal {ordinal}")
        parts = [part.strip() for part in raw_entry.split("::", 2)]
        if len(parts) < 2 or not parts[0] or not parts[1]:
            raise ValueError(f"idiom entry {ordinal} requires phrase and English explanation")
        examples: list[str] = []
        if len(parts) == 3:
            raw_examples = parts[2]
            if not raw_examples:
                raise ValueError(f"empty idiom example cell at ordinal {ordinal}")
            examples = [part.strip() for part in raw_examples.split("|")]
            if any(not example for example in examples):
                raise ValueError(f"empty idiom example at ordinal {ordinal}")
        parsed.append({
            "phrase_en": parts[0],
            "source_explanation_en": parts[1],
            "examples": examples,
        })
    return parsed


def load_jsonl(path: Path) -> list[dict]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def serialize_jsonl(rows: Iterable[dict]) -> str:
    ordered = sorted(rows, key=lambda row: str(row.get("idiom_id") or ""))
    return "".join(_canonical_json(row) + "\n" for row in ordered)


def _field(row: dict, lower: str, upper: str | None = None) -> object:
    if lower in row:
        return row.get(lower)
    return row.get(upper or lower.capitalize())


def _card_provenance(card: dict) -> str:
    explicit = card.get("idiom_provenance") or card.get("provenance")
    if isinstance(explicit, str):
        return explicit.strip()
    if explicit:
        return _canonical_json(explicit)
    sources: list[str] = []
    for field in ("source1", "source2", "Source1", "Source2"):
        source = _trim(card.get(field))
        if source and source not in sources:
            sources.append(source)
    return " + ".join(sources) or "unknown"


def _active_registry_index(registry_rows: list[dict]) -> dict[str, dict]:
    indexed: dict[str, dict] = {}
    for row in registry_rows:
        if not isinstance(row, dict):
            raise ValueError("registry rows must be objects")
        if row.get("status") != "active":
            continue
        guid = _trim(row.get("guid"))
        if not guid or guid in indexed:
            raise ValueError(f"duplicate or empty active registry GUID: {guid}")
        indexed[guid] = row
    return indexed


def _occurrence_sort_key(occurrence: object) -> tuple:
    if not isinstance(occurrence, dict):
        return ("\uffff", 1, 0, "", "", "", "", "", _canonical_json(occurrence))
    order = occurrence.get("order")
    valid_order = isinstance(order, int) and not isinstance(order, bool)
    return (
        _trim(occurrence.get("guid")),
        0 if valid_order else 1,
        order if valid_order else 0,
        "" if valid_order else _trim(order),
        _trim(occurrence.get("word")),
        _trim(occurrence.get("cefr")),
        _trim(occurrence.get("list")),
        _trim(occurrence.get("variant")),
        "",
    )


def _source_examples(occurrences: list[dict]) -> list[str]:
    examples: list[str] = []
    for occurrence in sorted(occurrences, key=_occurrence_sort_key):
        if not isinstance(occurrence, dict):
            continue
        example = _trim(occurrence.get("example"))
        if example and example not in examples:
            examples.append(example)
    return examples


def _coverage_fingerprint(occurrences: list[dict]) -> str:
    return _sha256(sorted(occurrences, key=_occurrence_sort_key))


def build_audit_rows(cards: list[dict], registry_rows: list[dict]) -> list[dict]:
    """Build one pending row per normalized phrase-and-source-meaning key."""
    registry_by_guid = _active_registry_index(registry_rows)
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    seen_cards: set[str] = set()

    for card in cards:
        guid = _trim(_field(card, "guid", "GUID"))
        if not guid or guid in seen_cards:
            raise ValueError(f"duplicate or empty built-card GUID: {guid}")
        seen_cards.add(guid)
        registry = registry_by_guid.get(guid)
        if registry is None:
            continue
        parsed = parse_serialized_idioms(_field(card, "idioms", "Idioms"))
        if len(parsed) > 2:
            raise ValueError(f"card {guid} has more than two selected idioms")

        word = _trim(registry.get("word") or _field(card, "word", "Word"))
        cefr = _trim(registry.get("cefr") or _field(card, "cefr", "CEFRLevel"))
        card_list = _trim(
            registry.get("list")
            or card.get("list")
            or primary_list_from_tags(_field(card, "tags", "Tags"), canonical=True)
        )
        variant = _trim(registry.get("variant") or card.get("variant"))
        provenance = _card_provenance(card)

        for order, idiom in enumerate(parsed, 1):
            if len(idiom["examples"]) > 1:
                raise ValueError(f"card {guid} idiom {order} has more than one example")
            phrase = idiom["phrase_en"]
            explanation = idiom["source_explanation_en"]
            examples = idiom["examples"]
            occurrence = {
                "guid": guid,
                "word": word,
                "cefr": cefr,
                "list": card_list,
                "variant": variant,
                "order": order,
                "example": examples[0] if examples else "",
                "provenance": provenance,
                "phrase_en": phrase,
                "source_explanation_en": explanation,
                "source_fingerprint": idiom_source_fingerprint(
                    phrase, explanation, examples
                ),
            }
            grouped[_semantic_key(phrase, explanation)].append(occurrence)

    rows: list[dict] = []
    for occurrences in grouped.values():
        occurrences = sorted(occurrences, key=_occurrence_sort_key)
        representative = occurrences[0]
        phrase = representative["phrase_en"]
        explanation = representative["source_explanation_en"]
        source_examples = _source_examples(occurrences)
        rows.append({
            "schema_version": AUDIT_SCHEMA_VERSION,
            "idiom_id": _idiom_id(phrase, explanation),
            "phrase_en": phrase,
            "source_explanation_en": explanation,
            "source_examples": source_examples,
            "occurrences": occurrences,
            "content_fingerprint": idiom_source_fingerprint(
                phrase, explanation, source_examples
            ),
            "coverage_fingerprint": _coverage_fingerprint(occurrences),
            "display_mode": "",
            "equivalence_kind": "",
            "explanation_en_simple": "",
            "explanation_vi": "",
            "decision": "pending",
            "confidence": "",
            "review_reason": "",
            "reviewer": "",
            "reviewed_at": "",
            "approval": "",
            "translation_provenance": "",
        })
    return sorted(rows, key=lambda row: row["idiom_id"])


def _has_control(value: str) -> bool:
    return any(unicodedata.category(char) == "Cc" for char in value)


def _invalid_text(
    value: object,
    *,
    required: bool = False,
    reject_example_pipe: bool = False,
) -> bool:
    if not isinstance(value, str) or (required and not value.strip()):
        return True
    if not isinstance(value, str):
        return True
    return (
        _has_control(value)
        or "$$" in value
        or "::" in value
        or (reject_example_pipe and "|" in value)
    )


def _validate_review(row: dict, idiom_id: str, require_complete: bool) -> list[str]:
    errors: list[str] = []
    mode = row.get("display_mode")
    kind = row.get("equivalence_kind")
    decision = row.get("decision")
    confidence = row.get("confidence")
    approval = row.get("approval")

    if mode not in DISPLAY_MODES:
        errors.append(f"invalid_display_mode:{idiom_id}")
    if kind not in EQUIVALENCE_KINDS:
        errors.append(f"invalid_equivalence_kind:{idiom_id}")
    if decision not in DECISIONS:
        errors.append(f"invalid_decision:{idiom_id}")
    if confidence not in CONFIDENCE_VALUES:
        errors.append(f"invalid_confidence:{idiom_id}")
    if approval not in APPROVAL_VALUES:
        errors.append(f"invalid_approval:{idiom_id}")

    for field in (
        "explanation_en_simple",
        "explanation_vi",
        "review_reason",
        "reviewer",
        "reviewed_at",
        "translation_provenance",
    ):
        if _invalid_text(row.get(field, ""), reject_example_pipe=True):
            errors.append(f"invalid_review_text:{idiom_id}:{field}")

    for field in _LOSSY_UNICODE_REVIEW_FIELDS:
        if has_suspected_lossy_unicode(row.get(field, "")):
            errors.append(f"suspected_lossy_unicode:{idiom_id}:{field}")

    if mode == "vi_equivalent":
        if kind not in {"proverb", "idiom", "saying"}:
            errors.append(f"vi_equivalent_requires_equivalence_kind:{idiom_id}")
        if not _trim(row.get("explanation_vi")):
            errors.append(f"vi_equivalent_requires_vietnamese:{idiom_id}")
        if _trim(row.get("explanation_en_simple")):
            errors.append(f"vi_equivalent_forbids_simple_english:{idiom_id}")
    elif mode == "bilingual_gloss":
        if kind != "none":
            errors.append(f"bilingual_gloss_requires_none_kind:{idiom_id}")
        if not _trim(row.get("explanation_en_simple")):
            errors.append(f"bilingual_gloss_requires_simple_english:{idiom_id}")
        if not _trim(row.get("explanation_vi")):
            errors.append(f"bilingual_gloss_requires_vietnamese:{idiom_id}")

    reviewed_at = _trim(row.get("reviewed_at"))
    if reviewed_at:
        try:
            date.fromisoformat(reviewed_at)
        except ValueError:
            errors.append(f"invalid_reviewed_at:{idiom_id}")

    if require_complete:
        if decision != "pass":
            errors.append(f"incomplete_decision:{idiom_id}:{decision}")
        if not mode:
            errors.append(f"missing_display_mode:{idiom_id}")
        if confidence not in {"high", "medium", "low"}:
            errors.append(f"missing_confidence:{idiom_id}")
        if confidence in {"medium", "low"} and approval != "approved":
            errors.append(f"approval_required:{idiom_id}")
        if approval == "rejected":
            errors.append(f"rejected_approval:{idiom_id}")
        for field in (
            "review_reason",
            "reviewer",
            "reviewed_at",
            "translation_provenance",
        ):
            if not _trim(row.get(field)):
                errors.append(f"missing_review_field:{idiom_id}:{field}")
    return errors


def validate_audit_rows(
    rows: list[dict],
    registry_rows: list[dict] | None = None,
    *,
    require_complete: bool = False,
) -> list[str]:
    """Validate structure, immutable source coverage, and review completeness."""
    errors: list[str] = []
    seen_ids: set[str] = set()
    seen_keys: set[tuple[str, str]] = set()
    seen_occurrences: set[tuple[str, int]] = set()
    occurrences_by_guid: dict[str, list[int]] = defaultdict(list)
    active_by_guid: dict[str, dict] = {}

    if registry_rows is not None:
        try:
            active_by_guid = _active_registry_index(registry_rows)
        except ValueError as exc:
            errors.append(f"invalid_registry:{exc}")

    registry_supplied = registry_rows is not None
    row_ids = [str(row.get("idiom_id") or "") for row in rows if isinstance(row, dict)]
    if row_ids != sorted(row_ids):
        errors.append("non_deterministic_row_order")

    for row in rows:
        if not isinstance(row, dict):
            errors.append("invalid_row_type")
            continue
        idiom_id = _trim(row.get("idiom_id"))
        if set(row) != ROW_FIELDS:
            errors.append(f"invalid_row_fields:{idiom_id}")
        if type(row.get("schema_version")) is not int or row.get("schema_version") != AUDIT_SCHEMA_VERSION:
            errors.append(f"invalid_schema_version:{idiom_id}")
        if not IDIOM_ID_RE.fullmatch(idiom_id) or idiom_id in seen_ids:
            errors.append(f"duplicate_or_invalid_idiom_id:{idiom_id}")
        seen_ids.add(idiom_id)

        phrase = row.get("phrase_en")
        explanation = row.get("source_explanation_en")
        if _invalid_text(phrase, required=True):
            errors.append(f"invalid_source_text:{idiom_id}:phrase_en")
        if _invalid_text(explanation, required=True, reject_example_pipe=True):
            errors.append(f"invalid_source_text:{idiom_id}:source_explanation_en")
        key = _semantic_key(phrase, explanation)
        if not all(key) or key in seen_keys:
            errors.append(f"duplicate_or_empty_semantic_key:{idiom_id}")
        seen_keys.add(key)
        if idiom_id != _idiom_id(phrase, explanation):
            errors.append(f"idiom_id_mismatch:{idiom_id}")

        occurrences = row.get("occurrences")
        if not isinstance(occurrences, list) or not occurrences:
            errors.append(f"invalid_occurrences:{idiom_id}")
            occurrences = []
        elif occurrences != sorted(occurrences, key=_occurrence_sort_key):
            errors.append(f"non_deterministic_occurrence_order:{idiom_id}")

        for occurrence in occurrences:
            if not isinstance(occurrence, dict):
                errors.append(f"invalid_occurrence_type:{idiom_id}")
                continue
            guid = _trim(occurrence.get("guid"))
            order = occurrence.get("order")
            if set(occurrence) != OCCURRENCE_FIELDS:
                errors.append(f"invalid_occurrence_fields:{idiom_id}:{guid}")
            if not guid or type(order) is not int or order < 1:
                errors.append(f"invalid_occurrence_identity:{idiom_id}:{guid}:{order}")
                continue
            occurrence_key = (guid, order)
            if occurrence_key in seen_occurrences:
                errors.append(f"duplicate_occurrence:{guid}:{order}")
            seen_occurrences.add(occurrence_key)
            occurrences_by_guid[guid].append(order)

            for field in (
                "guid",
                "word",
                "cefr",
                "list",
                "variant",
                "provenance",
                "phrase_en",
                "source_explanation_en",
            ):
                required = field in {
                    "guid", "word", "cefr", "list", "provenance",
                    "phrase_en", "source_explanation_en",
                }
                reject_pipe = field not in {"guid", "phrase_en"}
                if _invalid_text(
                    occurrence.get(field),
                    required=required,
                    reject_example_pipe=reject_pipe,
                ):
                    errors.append(f"invalid_occurrence_text:{idiom_id}:{guid}:{field}")
            if _invalid_text(
                occurrence.get("example", ""), reject_example_pipe=True
            ):
                errors.append(f"invalid_occurrence_text:{idiom_id}:{guid}:example")
            occurrence_key_text = _semantic_key(
                occurrence.get("phrase_en"),
                occurrence.get("source_explanation_en"),
            )
            if occurrence_key_text != key:
                errors.append(f"occurrence_semantic_key_mismatch:{idiom_id}:{guid}:{order}")
            occurrence_examples = (
                [_trim(occurrence.get("example"))]
                if _trim(occurrence.get("example"))
                else []
            )
            expected_source_fingerprint = idiom_source_fingerprint(
                occurrence.get("phrase_en") or "",
                occurrence.get("source_explanation_en") or "",
                occurrence_examples,
            )
            if occurrence.get("source_fingerprint") != expected_source_fingerprint:
                errors.append(f"source_fingerprint_mismatch:{idiom_id}:{guid}:{order}")

            if registry_supplied:
                registry = active_by_guid.get(guid)
                if registry is None:
                    errors.append(f"unknown_occurrence_guid:{idiom_id}:{guid}")
                else:
                    for field in ("word", "cefr", "list", "variant"):
                        if _trim(occurrence.get(field)) != _trim(registry.get(field)):
                            errors.append(f"occurrence_identity_mismatch:{idiom_id}:{guid}:{field}")

        source_examples = row.get("source_examples")
        if not isinstance(source_examples, list):
            errors.append(f"invalid_source_examples:{idiom_id}")
            source_examples = []
        else:
            if all(isinstance(example, str) for example in source_examples) and len(source_examples) != len(set(source_examples)):
                errors.append(f"duplicate_source_examples:{idiom_id}")
            for example in source_examples:
                if _invalid_text(example, required=True, reject_example_pipe=True):
                    errors.append(f"invalid_source_example:{idiom_id}")
            if source_examples != _source_examples(occurrences):
                errors.append(f"source_examples_mismatch:{idiom_id}")

        expected_content = idiom_source_fingerprint(
            phrase or "", explanation or "", source_examples
        )
        if row.get("content_fingerprint") != expected_content:
            errors.append(f"content_fingerprint_mismatch:{idiom_id}")
        if row.get("coverage_fingerprint") != _coverage_fingerprint(occurrences):
            errors.append(f"coverage_fingerprint_mismatch:{idiom_id}")

        errors.extend(_validate_review(row, idiom_id, require_complete))

    for guid, orders in occurrences_by_guid.items():
        if sorted(orders) != list(range(1, len(orders) + 1)) or len(orders) > 2:
            errors.append(f"invalid_card_idiom_order:{guid}")
    return errors


def _excel_value(column: str, value: object) -> object:
    if column in {"source_examples", "occurrences"}:
        return _canonical_json(value)
    return value


def _append_literal_row(sheet, values: list[object]) -> None:
    sheet.append(values)
    row_number = sheet.max_row
    for column, value in enumerate(values, 1):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            sheet.cell(row_number, column).data_type = "s"


def export_workbook(audit_rows: list[dict], path: Path) -> None:
    """Export an editable review view while keeping JSONL canonical."""
    errors = validate_audit_rows(audit_rows)
    if errors:
        raise ValueError("Cannot export invalid Idiom Audit:\n" + "\n".join(errors))

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    review = wb.create_sheet("Review")
    instructions = wb.create_sheet("Instructions")

    summary = audit_summary(audit_rows)
    overview.append(["Metric", "Value"])
    for key, value in summary.items():
        overview.append([key, value])
    for row in overview.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.row == 1)
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 18

    review.append(list(REVIEW_COLUMNS))
    for row in audit_rows:
        _append_literal_row(
            review,
            [_excel_value(column, row.get(column, "")) for column in REVIEW_COLUMNS],
        )
    review.freeze_panes = "A2"
    review.auto_filter.ref = review.dimensions
    review.sheet_view.showGridLines = False
    for cell in review[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="9C6500")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in review.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "idiom_id": 31,
        "phrase_en": 36,
        "source_explanation_en": 55,
        "source_examples": 55,
        "occurrences": 50,
        "explanation_en_simple": 45,
        "explanation_vi": 45,
        "review_reason": 45,
        "translation_provenance": 30,
    }
    for name, width in widths.items():
        review.column_dimensions[get_column_letter(REVIEW_COLUMNS.index(name) + 1)].width = width
    for name in ("schema_version", "content_fingerprint", "coverage_fingerprint"):
        review.column_dimensions[get_column_letter(REVIEW_COLUMNS.index(name) + 1)].hidden = True

    validations = {
        "display_mode": DISPLAY_MODES[1:],
        "equivalence_kind": EQUIVALENCE_KINDS[1:],
        "decision": DECISIONS,
        "confidence": CONFIDENCE_VALUES[1:],
        "approval": APPROVAL_VALUES[1:],
    }
    last_row = max(2, len(audit_rows) + 1)
    for column, values in validations.items():
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=column in {"display_mode", "equivalence_kind", "confidence", "approval"},
        )
        review.add_data_validation(validation)
        letter = get_column_letter(REVIEW_COLUMNS.index(column) + 1)
        validation.add(f"{letter}2:{letter}{last_row}")
    decision_letter = get_column_letter(REVIEW_COLUMNS.index("decision") + 1)
    review.conditional_formatting.add(
        f"A2:{get_column_letter(len(REVIEW_COLUMNS))}{last_row}",
        FormulaRule(
            formula=[f'${decision_letter}2="uncertain"'],
            fill=PatternFill("solid", fgColor="F4CCCC"),
        ),
    )

    instructions.append(["Bilingual Idiom Audit"])
    instructions.append(["Edit only columns from display_mode through translation_provenance."])
    instructions.append(["Use vi_equivalent for an established Vietnamese proverb, idiom, saying, or fixed figurative expression with an equivalent or clearly related meaning; exact imagery and pragmatic scope need not match."])
    instructions.append(["Canonical examples: get back on the rails -> đâu lại vào đấy; be at odds -> trống đánh xuôi, kèn thổi ngược."])
    instructions.append(["For vi_equivalent choose proverb/idiom/saying and leave simple English empty; use bilingual_gloss only when no clear established semantic relation exists or the candidate would mislead."])
    instructions.append(["For bilingual_gloss choose none and write short, natural learner glosses in both English and Vietnamese; keep only the core meaning, and do not mirror English sentence structure in Vietnamese."])
    instructions.append(["There is no hard word limit: keep extra detail only when shortening would lose a material condition or restriction."])
    instructions.append(["Do not repeat somebody/something or other visible learning-pattern slots as placeholder subjects, objects, or subordinate clauses; prefer compact lexical pairs such as twist somebody’s arm -> persuade/pressure / thuyết phục/nài ép."])
    instructions.append(["Unchanged is not reviewed by default. To retain an existing bilingual_gloss, record the shorter wording considered and the exact material meaning it would lose, or cite an exact user-locked canonical pair; never bulk-pass unchanged rows."])
    instructions.append(["Cambridge-backed canonical fallback: put somebody to the sword -> kill / giết."])
    instructions.append(["Canonical bilingual examples: an old wives’ tale -> an old belief that is not true / quan niệm dân gian sai lầm; shake/rock the foundations ... -> seriously weaken something at its core / làm lung lay tận gốc."])
    instructions.column_dimensions["A"].width = 115
    for row in instructions.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.row == 1)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _workbook_immutable_value(column: str, value: object) -> object:
    if column in {"source_examples", "occurrences"}:
        if not isinstance(value, str):
            raise ValueError(f"Immutable JSON column {column!r} is not text")
        try:
            return json.loads(value)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Immutable JSON column {column!r} is malformed") from exc
    if column == "schema_version":
        return int(value)
    return "" if value is None else str(value)


def import_workbook(audit_rows: list[dict], path: Path) -> list[dict]:
    """Import editable columns after exact immutable/fingerprint verification."""
    originals = {row["idiom_id"]: row for row in audit_rows}
    wb = load_workbook(path, data_only=False)
    if "Review" not in wb.sheetnames:
        raise ValueError("Workbook is missing the Review sheet")
    sheet = wb["Review"]
    headers = [cell.value for cell in sheet[1]]
    if headers != list(REVIEW_COLUMNS):
        raise ValueError("Review sheet columns do not match the Idiom Audit contract")

    updates: dict[str, dict] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        raw = {header: value for header, value in zip(headers, values)}
        idiom_id = _trim(raw.get("idiom_id"))
        original = originals.get(idiom_id)
        if original is None:
            raise ValueError(f"Unknown review row: {idiom_id}")
        if idiom_id in updates:
            raise ValueError(f"Duplicate review row: {idiom_id}")
        for column in IMMUTABLE_COLUMNS:
            actual = _workbook_immutable_value(column, raw.get(column))
            if actual != original[column]:
                raise ValueError(f"Immutable column {column!r} changed for {idiom_id}")
        updates[idiom_id] = {
            column: "" if raw.get(column) is None else str(raw.get(column))
            for column in EDITABLE_COLUMNS
        }
    if set(updates) != set(originals):
        missing = sorted(set(originals) - set(updates))
        raise ValueError(f"Workbook is missing review rows: {missing[:5]}")

    result = copy.deepcopy(audit_rows)
    for row in result:
        row.update(updates[row["idiom_id"]])
    errors = validate_audit_rows(result)
    if errors:
        raise ValueError("Workbook review is invalid:\n" + "\n".join(errors))
    return result


def apply_review_bundle(audit_rows: list[dict], decisions: list[dict]) -> list[dict]:
    """Apply fingerprint-bound, idiom-keyed editable updates transactionally."""
    result = copy.deepcopy(audit_rows)
    by_id = {row["idiom_id"]: row for row in result}
    seen: set[str] = set()
    required_keys = {"idiom_id", "content_fingerprint", "coverage_fingerprint"}
    allowed_keys = required_keys | set(EDITABLE_COLUMNS)

    for decision in decisions:
        if not isinstance(decision, dict):
            raise ValueError("Review bundle rows must be objects")
        idiom_id = _trim(decision.get("idiom_id"))
        if idiom_id not in by_id or idiom_id in seen:
            raise ValueError(f"Unknown or duplicate review idiom_id: {idiom_id}")
        seen.add(idiom_id)
        unknown = set(decision) - allowed_keys
        missing = required_keys - set(decision)
        if unknown or missing:
            raise ValueError(
                f"Invalid review bundle fields for {idiom_id}: missing={sorted(missing)} unknown={sorted(unknown)}"
            )
        target = by_id[idiom_id]
        for fingerprint in ("content_fingerprint", "coverage_fingerprint"):
            if decision[fingerprint] != target[fingerprint]:
                raise ValueError(f"Stale review bundle fingerprint for {idiom_id}: {fingerprint}")
        editable = set(decision) & set(EDITABLE_COLUMNS)
        if not editable:
            raise ValueError(f"Review bundle has no editable fields for {idiom_id}")
        for field in editable:
            value = decision[field]
            target[field] = "" if value is None else str(value)

    errors = validate_audit_rows(result)
    if errors:
        raise ValueError("Review bundle is invalid:\n" + "\n".join(errors))
    return result


def promoted_idioms_by_guid(rows: list[dict]) -> dict[str, list[dict]]:
    """Return complete reviewed idioms aligned to each card occurrence."""
    errors = validate_audit_rows(rows, require_complete=True)
    if errors:
        raise ValueError("Idiom Audit is not promotion-ready:\n" + "\n".join(errors))

    promoted: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        for occurrence in row["occurrences"]:
            examples = [occurrence["example"]] if occurrence["example"] else []
            explanation_en = (
                occurrence["source_explanation_en"]
                if row["display_mode"] == "vi_equivalent"
                else row["explanation_en_simple"]
            )
            promoted[occurrence["guid"]].append({
                "idiom_id": row["idiom_id"],
                "order": occurrence["order"],
                "source_fingerprint": idiom_source_fingerprint(
                    occurrence["phrase_en"],
                    occurrence["source_explanation_en"],
                    examples,
                ),
                "phrase_en": occurrence["phrase_en"],
                "display_mode": row["display_mode"],
                "explanation_en": explanation_en,
                "explanation_vi": row["explanation_vi"],
                "examples": examples,
                "translation_provenance": row["translation_provenance"],
            })
    return {
        guid: sorted(items, key=lambda item: item["order"])
        for guid, items in sorted(promoted.items())
    }


def audit_summary(rows: list[dict]) -> dict[str, int]:
    decisions = Counter(row.get("decision") or "missing" for row in rows)
    modes = Counter(row.get("display_mode") or "unassigned" for row in rows)
    return {
        "semantic_rows": len(rows),
        "occurrences": sum(len(row.get("occurrences") or []) for row in rows),
        "cards": len({
            occurrence.get("guid")
            for row in rows
            for occurrence in row.get("occurrences") or []
            if occurrence.get("guid")
        }),
        "pending": decisions["pending"],
        "pass": decisions["pass"],
        "uncertain": decisions["uncertain"],
        "rejected": decisions["rejected"],
        "vi_equivalent": modes["vi_equivalent"],
        "bilingual_gloss": modes["bilingual_gloss"],
        "unassigned": modes["unassigned"],
    }
