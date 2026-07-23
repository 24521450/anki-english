"""Fingerprint-bound review of Oxford phrasal-verb routing."""
from __future__ import annotations

import copy
import hashlib
import json
import re
from datetime import date
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation


SCHEMA_VERSION = 1
DISPOSITIONS = (
    "pending",
    "distinct_secondary",
    "same_owned_sense",
    "existing_phrase_card",
    "core_pattern",
    "excluded_not_relevant",
    "uncertain",
)
APPROVALS = ("", "approved", "rejected")
ROW_FIELDS = {
    "schema_version", "route_id", "parent_guid", "parent_word", "phrase",
    "target_url", "target_source_files", "target_source_sense_ids",
    "structural_collision_guids", "source_fingerprint", "input_fingerprint",
    "disposition", "target_guid", "owned_semantic_sense_ids", "reason",
    "reviewer", "reviewed_at", "approval",
}
IMMUTABLE_COLUMNS = (
    "schema_version", "route_id", "parent_guid", "parent_word", "phrase",
    "target_url", "target_source_files", "target_source_sense_ids",
    "structural_collision_guids", "source_fingerprint", "input_fingerprint",
)
EDITABLE_COLUMNS = (
    "disposition", "target_guid", "owned_semantic_sense_ids", "reason",
    "reviewer", "reviewed_at", "approval",
)
WORKBOOK_COLUMNS = IMMUTABLE_COLUMNS + EDITABLE_COLUMNS
_TOKEN_RE = re.compile(r"[^\W_]+(?:['’][^\W_]+)?", re.UNICODE)


def _canonical_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _sha256(value: object) -> str:
    return hashlib.sha256(_canonical_json(value).encode("utf-8")).hexdigest()


def _text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _norm(value: object) -> str:
    return _text(value).casefold()


def load_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def serialize_rows(rows: Iterable[dict]) -> str:
    return "".join(_canonical_json(row) + "\n" for row in sorted(rows, key=lambda item: item["route_id"]))


def expand_explicit_slash_phrases(surface: object) -> list[str]:
    """Expand slash tokens only; do not stem, lemmatize, or infer variants."""
    tokens = _text(surface).split()
    if not tokens:
        return []
    phrases = [""]
    for token in tokens:
        alternatives = token.split("/")
        if any(not alternative for alternative in alternatives):
            return []
        phrases = [f"{prefix} {alternative}".strip() for prefix in phrases for alternative in alternatives]
    return phrases


def phrase_starts_learner_surface(phrase: object, surface: object) -> bool:
    """Match the complete phrase at token zero, with explicit slash expansion only."""
    phrase_tokens = _TOKEN_RE.findall(_norm(phrase))
    if not phrase_tokens:
        return False
    for expanded in expand_explicit_slash_phrases(surface):
        tokens = _TOKEN_RE.findall(_norm(expanded))
        if tokens[:len(phrase_tokens)] == phrase_tokens:
            return True
    return False


def structural_phrase_collisions(phrase: object, registry_rows: Sequence[dict]) -> list[str]:
    collisions = []
    for row in registry_rows:
        if row.get("status") != "active":
            continue
        if phrase_starts_learner_surface(phrase, row.get("word")):
            guid = _text(row.get("guid"))
            if guid:
                collisions.append(guid)
    return sorted(set(collisions))


def _source_sense_ids(record: Mapping[str, object]) -> list[str]:
    ids = []
    for pos_index, pos_data in enumerate(record.get("pos_data") or [], 1):
        for sense_index, definition in enumerate(pos_data.get("definitions") or [], 1):
            ids.append("pvsrc_" + _sha256({
                "word": record.get("word"),
                "source_files": record.get("source_files") or [],
                "pos_index": pos_index,
                "sense_index": sense_index,
                "definition": definition,
            })[:24])
    return ids


def _route_sources(
    oxford_rows: Sequence[dict],
    *,
    included_routes: set[tuple[str, str]] | None = None,
) -> list[tuple[dict, dict]]:
    active_rows = [row for row in oxford_rows if not row.get("_skip")]
    by_url: dict[str, dict] = {}
    for row in active_rows:
        for pos_data in row.get("pos_data") or []:
            source_url = _text(pos_data.get("source_url")).split("#", 1)[0].rstrip("/")
            if source_url:
                by_url[source_url] = row
    result = []
    seen: set[tuple[str, str]] = set()
    for parent in active_rows:
        for link in parent.get("phrasal_verb_links") or []:
            phrase = _text(link.get("phrase"))
            if included_routes is not None and (_norm(parent.get("word")), _norm(phrase)) not in included_routes:
                continue
            target_url = _text(link.get("url")).split("#", 1)[0].rstrip("/")
            target = by_url.get(target_url)
            if target is None or _norm(target.get("word")) != _norm(phrase):
                raise ValueError(f"missing authoritative Oxford target: {phrase} ({target_url})")
            expected_slug = _norm(phrase).replace(" ", "-")
            target_slug = target_url.rsplit("/", 1)[-1]
            if not re.fullmatch(re.escape(expected_slug) + r"(?:_\d+)?", target_slug):
                raise ValueError(f"invalid authoritative Oxford target URL: {phrase}")
            if not _source_sense_ids(target):
                raise ValueError(f"unhydrated phrasal target: {phrase}")
            identity = (_norm(parent.get("word")), target_url)
            if identity not in seen:
                seen.add(identity)
                result.append((parent, target))
    return sorted(result, key=lambda pair: (_norm(pair[0].get("word")), _norm(pair[1].get("word"))))


def _active_cards(registry_rows: Sequence[dict]) -> dict[str, dict]:
    result = {}
    for row in registry_rows:
        if row.get("status") != "active":
            continue
        guid = _text(row.get("guid"))
        if not guid or guid in result:
            raise ValueError(f"duplicate or empty active Card Registry GUID: {guid}")
        result[guid] = row
    return result


def build_audit_rows(
    registry_rows: Sequence[dict],
    oxford_rows: Sequence[dict],
    *,
    existing_rows: Sequence[dict] = (),
    collocation_audit_rows: Sequence[dict] | None = None,
) -> list[dict]:
    active = _active_cards(registry_rows)
    cards_by_word: dict[str, list[dict]] = {}
    for card in active.values():
        cards_by_word.setdefault(_norm(card.get("word")), []).append(card)
    included_routes: set[tuple[str, str]] | None = None
    if collocation_audit_rows is not None:
        surfaces_by_guid: dict[str, list[str]] = {}
        for row in collocation_audit_rows:
            guid = _text(row.get("guid"))
            surfaces_by_guid[guid] = [
                _text(item.get("text"))
                for field in ("current_items", "mandatory_candidates", "final_items")
                for item in row.get(field) or []
                if _text(item.get("text"))
            ]
        included_routes = set()
        for parent in oxford_rows:
            parent_cards = cards_by_word.get(_norm(parent.get("word")), [])
            if not parent_cards:
                continue
            for link in parent.get("phrasal_verb_links") or []:
                phrase = _text(link.get("phrase"))
                exact_card_collision = bool(structural_phrase_collisions(phrase, registry_rows))
                audit_collision = any(
                    phrase_starts_learner_surface(phrase, surface)
                    for card in parent_cards
                    for surface in surfaces_by_guid.get(_text(card.get("guid")), [])
                )
                if exact_card_collision or audit_collision:
                    included_routes.add((_norm(parent.get("word")), _norm(phrase)))
    existing = {row.get("route_id"): row for row in existing_rows if isinstance(row, dict)}
    rows = []
    for parent, target in _route_sources(oxford_rows, included_routes=included_routes):
        parent_cards = cards_by_word.get(_norm(parent.get("word")), [])
        if collocation_audit_rows is not None and parent_cards:
            phrase = _text(target.get("word"))
            matched_parent_cards = [
                card for card in parent_cards
                if any(
                    phrase_starts_learner_surface(phrase, surface)
                    for surface in surfaces_by_guid.get(_text(card.get("guid")), [])
                )
            ]
            if matched_parent_cards:
                parent_cards = matched_parent_cards
        if not parent_cards:
            parent_cards = [
                active[guid]
                for guid in structural_phrase_collisions(target.get("word"), registry_rows)
                if _norm(active[guid].get("word")) == _norm(target.get("word"))
            ]
        if not parent_cards:
            base = _norm(parent.get("word"))
            parent_cards = [
                card for card in active.values()
                if _norm(card.get("word")).startswith(base + " ")
                and "phrasal verb" in _norm(card.get("pos"))
            ]
        if not parent_cards:
            continue
        if len(parent_cards) != 1:
            raise ValueError(
                f"phrasal route requires one active base or exact-phrase card: {parent.get('word')}"
            )
        card = parent_cards[0]
        phrase = _text(target.get("word"))
        target_url = _text((target.get("pos_data") or [{}])[0].get("source_url"))
        source_payload = {
            "parent_source_files": parent.get("source_files") or [],
            "phrase": phrase,
            "target_url": target_url,
            "target_source_files": target.get("source_files") or [],
            "target_source_sense_ids": _source_sense_ids(target),
        }
        route_id = "pvr_" + _sha256({"parent_guid": card["guid"], **source_payload})[:24]
        immutable = {
            "schema_version": SCHEMA_VERSION,
            "route_id": route_id,
            "parent_guid": card["guid"],
            "parent_word": _text(card.get("word")),
            "phrase": phrase,
            "target_url": target_url,
            "target_source_files": list(target.get("source_files") or []),
            "target_source_sense_ids": source_payload["target_source_sense_ids"],
            "structural_collision_guids": structural_phrase_collisions(phrase, registry_rows),
            "source_fingerprint": _sha256(source_payload),
        }
        immutable["input_fingerprint"] = _sha256(immutable)
        row = {**immutable, "disposition": "pending", "target_guid": "",
               "owned_semantic_sense_ids": [], "reason": "", "reviewer": "",
               "reviewed_at": "", "approval": ""}
        old = existing.get(route_id)
        if old and all(old.get(field) == row.get(field) for field in IMMUTABLE_COLUMNS):
            for field in EDITABLE_COLUMNS:
                row[field] = copy.deepcopy(old.get(field))
        rows.append(row)
    return sorted(rows, key=lambda row: row["route_id"])


def _collocation_items_by_guid(rows: Sequence[dict]) -> dict[str, list[str]]:
    return {
        _text(row.get("guid")): [_text(item.get("text")) for item in row.get("final_items") or []]
        for row in rows if isinstance(row, dict)
    }


def validate_audit_rows(
    rows: Sequence[dict],
    registry_rows: Sequence[dict],
    *,
    collocation_audit_rows: Sequence[dict] = (),
    require_complete: bool = False,
) -> list[str]:
    errors = []
    active = _active_cards(registry_rows)
    collisions_by_parent = _collocation_items_by_guid(collocation_audit_rows)
    if [row.get("route_id") for row in rows] != sorted(row.get("route_id") for row in rows):
        errors.append("non_deterministic_phrasal_route_order")
    seen = set()
    for row in rows:
        route_id = _text(row.get("route_id"))
        if set(row) != ROW_FIELDS:
            errors.append(f"invalid_phrasal_route_fields:{route_id}")
        if row.get("schema_version") != SCHEMA_VERSION:
            errors.append(f"invalid_phrasal_route_schema:{route_id}")
        if not route_id or route_id in seen:
            errors.append(f"duplicate_or_empty_phrasal_route:{route_id}")
        seen.add(route_id)
        parent_guid = _text(row.get("parent_guid"))
        if parent_guid not in active:
            errors.append(f"unknown_phrasal_parent_guid:{route_id}")
        expected_collisions = structural_phrase_collisions(row.get("phrase"), registry_rows)
        if row.get("structural_collision_guids") != expected_collisions:
            errors.append(f"stale_phrasal_structural_collisions:{route_id}")
        immutable = {field: row.get(field) for field in IMMUTABLE_COLUMNS if field != "input_fingerprint"}
        if row.get("input_fingerprint") != _sha256(immutable):
            errors.append(f"phrasal_input_fingerprint_mismatch:{route_id}")
        disposition = row.get("disposition")
        if disposition not in DISPOSITIONS:
            errors.append(f"invalid_phrasal_disposition:{route_id}")
            continue
        if require_complete and disposition in {"pending", "uncertain"}:
            errors.append(f"incomplete_phrasal_route:{route_id}:{disposition}")
        resolved = disposition not in {"pending", "uncertain"}
        if resolved:
            if row.get("approval") != "approved" or not _text(row.get("reviewer")) or not _text(row.get("reason")):
                errors.append(f"unapproved_phrasal_route:{route_id}")
            try:
                date.fromisoformat(_text(row.get("reviewed_at")))
            except ValueError:
                errors.append(f"invalid_phrasal_review_date:{route_id}")
        target_guid = _text(row.get("target_guid"))
        owned = row.get("owned_semantic_sense_ids")
        if not isinstance(owned, list) or any(not isinstance(value, str) or not value for value in owned):
            errors.append(f"invalid_owned_semantic_senses:{route_id}")
            owned = []
        if disposition == "same_owned_sense" and not owned:
            errors.append(f"same_owned_sense_without_semantic_sense:{route_id}")
        if disposition == "existing_phrase_card":
            if target_guid not in active or target_guid not in expected_collisions:
                errors.append(f"invalid_existing_phrase_card_target:{route_id}")
        elif disposition == "distinct_secondary" and not target_guid:
            errors.append(f"distinct_secondary_without_target_guid:{route_id}")
        elif disposition in {"same_owned_sense", "core_pattern", "excluded_not_relevant"} and target_guid:
            errors.append(f"unexpected_phrasal_target_guid:{route_id}")
        if disposition == "distinct_secondary" and row.get("approval") == "approved":
            for chip in collisions_by_parent.get(parent_guid, []):
                if phrase_starts_learner_surface(row.get("phrase"), chip):
                    errors.append(f"distinct_route_retained_parent_collocation:{route_id}:{chip}")
    return errors


def validate_current_audit(
    rows: Sequence[dict], registry_rows: Sequence[dict], oxford_rows: Sequence[dict],
    *, collocation_audit_rows: Sequence[dict] = (), require_complete: bool = False,
) -> list[str]:
    errors = validate_audit_rows(rows, registry_rows, collocation_audit_rows=collocation_audit_rows,
                                 require_complete=require_complete)
    try:
        fresh = build_audit_rows(
            registry_rows,
            oxford_rows,
            collocation_audit_rows=collocation_audit_rows,
        )
    except ValueError as exc:
        return [*errors, f"invalid_current_phrasal_sources:{exc}"]
    projected = build_audit_rows(
        registry_rows,
        oxford_rows,
        existing_rows=rows,
        collocation_audit_rows=collocation_audit_rows,
    )
    if serialize_rows(rows) != serialize_rows(projected):
        errors.append("stale_phrasal_verb_audit_projection")
    if {row["route_id"] for row in rows} != {row["route_id"] for row in fresh}:
        errors.append("stale_phrasal_verb_route_coverage")
    return errors


def export_workbook(rows: Sequence[dict], path: Path) -> None:
    errors = validate_audit_rows(rows, []) if not rows else []
    if errors:
        raise ValueError("invalid Phrasal Verb Routing Audit: " + "; ".join(errors[:5]))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Routes"
    sheet.append(WORKBOOK_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([
            json.dumps(row[column], ensure_ascii=False) if isinstance(row[column], list) else row[column]
            for column in WORKBOOK_COLUMNS
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.protection.sheet = True
    sheet.protection.password = "phrasal-routing"
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row):
            cell.protection = Protection(locked=index < len(IMMUTABLE_COLUMNS))
    disposition = DataValidation(type="list", formula1='"' + ",".join(DISPOSITIONS) + '"')
    approval = DataValidation(type="list", formula1='"' + ",".join(APPROVALS) + '"')
    sheet.add_data_validation(disposition)
    sheet.add_data_validation(approval)
    disposition.add(f"L2:L{max(2, sheet.max_row)}")
    approval.add(f"R2:R{max(2, sheet.max_row)}")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def import_workbook(rows: Sequence[dict], path: Path) -> list[dict]:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["Routes"]:
        raise ValueError("Phrasal routing workbook must contain only Routes")
    sheet = workbook["Routes"]
    headers = tuple(cell.value for cell in sheet[1])
    if headers != WORKBOOK_COLUMNS:
        raise ValueError("Phrasal routing workbook columns changed")
    by_id = {row["route_id"]: row for row in rows}
    result = copy.deepcopy(by_id)
    seen = set()
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        item = dict(zip(WORKBOOK_COLUMNS, values))
        route_id = _text(item["route_id"])
        if route_id not in by_id or route_id in seen:
            raise ValueError(f"unknown or duplicate workbook route: {route_id}")
        seen.add(route_id)
        for field in IMMUTABLE_COLUMNS:
            value = item[field]
            if field in {"target_source_files", "target_source_sense_ids", "structural_collision_guids"}:
                value = json.loads(value or "[]")
            if value != by_id[route_id][field]:
                raise ValueError(f"immutable workbook value changed: {route_id}:{field}")
        updated = result[route_id]
        for field in EDITABLE_COLUMNS:
            value = item[field] if item[field] is not None else ""
            if field == "owned_semantic_sense_ids":
                value = json.loads(value or "[]")
            updated[field] = value
    if seen != set(by_id):
        raise ValueError("workbook route coverage changed")
    return sorted(result.values(), key=lambda row: row["route_id"])
