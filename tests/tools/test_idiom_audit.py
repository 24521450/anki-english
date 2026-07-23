import copy
import hashlib
import json

import pytest
from openpyxl import load_workbook

from src.config import ProjectPaths
from src.deck_builder.idiom_audit import (
    EDITABLE_COLUMNS,
    REVIEW_COLUMNS,
    apply_review_bundle,
    build_audit_rows,
    export_workbook,
    idiom_source_fingerprint,
    import_workbook,
    load_jsonl,
    parse_serialized_idioms,
    promoted_idioms_by_guid,
    serialize_jsonl,
    validate_audit_rows,
)
from tools.idiom_audit import main


def _registry(*guids):
    return [
        {
            "guid": guid,
            "word": f"word-{guid}",
            "cefr": "C1",
            "list": "Oxford_5000",
            "variant": "",
            "pos": "noun",
            "status": "active",
            "deck_override": None,
        }
        for guid in guids
    ]


def _card(guid, idioms, **updates):
    row = {
        "guid": guid,
        "word": f"word-{guid}",
        "cefr": "C1",
        "pos": "noun",
        "idioms": idioms,
        "tags": "Oxford_5000 CEFR::C1",
        "source1": "Oxford",
        "source2": "Oxford",
    }
    row.update(updates)
    return row


def _complete(row, *, mode="vi_equivalent", confidence="high"):
    row.update(
        {
            "display_mode": mode,
            "equivalence_kind": "proverb" if mode == "vi_equivalent" else "none",
            "explanation_en_simple": "" if mode == "vi_equivalent" else "weaken something deeply",
            "explanation_vi": "Không vào hang cọp, sao bắt được cọp con"
            if mode == "vi_equivalent"
            else "làm lung lay tận gốc",
            "decision": "pass",
            "confidence": confidence,
            "review_reason": "Meaning and register were reviewed.",
            "reviewer": "reviewer",
            "reviewed_at": "2026-07-16",
            "approval": "approved" if confidence != "high" else "",
            "translation_provenance": "manual_bilingual_review",
        }
    )
    return row


def _write_jsonl(path, rows):
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def test_parser_and_source_fingerprint_preserve_exact_trimmed_source_text():
    value = (
        " shake/rock the foundations of something | shake/rock something to its foundations "
        " :: seriously weaken something at its core :: The scandal rocked the institution. "
    )

    assert parse_serialized_idioms(value) == [
        {
            "phrase_en": "shake/rock the foundations of something | shake/rock something to its foundations",
            "source_explanation_en": "seriously weaken something at its core",
            "examples": ["The scandal rocked the institution."],
        }
    ]
    payload = {
        "examples": ["The scandal rocked the institution."],
        "phrase_en": "shake/rock the foundations of something | shake/rock something to its foundations",
        "source_explanation_en": "seriously weaken something at its core",
    }
    expected = hashlib.sha256(
        json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    assert idiom_source_fingerprint(
        f"  {payload['phrase_en']} ",
        f" {payload['source_explanation_en']}  ",
        [f" {payload['examples'][0]} "],
    ) == expected

    with pytest.raises(ValueError, match="requires phrase"):
        parse_serialized_idioms("phrase only")
    with pytest.raises(ValueError, match="empty idiom entry"):
        parse_serialized_idioms("phrase :: meaning$$")


def test_scaffold_groups_normalized_semantic_keys_and_is_deterministic():
    cards = [
        _card(
            "g2",
            "  Nothing   Ventured, Nothing Gained :: You must take risks to succeed  ",
        ),
        _card(
            "g1",
            "nothing ventured, nothing gained :: you must take risks to succeed :: Try it and see.",
        ),
    ]
    registry = _registry("g1", "g2")

    rows = build_audit_rows(cards, registry)
    reversed_rows = build_audit_rows(list(reversed(cards)), list(reversed(registry)))

    assert len(rows) == 1
    assert rows == reversed_rows
    assert rows[0]["idiom_id"].startswith("idm_")
    assert [item["guid"] for item in rows[0]["occurrences"]] == ["g1", "g2"]
    assert rows[0]["source_examples"] == ["Try it and see."]
    assert serialize_jsonl(rows) == serialize_jsonl(reversed_rows)
    assert validate_audit_rows(rows, registry) == []


def test_validation_fails_closed_for_fingerprints_modes_and_approval():
    registry = _registry("g1")
    rows = build_audit_rows(
        [_card("g1", "nothing ventured, nothing gained :: take risks to succeed")],
        registry,
    )

    pending_errors = validate_audit_rows(rows, registry, require_complete=True)
    assert any(error.startswith("incomplete_decision:") for error in pending_errors)
    assert not any(error.startswith("approval_required:") for error in pending_errors)

    complete = copy.deepcopy(rows)
    _complete(complete[0], confidence="medium")
    complete[0]["approval"] = ""
    assert any(
        error.startswith("approval_required:")
        for error in validate_audit_rows(complete, registry, require_complete=True)
    )
    complete[0]["approval"] = "approved"
    assert validate_audit_rows(complete, registry, require_complete=True) == []

    stale = copy.deepcopy(complete)
    stale[0]["occurrences"][0]["example"] = "Changed example."
    errors = validate_audit_rows(stale, registry, require_complete=True)
    assert any(error.startswith("source_fingerprint_mismatch:") for error in errors)
    assert any(error.startswith("coverage_fingerprint_mismatch:") for error in errors)

    malformed = copy.deepcopy(complete)
    malformed[0]["explanation_vi"] = "bad::delimiter"
    assert any(
        error.endswith(":explanation_vi")
        for error in validate_audit_rows(malformed, registry, require_complete=True)
    )


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("explanation_en_simple", "someone?s condition"),
        ("explanation_vi", "ho?n to?n kh?e m?nh/nguy?n v?n"),
        ("review_reason", "?Ph?n meaning was retained."),
        ("translation_provenance", "manual\ufffdreview"),
    ],
)
def test_validation_rejects_suspected_lossy_unicode_in_review_text(field, value):
    rows = build_audit_rows(
        [_card("g1", "sound as a bell :: completely healthy or undamaged")],
        _registry("g1"),
    )
    _complete(rows[0])
    rows[0][field] = value

    assert (
        f"suspected_lossy_unicode:{rows[0]['idiom_id']}:{field}"
        in validate_audit_rows(rows, _registry("g1"), require_complete=True)
    )


def test_validation_allows_unicode_and_terminal_question_punctuation():
    rows = build_audit_rows(
        [_card("g1", "sound as a bell :: completely healthy or undamaged")],
        _registry("g1"),
    )
    _complete(rows[0])
    rows[0]["explanation_vi"] = "hoàn toàn khỏe mạnh/nguyên vẹn"
    rows[0]["review_reason"] = "Ai trả giá nào?"
    rows[0]["translation_provenance"] = "rà_soát_thủ_công"

    assert validate_audit_rows(rows, _registry("g1"), require_complete=True) == []
    assert promoted_idioms_by_guid(rows)["g1"][0]["explanation_vi"] == (
        "hoàn toàn khỏe mạnh/nguyên vẹn"
    )


def test_promotion_rejects_suspected_lossy_unicode():
    rows = build_audit_rows(
        [_card("g1", "sound as a bell :: completely healthy or undamaged")],
        _registry("g1"),
    )
    _complete(rows[0])
    rows[0]["explanation_vi"] = "ho?n to?n kh?e m?nh/nguy?n v?n"

    with pytest.raises(
        ValueError,
        match=rf"suspected_lossy_unicode:{rows[0]['idiom_id']}:explanation_vi",
    ):
        promoted_idioms_by_guid(rows)


def test_promotion_uses_source_fallback_for_vi_equivalent_and_simple_english_for_gloss():
    registry = _registry("g1", "g2")
    rows = build_audit_rows(
        [
            _card(
                "g1",
                "nothing ventured, nothing gained :: used to say success requires risk",
            ),
            _card(
                "g2",
                "shake something to its foundations :: damage something at its core :: The scandal shook it.",
            ),
        ],
        registry,
    )
    by_phrase = {row["phrase_en"]: row for row in rows}
    _complete(by_phrase["nothing ventured, nothing gained"])
    _complete(by_phrase["shake something to its foundations"], mode="bilingual_gloss")

    promoted = promoted_idioms_by_guid(rows)

    assert promoted["g1"][0]["explanation_en"] == "used to say success requires risk"
    assert promoted["g1"][0]["explanation_vi"] == "Không vào hang cọp, sao bắt được cọp con"
    assert promoted["g2"][0]["explanation_en"] == "weaken something deeply"
    assert promoted["g2"][0]["examples"] == ["The scandal shook it."]
    assert set(promoted["g2"][0]) == {
        "idiom_id",
        "order",
        "source_fingerprint",
        "phrase_en",
        "display_mode",
        "explanation_en",
        "explanation_vi",
        "examples",
        "translation_provenance",
    }


def test_workbook_round_trip_updates_only_editable_columns(tmp_path):
    rows = build_audit_rows(
        [_card("g1", "nothing ventured, nothing gained :: take risks to succeed")],
        _registry("g1"),
    )
    path = tmp_path / "idioms.xlsx"
    export_workbook(rows, path)
    workbook = load_workbook(path)
    instructions = "\n".join(
        str(row[0].value or "") for row in workbook["Instructions"].iter_rows()
    )
    assert "exact imagery and pragmatic scope need not match" in instructions
    assert "get back on the rails -> đâu lại vào đấy" in instructions
    assert "be at odds -> trống đánh xuôi, kèn thổi ngược" in instructions
    assert "short, natural learner glosses in both English and Vietnamese" in instructions
    assert "do not mirror English sentence structure in Vietnamese" in instructions
    assert "There is no hard word limit" in instructions
    assert "Do not repeat somebody/something" in instructions
    assert "persuade/pressure / thuyết phục/nài ép" in instructions
    assert "Unchanged is not reviewed by default" in instructions
    assert "never bulk-pass unchanged rows" in instructions
    assert "put somebody to the sword -> kill / giết" in instructions
    assert "an old belief that is not true / quan niệm dân gian sai lầm" in instructions
    sheet = workbook["Review"]
    for field, value in {
        "display_mode": "vi_equivalent",
        "equivalence_kind": "proverb",
        "explanation_vi": "Không vào hang cọp, sao bắt được cọp con",
        "decision": "pass",
        "confidence": "high",
        "review_reason": "Reviewed equivalent.",
        "reviewer": "reviewer",
        "reviewed_at": "2026-07-16",
        "translation_provenance": "manual_bilingual_review",
    }.items():
        sheet.cell(2, REVIEW_COLUMNS.index(field) + 1).value = value
    workbook.save(path)

    imported = import_workbook(rows, path)

    assert imported[0]["decision"] == "pass"
    assert imported[0]["content_fingerprint"] == rows[0]["content_fingerprint"]
    assert rows[0]["decision"] == "pending"

    workbook = load_workbook(path)
    workbook["Review"].cell(2, REVIEW_COLUMNS.index("phrase_en") + 1).value = "tampered"
    workbook.save(path)
    with pytest.raises(ValueError, match="Immutable column 'phrase_en'"):
        import_workbook(rows, path)


def test_review_bundle_is_fingerprint_bound_and_transactional():
    rows = build_audit_rows(
        [
            _card("g1", "first phrase :: first meaning"),
            _card("g2", "second phrase :: second meaning"),
        ],
        _registry("g1", "g2"),
    )
    first, second = rows
    valid_update = {
        "idiom_id": first["idiom_id"],
        "content_fingerprint": first["content_fingerprint"],
        "coverage_fingerprint": first["coverage_fingerprint"],
        "display_mode": "bilingual_gloss",
        "equivalence_kind": "none",
        "explanation_en_simple": "simple meaning",
        "explanation_vi": "nghĩa đơn giản",
    }
    stale_update = {
        "idiom_id": second["idiom_id"],
        "content_fingerprint": "0" * 64,
        "coverage_fingerprint": second["coverage_fingerprint"],
        "explanation_vi": "không được áp dụng",
    }

    with pytest.raises(ValueError, match="Stale review bundle fingerprint"):
        apply_review_bundle(rows, [valid_update, stale_update])
    assert all(row["display_mode"] == "" for row in rows)

    updated = apply_review_bundle(rows, [valid_update])
    assert updated[0]["display_mode"] == "bilingual_gloss"
    assert rows[0]["display_mode"] == ""
    assert set(EDITABLE_COLUMNS).issuperset(set(valid_update) - {
        "idiom_id", "content_fingerprint", "coverage_fingerprint"
    })


def test_cli_scaffold_and_apply_review_dry_run_do_not_mutate_canonical_jsonl(tmp_path):
    notes = tmp_path / "notes.jsonl"
    registry_path = tmp_path / "registry.jsonl"
    audit = tmp_path / "audit.jsonl"
    decisions = tmp_path / "decisions.jsonl"
    _write_jsonl(notes, [_card("g1", "first phrase :: first meaning")])
    registry = _registry("g1")
    _write_jsonl(registry_path, registry)

    assert main([
        "--audit", str(audit),
        "--registry", str(registry_path),
        "scaffold",
        "--notes", str(notes),
    ]) == 0
    before = audit.read_bytes()
    row = json.loads(before.decode("utf-8"))
    _write_jsonl(decisions, [{
        "idiom_id": row["idiom_id"],
        "content_fingerprint": row["content_fingerprint"],
        "coverage_fingerprint": row["coverage_fingerprint"],
        "display_mode": "bilingual_gloss",
        "equivalence_kind": "none",
        "explanation_en_simple": "simple meaning",
        "explanation_vi": "nghĩa đơn giản",
    }])

    assert main([
        "--audit", str(audit),
        "--registry", str(registry_path),
        "apply-review",
        "--input", str(decisions),
        "--dry-run",
    ]) == 0
    assert audit.read_bytes() == before


def test_cli_validate_fails_when_the_canonical_ledger_is_missing(tmp_path, capsys):
    registry = tmp_path / "registry.jsonl"
    _write_jsonl(registry, _registry("g1"))

    assert main([
        "--audit", str(tmp_path / "missing.jsonl"),
        "--registry", str(registry),
        "validate",
    ]) == 1

    assert "No such file or directory" in capsys.readouterr().err


def test_cli_report_includes_exceptions_and_deterministic_sample(tmp_path):
    registry_path = tmp_path / "registry.jsonl"
    audit = tmp_path / "audit.jsonl"
    report = tmp_path / "report.md"
    registry = _registry("g1", "g2")
    rows = build_audit_rows([
        _card("g1", "first phrase :: first meaning"),
        _card("g2", "second phrase :: second meaning"),
    ], registry)
    _complete(rows[0])
    _complete(rows[1], mode="bilingual_gloss", confidence="medium")
    rows[1].update({
        "decision": "uncertain",
        "approval": "",
        "review_reason": "The candidate is only a literal paraphrase, not an established saying.",
    })
    _write_jsonl(registry_path, registry)
    audit.write_text(serialize_jsonl(rows), encoding="utf-8")

    assert main([
        "--audit", str(audit),
        "--registry", str(registry_path),
        "report",
        "--output", str(report),
        "--sample-size", "1",
    ]) == 0

    text = report.read_text(encoding="utf-8")
    assert "## Review exceptions" in text
    assert "The candidate is only a literal paraphrase, not an established saying." in text
    assert "## Deterministic high-confidence sample (1)" in text


def test_canonical_ledger_keeps_user_approved_idiom_mappings():
    rows = {
        row["phrase_en"]: row
        for row in load_jsonl(ProjectPaths().bilingual_idiom_audit)
    }
    expected = {
        "get back on the rails": "đâu lại vào đấy",
        "be at odds (with somebody) (over/on something)":
            "trống đánh xuôi, kèn thổi ngược",
    }

    for phrase, vietnamese in expected.items():
        row = rows[phrase]
        assert row["display_mode"] == "vi_equivalent"
        assert row["equivalence_kind"] == "idiom"
        assert row["explanation_en_simple"] == ""
        assert row["explanation_vi"] == vietnamese
        assert row["decision"] == "pass"
        assert row["confidence"] == "high"

    foundations = rows[
        "shake/rock the foundations of something | "
        "shake/rock something to its foundations"
    ]
    assert foundations["display_mode"] == "bilingual_gloss"
    assert foundations["equivalence_kind"] == "none"
    assert foundations["explanation_en_simple"] == (
        "seriously weaken something at its core"
    )
    assert foundations["explanation_vi"] == "làm lung lay tận gốc"

    old_wives_tale = rows["an old wives’ tale"]
    assert old_wives_tale["display_mode"] == "bilingual_gloss"
    assert old_wives_tale["equivalence_kind"] == "none"
    assert old_wives_tale["explanation_en_simple"] == (
        "an old belief that is not true"
    )
    assert old_wives_tale["explanation_vi"] == "quan niệm dân gian sai lầm"


def test_canonical_bilingual_gloss_quality_regressions():
    rows = {
        row["phrase_en"]: row
        for row in load_jsonl(ProjectPaths().bilingual_idiom_audit)
    }
    expected = {
        "sign/take the pledge": (
            "promise never to drink alcohol",
            "cam kết không bao giờ uống rượu",
        ),
        "put an animal, a bird, etc. out of its misery": (
            "kill a suffering animal humanely when it cannot recover",
            "giết một con vật đang đau đớn theo cách nhân đạo khi không thể chữa khỏi",
        ),
        "the naked eye": (
            "normal sight without an instrument",
            "mắt thường, không dùng dụng cụ hỗ trợ",
        ),
        "kindly adjust": (
            "please excuse the inconvenience and adapt as needed",
            "xin thông cảm và chủ động điều chỉnh cho phù hợp",
        ),
        "without prejudice (to something)": (
            "without affecting other legal rights / issues",
            "không ảnh hưởng đến quyền / vấn đề pháp lý khác",
        ),
        "in their/our/its/your midst": (
            "among a group",
            "ở giữa một nhóm người",
        ),
    }

    for phrase, (english, vietnamese) in expected.items():
        row = rows[phrase]
        assert row["display_mode"] == "bilingual_gloss"
        assert row["equivalence_kind"] == "none"
        assert row["explanation_en_simple"] == english
        assert row["explanation_vi"] == vietnamese
        assert row["decision"] == "pass"
        assert row["confidence"] == "high"

    assert "off the top of your head" not in rows


def test_canonical_bilingual_gloss_concision_rereview_anchors():
    rows = {
        row["phrase_en"]: row
        for row in load_jsonl(ProjectPaths().bilingual_idiom_audit)
    }
    expected = {
        "put somebody to the sword": ("kill", "giết"),
        "twist somebody’s arm": (
            "persuade/pressure",
            "thuyết phục/nài ép",
        ),
        "out of joint": ("dislocated", "trật khớp"),
        "in default of something": ("for lack of", "vì thiếu"),
    }

    for phrase, (english, vietnamese) in expected.items():
        row = rows[phrase]
        assert row["display_mode"] == "bilingual_gloss"
        assert row["equivalence_kind"] == "none"
        assert row["explanation_en_simple"] == english
        assert row["explanation_vi"] == vietnamese

    dawn = rows["at the crack of dawn"]
    assert dawn["display_mode"] == "vi_equivalent"
    assert dawn["equivalence_kind"] == "idiom"
    assert dawn["explanation_en_simple"] == ""
    assert dawn["explanation_vi"] == "sáng tinh mơ"
