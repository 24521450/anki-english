import copy

import pytest
from openpyxl import load_workbook

from src.deck_builder.phrasal_verb_audit import (
    build_audit_rows,
    export_workbook,
    import_workbook,
    phrase_starts_learner_surface,
    structural_phrase_collisions,
    validate_audit_rows,
)


def _registry():
    return [
        {"guid": "g-parent", "word": "derive", "status": "active"},
        {"guid": "g-phrase", "word": "derive from sth", "status": "active"},
        {"guid": "g-substring", "word": "rederive from", "status": "active"},
    ]


def _oxford():
    definition = {"text": "to come from something", "examples": []}
    target_file = "oxford_derive-from_(phrasal_verb).html"
    return [
        {
            "word": "derive", "source_files": ["oxford_derive.html"],
            "phrasal_verb_links": [{
                "phrase": "derive from",
                "url": "https://www.oxfordlearnersdictionaries.com/definition/english/derive-from",
            }],
            "pos_data": [{"source_url": "https://example/derive", "definitions": []}],
            "_skip": False,
        },
        {
            "word": "derive from", "source_files": [target_file],
            "pos_data": [{
                "source_url": "https://www.oxfordlearnersdictionaries.com/definition/english/derive-from",
                "definitions": [definition],
            }],
            "_skip": False,
        },
    ]


def _approve(row, disposition="existing_phrase_card"):
    row.update({
        "disposition": disposition,
        "target_guid": "g-phrase" if disposition == "existing_phrase_card" else "",
        "reason": "Route this exact hydrated Oxford phrase deliberately.",
        "reviewer": "reviewer", "reviewed_at": "2026-07-23", "approval": "approved",
    })


def test_structural_match_is_start_anchored_and_expands_only_explicit_slash():
    assert phrase_starts_learner_surface("derive from", "derive from/out sth")
    assert not phrase_starts_learner_surface("derive from", "rederive from")
    assert not phrase_starts_learner_surface("derive from", "derive funding")
    assert structural_phrase_collisions("derive from", _registry()) == ["g-phrase"]


def test_scaffold_requires_authoritative_hydrated_target_and_reuses_bound_review():
    rows = build_audit_rows(_registry(), _oxford())
    assert len(rows) == 1
    assert rows[0]["target_source_sense_ids"]
    assert rows[0]["structural_collision_guids"] == ["g-phrase"]
    _approve(rows[0])
    assert build_audit_rows(_registry(), _oxford(), existing_rows=rows)[0]["approval"] == "approved"

    stale = copy.deepcopy(_oxford())
    stale[0]["phrasal_verb_links"][0]["url"] = (
        "https://www.oxfordlearnersdictionaries.com/definition/english/wrong-target"
    )
    with pytest.raises(ValueError, match="authoritative Oxford target"):
        build_audit_rows(_registry(), stale)


def test_complete_validation_enforces_disposition_payload_and_distinct_route_chip_gate():
    rows = build_audit_rows(_registry(), _oxford())
    assert any(error.startswith("incomplete_phrasal_route:") for error in
               validate_audit_rows(rows, _registry(), require_complete=True))

    _approve(rows[0], "distinct_secondary")
    rows[0]["target_guid"] = "new-guid"
    collocations = [{"guid": "g-parent", "final_items": [{"text": "derive from sth"}]}]
    errors = validate_audit_rows(rows, _registry(), collocation_audit_rows=collocations,
                                 require_complete=True)
    assert any(error.startswith("distinct_route_retained_parent_collocation:") for error in errors)


def test_workbook_locks_immutable_columns_and_rejects_tampering(tmp_path):
    rows = build_audit_rows(_registry(), _oxford())
    path = tmp_path / "routes.xlsx"
    export_workbook(rows, path)
    workbook = load_workbook(path)
    sheet = workbook["Routes"]
    assert sheet.protection.sheet
    assert sheet["A2"].protection.locked
    assert not sheet["L2"].protection.locked
    sheet["L2"] = "core_pattern"
    sheet["N2"] = "[\"sem_1\"]"
    sheet["O2"] = "This is the reviewed core learning pattern."
    sheet["P2"] = "reviewer"
    sheet["Q2"] = "2026-07-23"
    sheet["R2"] = "approved"
    workbook.save(path)
    assert import_workbook(rows, path)[0]["disposition"] == "core_pattern"

    workbook = load_workbook(path)
    workbook["Routes"]["E2"] = "tampered phrase"
    workbook.save(path)
    with pytest.raises(ValueError, match="immutable workbook value changed"):
        import_workbook(rows, path)
