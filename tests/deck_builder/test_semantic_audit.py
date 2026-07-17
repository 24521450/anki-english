from copy import deepcopy

import pytest
from openpyxl import load_workbook

from src.deck_builder.semantic_audit import (
    COVERAGE_COLUMNS,
    REVIEW_COLUMNS,
    apply_review_bundle,
    build_audit_rows,
    export_workbook,
    flatten_review_rows,
    import_workbook,
    split_definition_chunk,
    validate_audit_rows,
)


def _registry(**overrides):
    row = {
        "word": "equate", "cefr": "C1", "list": "Oxford_5000", "variant": "",
        "pos": "verb", "guid": "guid-1", "status": "active", "deck_override": None,
    }
    row.update(overrides)
    return row


def _card(**overrides):
    row = {
        "guid": "guid-1", "word": "equate", "pos": "verb", "cefr": "C1",
        "definition": "consider equal (đánh đồng)", "example": "You should not equate wealth with happiness.",
        "idioms": "", "tags": "Oxford_5000 CEFR::C1",
    }
    row.update(overrides)
    return row


def _oxford():
    return [{
        "word": "equate", "homonym_index": None, "source_files": ["oxford_equate_(verb).html"],
        "oxford_badge": "C1",
        "pos_data": [{
            "pos": "verb",
            "definitions": [{
                "sensenum_local": None, "text": "to think that something is the same as something else",
                "cefr": "C1", "examples": [{"text": "You should not equate wealth with happiness."}],
                "register_tags": [], "domain": None,
            }],
        }],
    }]


def _audit():
    return build_audit_rows([_card()], [_registry()], _oxford())


def test_split_definition_keeps_natural_vietnamese():
    assert split_definition_chunk("consider equal (đánh đồng)") == ("consider equal", "đánh đồng")
    assert split_definition_chunk(
        "death (the grave) (cái chết (the grave))"
    ) == ("death (the grave)", "cái chết (the grave)")
    assert split_definition_chunk("slope") == ("slope", "")


def test_scaffold_prefers_separate_definition_vi_field():
    rows = build_audit_rows([
        _card(definition_vi="đánh đồng chính xác")
    ], [_registry()], _oxford())

    assert rows[0]["semantic_senses"][0]["current"] == {
        "definition_en": "consider equal",
        "definition_vi": "đánh đồng chính xác",
        "examples": ["You should not equate wealth with happiness."],
    }


def test_scaffold_is_pending_and_does_not_guess_source_mapping():
    rows = _audit()
    sense = rows[0]["semantic_senses"][0]
    assert sense["decision"] == "pending"
    assert sense["source_sense_ids"] == []
    assert len(rows[0]["coverage"]["candidate_source_sense_ids"]) == 1
    assert validate_audit_rows(rows, [_registry()]) == []


def test_source_sense_ids_do_not_depend_on_sensenum_presence():
    first = _audit()[0]["source_senses"][0]["source_sense_id"]
    source = _oxford()
    source[0]["pos_data"][0]["definitions"][0]["sensenum_local"] = "1"
    second = build_audit_rows([_card()], [_registry()], source)[0]["source_senses"][0]["source_sense_id"]
    assert first != second


def test_complete_gate_requires_mapping_review_and_approval():
    rows = _audit()
    sense = rows[0]["semantic_senses"][0]
    source_id = rows[0]["coverage"]["candidate_source_sense_ids"][0]
    sense["source_sense_ids"] = [source_id]
    rows[0]["source_coverage"][0].update({
        "disposition": "mapped", "target_semantic_sense_ids": [sense["semantic_sense_id"]],
    })
    sense["checks"] = {key: "pass" for key in sense["checks"]}
    sense["decision"] = "pass"
    sense["reviewer"] = "chatgpt-5.6-sol-ultra"
    sense["cambridge"]["match"] = "exact"
    sense["cambridge"]["translation_provenance"] = "cambridge_reference"
    rows[0]["coverage"]["status"] = "pass"
    assert validate_audit_rows(rows, [_registry()], require_complete=True) == []


def test_validator_rejects_invalid_review_shape_and_corrupt_vietnamese():
    rows = _audit()
    sense = rows[0]["semantic_senses"][0]
    sense["checks"] = {
        "english_semantics": "pass",
        "vietnamese_semantics": "pass",
        "simplicity": "pass",
    }
    sense["decision"] = "repair_proposed"
    sense["proposed"]["definition_vi"] = "ngh?a b? l?i"
    sense["cambridge"]["match"] = "related"

    errors = validate_audit_rows(rows, [_registry()])
    assert f"invalid_check_set:guid-1:{sense['semantic_sense_id']}" in errors
    assert f"repair_without_repair_check:guid-1:{sense['semantic_sense_id']}" in errors
    assert f"corrupt_vietnamese_text:guid-1:{sense['semantic_sense_id']}:proposed" in errors
    assert f"invalid_cambridge_match:guid-1:{sense['semantic_sense_id']}" in errors


def test_complete_gate_rejects_pending_cambridge_review():
    rows = _audit()
    card = rows[0]
    sense = card["semantic_senses"][0]
    source_id = card["source_coverage"][0]["source_sense_id"]
    card["source_coverage"][0].update({
        "disposition": "mapped", "target_semantic_sense_ids": [sense["semantic_sense_id"]],
    })
    sense["source_sense_ids"] = [source_id]
    sense["checks"] = {key: "pass" for key in sense["checks"]}
    sense["decision"] = "pass"
    sense["reviewer"] = "chatgpt-5.6-sol-ultra"
    card["coverage"]["status"] = "pass"

    errors = validate_audit_rows(rows, [_registry()], require_complete=True)
    assert f"open_cambridge_match:guid-1:{sense['semantic_sense_id']}" in errors
    assert f"missing_translation_provenance:guid-1:{sense['semantic_sense_id']}" in errors


def test_complete_gate_rejects_unaccounted_source_sense():
    rows = _audit()
    rows[0]["coverage"]["status"] = "pass"
    errors = validate_audit_rows(rows, [_registry()], require_complete=True)
    assert any(error.startswith("unaccounted_source_sense:guid-1:") for error in errors)


def test_only_real_idiom_only_card_can_be_not_applicable():
    card = _card(definition="", example="", idioms="in accordance with :: as required :: example")
    rows = build_audit_rows([card], [_registry()], _oxford())
    assert rows[0]["coverage"] == {
        "status": "not_applicable", "reason": "idiom_only",
        "candidate_source_sense_ids": rows[0]["coverage"]["candidate_source_sense_ids"],
        "expected_same_cefr_source_sense_ids": rows[0]["coverage"]["expected_same_cefr_source_sense_ids"],
    }
    assert validate_audit_rows(rows, [_registry()]) == []
    rows[0]["current"]["idioms"] = ""
    assert "invalid_empty_card:guid-1" in validate_audit_rows(rows, [_registry()])


def test_idiom_only_card_rejects_added_standalone_semantic_sense():
    card = _card(definition="", example="", idioms="in accordance with :: as required :: example")
    rows = build_audit_rows([card], [_registry()], _oxford())
    with pytest.raises(ValueError, match="Cannot add semantic senses to idiom-only card"):
        apply_review_bundle(rows, [{
            "guid": "guid-1",
            "add_senses": [{
                "semantic_sense_id": "sem-invalid",
                "order": 1,
                "checks": {"english_semantics": "repair"},
                "proposed": {"definition_en": "invalid standalone meaning"},
            }],
        }])


def test_same_source_sense_cannot_be_mapped_to_multiple_cards():
    first = _audit()[0]
    source_id = first["source_coverage"][0]["source_sense_id"]
    first_semantic_id = first["semantic_senses"][0]["semantic_sense_id"]
    first["source_coverage"][0].update({
        "disposition": "mapped", "target_semantic_sense_ids": [first_semantic_id],
    })
    first["semantic_senses"][0]["source_sense_ids"] = [source_id]

    second = deepcopy(first)
    second["guid"] = "guid-2"
    second_semantic_id = "sem-guid-2"
    second["semantic_senses"][0]["semantic_sense_id"] = second_semantic_id
    second["source_coverage"][0]["target_semantic_sense_ids"] = [second_semantic_id]
    registry = [_registry(), _registry(guid="guid-2")]

    errors = validate_audit_rows([first, second], registry)
    assert f"source_mapped_to_multiple_cards:{source_id}:guid-1:guid-2" in errors


def test_duplicate_source_wording_needs_explicit_reason_for_mixed_disposition():
    rows = _audit()
    card = rows[0]
    first_source = card["source_senses"][0]
    first_id = first_source["source_sense_id"]
    duplicate_id = "ox-duplicate"
    duplicate = deepcopy(first_source)
    duplicate["source_sense_id"] = duplicate_id
    duplicate["pos"] = "adjective"
    duplicate["cefr_resolved"] = "C2"
    card["source_senses"].append(duplicate)
    card["coverage"]["candidate_source_sense_ids"].append(duplicate_id)
    semantic_id = card["semantic_senses"][0]["semantic_sense_id"]
    card["source_coverage"][0].update({
        "disposition": "mapped", "target_semantic_sense_ids": [semantic_id],
    })
    card["semantic_senses"][0]["source_sense_ids"] = [first_id]
    card["source_coverage"].append({
        "source_sense_id": duplicate_id,
        "disposition": "excluded",
        "target_semantic_sense_ids": [],
        "reason": "Distinct source sense outside this card.",
    })

    errors = validate_audit_rows(rows, [_registry()])
    assert any(error.startswith("contradictory_duplicate_source_disposition:guid-1:") for error in errors)

    card["source_coverage"][-1]["reason"] = "Duplicate wording belongs to the wrong POS."
    assert not any(
        error.startswith("contradictory_duplicate_source_disposition:guid-1:")
        for error in validate_audit_rows(rows, [_registry()])
    )


def test_xlsx_round_trip_preserves_unicode_and_literal_formula(tmp_path):
    rows = _audit()
    path = tmp_path / "audit.xlsx"
    export_workbook(rows, path)
    workbook = load_workbook(path)
    sheet = workbook["Review"]
    instructions = "\n".join(
        str(cell.value or "") for row in workbook["Instructions"] for cell in row
    )
    assert "independent Lexical Glosses" in instructions
    assert "generic claim that it 'preserves nuance' is not sufficient" in instructions
    assert "changing only punctuation or word order" in instructions
    headers = [cell.value for cell in sheet[1]]
    values = {name: headers.index(name) + 1 for name in headers}
    sheet.cell(2, values["cambridge_match"]).value = "exact"
    sheet.cell(2, values["english_check"]).value = "pass"
    sheet.cell(2, values["vietnamese_check"]).value = "repair"
    sheet.cell(2, values["simplicity_check"]).value = "pass"
    sheet.cell(2, values["example_pos_check"]).value = "pass"
    sheet.cell(2, values["decision"]).value = "repair_proposed"
    sheet.cell(2, values["proposed_vi"]).value = "=đánh đồng"
    sheet.cell(2, values["reviewer"]).value = "chatgpt-5.6-sol-ultra"
    coverage_sheet = workbook["Source Coverage"]
    coverage_headers = [cell.value for cell in coverage_sheet[1]]
    coverage_values = {name: coverage_headers.index(name) + 1 for name in coverage_headers}
    coverage_sheet.cell(2, coverage_values["disposition"]).value = "mapped"
    coverage_sheet.cell(2, coverage_values["target_semantic_sense_ids"]).value = rows[0]["semantic_senses"][0]["semantic_sense_id"]
    workbook.save(path)

    updated = import_workbook(deepcopy(rows), path)
    sense = updated[0]["semantic_senses"][0]
    assert sense["proposed"]["definition_vi"] == "=đánh đồng"
    assert sense["source_sense_ids"] == rows[0]["coverage"]["candidate_source_sense_ids"]
    assert updated[0]["coverage"]["status"] == "repair_proposed"


def test_xlsx_import_rejects_immutable_change(tmp_path):
    rows = _audit()
    path = tmp_path / "audit.xlsx"
    export_workbook(rows, path)
    workbook = load_workbook(path)
    sheet = workbook["Review"]
    word_column = list(REVIEW_COLUMNS).index("word") + 1
    sheet.cell(2, word_column).value = "wrong"
    workbook.save(path)
    with pytest.raises(ValueError, match="Immutable column"):
        import_workbook(rows, path)


def test_flatten_uses_candidates_without_claiming_mapping():
    row = flatten_review_rows(_audit())[0]
    assert row["candidate_source_sense_ids"].startswith("ox_")
    assert "same as" in row["candidate_source_definitions"]


def test_apply_review_bundle_records_explicit_mapping_and_verdict():
    rows = _audit()
    card = rows[0]
    sense = card["semantic_senses"][0]
    source_id = card["coverage"]["candidate_source_sense_ids"][0]
    apply_review_bundle(rows, [{
        "guid": card["guid"],
        "source_coverage": [{
            "source_sense_id": source_id,
            "disposition": "mapped",
            "target_semantic_sense_ids": [sense["semantic_sense_id"]],
            "reason": "same Oxford sense",
        }],
        "senses": [{
            "semantic_sense_id": sense["semantic_sense_id"],
            "checks": {
                "english_semantics": "pass", "vietnamese_semantics": "pass",
                "simplicity": "pass", "example_pos_alignment": "pass",
            },
            "decision": "pass",
            "cambridge": {"match": "exact", "summary": "đánh đồng", "translation_provenance": "cambridge_reference"},
            "confidence": "high", "review_reason": "Meaning and example align.",
        }],
    }])
    assert sense["source_sense_ids"] == [source_id]
    assert card["coverage"]["status"] == "pass"
    assert sense["reviewer"] == "chatgpt-5.6-sol-ultra"


def test_apply_review_bundle_can_add_a_missing_semantic_sense():
    rows = _audit()
    card = rows[0]
    source_id = card["coverage"]["candidate_source_sense_ids"][0]
    new_id = "sem_reviewer_added"

    apply_review_bundle(rows, [{
        "guid": card["guid"],
        "add_senses": [{
            "semantic_sense_id": new_id,
            "order": 2,
            "checks": {
                "english_semantics": "repair", "vietnamese_semantics": "repair",
                "simplicity": "pass", "example_pos_alignment": "repair",
            },
            "decision": "repair_proposed",
            "proposed": {
                "definition_en": "a distinct missing meaning",
                "definition_vi": "một nghĩa còn thiếu",
                "examples": ["A separate example."],
            },
            "cambridge": {
                "match": "exact", "summary": "A separate Cambridge sense.",
                "translation_provenance": "cambridge_reference",
            },
            "confidence": "high",
            "review_reason": "The legacy gloss merged two distinct meanings.",
        }],
        "source_coverage": [{
            "source_sense_id": source_id,
            "disposition": "mapped",
            "target_semantic_sense_ids": [new_id],
            "reason": "Maps to the newly recovered sense.",
        }],
    }])

    assert card["semantic_senses"][1]["semantic_sense_id"] == new_id
    assert card["semantic_senses"][1]["source_sense_ids"] == [source_id]
    assert card["coverage"]["status"] == "pending"
    assert validate_audit_rows(rows, [_registry()]) == []


def test_apply_review_bundle_can_insert_a_missing_sense_before_an_existing_one():
    rows = _audit()
    card = rows[0]
    existing_id = card["semantic_senses"][0]["semantic_sense_id"]

    apply_review_bundle(rows, [{
        "guid": card["guid"],
        "add_senses": [{
            "semantic_sense_id": "sem_inserted", "order": 1,
            "checks": {key: "repair" for key in card["semantic_senses"][0]["checks"]},
            "proposed": {"definition_en": "earlier meaning", "definition_vi": "nghĩa trước", "examples": ["Earlier."]},
        }],
        "senses": [{"semantic_sense_id": existing_id, "order": 2}],
    }])

    assert [sense["semantic_sense_id"] for sense in card["semantic_senses"]] == ["sem_inserted", existing_id]


def _audit_with_specialized_second_sense():
    rows = _audit()
    card = rows[0]
    retained_sense = card["semantic_senses"][0]
    retained_source_id = card["source_senses"][0]["source_sense_id"]
    card["source_coverage"][0].update({
        "disposition": "mapped",
        "target_semantic_sense_ids": [retained_sense["semantic_sense_id"]],
        "reason": "Retained learner-relevant meaning.",
    })
    retained_sense["source_sense_ids"] = [retained_source_id]
    retained_sense["checks"] = {key: "pass" for key in retained_sense["checks"]}
    retained_sense["decision"] = "pass"
    retained_sense["reviewer"] = "test-reviewer"
    retained_sense["cambridge"].update({
        "match": "exact",
        "translation_provenance": "cambridge_reference",
    })
    card["coverage"]["status"] = "pass"
    source = deepcopy(card["source_senses"][0])
    source["source_sense_id"] = "ox_specialized"
    source["definition"] = "a narrowly specialized technical meaning"
    source["domain"] = "computing"
    card["source_senses"].append(source)
    card["coverage"]["candidate_source_sense_ids"].append(source["source_sense_id"])
    card["source_coverage"].append({
        "source_sense_id": source["source_sense_id"],
        "disposition": "mapped",
        "target_semantic_sense_ids": ["sem_specialized"],
        "reason": "Temporarily mapped before relevance review.",
    })
    sense = deepcopy(card["semantic_senses"][0])
    sense.update({
        "semantic_sense_id": "sem_specialized",
        "order": 2,
        "source_sense_ids": [source["source_sense_id"]],
    })
    card["semantic_senses"].append(sense)
    return rows


def test_apply_review_bundle_can_remove_a_specialized_sense_and_compact_order():
    rows = _audit_with_specialized_second_sense()
    card = rows[0]
    retained_id = card["semantic_senses"][0]["semantic_sense_id"]

    apply_review_bundle(rows, [{
        "guid": card["guid"],
        "remove_senses": ["sem_specialized"],
        "source_coverage": [{
            "source_sense_id": "ox_specialized",
            "disposition": "excluded",
            "target_semantic_sense_ids": [],
            "reason": "Overly specialized computing sense outside learner scope.",
        }],
    }])

    assert [sense["semantic_sense_id"] for sense in card["semantic_senses"]] == [retained_id]
    assert [sense["order"] for sense in card["semantic_senses"]] == [1]
    assert card["source_coverage"][-1] == {
        "source_sense_id": "ox_specialized",
        "disposition": "excluded",
        "target_semantic_sense_ids": [],
        "reason": "Overly specialized computing sense outside learner scope.",
    }
    assert validate_audit_rows(rows, [_registry()]) == []
    assert validate_audit_rows(rows, [_registry()], require_complete=True) == []


def test_apply_review_bundle_rejects_removal_while_a_source_still_targets_it():
    rows = _audit_with_specialized_second_sense()

    with pytest.raises(ValueError, match="Source coverage still targets removed semantic sense"):
        apply_review_bundle(rows, [{
            "guid": rows[0]["guid"],
            "remove_senses": ["sem_specialized"],
        }])


def test_apply_review_bundle_rejects_removing_every_semantic_sense():
    rows = _audit()
    semantic_id = rows[0]["semantic_senses"][0]["semantic_sense_id"]

    with pytest.raises(ValueError, match="Cannot remove every semantic sense"):
        apply_review_bundle(rows, [{
            "guid": rows[0]["guid"],
            "remove_senses": [semantic_id],
        }])


def test_apply_review_bundle_rejects_non_list_remove_senses():
    rows = _audit_with_specialized_second_sense()

    with pytest.raises(ValueError, match="remove_senses must be a list"):
        apply_review_bundle(rows, [{
            "guid": rows[0]["guid"],
            "remove_senses": "sem_specialized",
        }])


@pytest.mark.parametrize("bad_value", [None, "", {}, 0, False])
def test_apply_review_bundle_rejects_falsy_non_list_remove_senses(bad_value):
    rows = _audit_with_specialized_second_sense()

    with pytest.raises(ValueError, match="remove_senses must be a list"):
        apply_review_bundle(rows, [{
            "guid": rows[0]["guid"],
            "remove_senses": bad_value,
        }])
