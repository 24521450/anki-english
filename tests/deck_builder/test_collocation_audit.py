import copy
import json

import pytest
from openpyxl import load_workbook

from src.deck_builder.collocation_audit import (
    CARD_COLUMNS,
    CURRENT_REVIEW_COLUMNS,
    EVIDENCE_COLUMNS,
    apply_collocation_registry,
    build_audit_rows,
    collocation_final_item_id,
    collocation_text_matches_evidence,
    export_workbook,
    import_workbook,
    promote_audit_rows,
    registry_payload_by_guid,
    serialize_audit_rows,
    serialize_registry_rows,
    validate_audit_rows,
    validate_current_audit,
    validate_registry_rows,
    _validate_final_items,
)
from src.deck_builder.build_contracts import BuiltCard
from src.deck_builder.simplify_senses import _flatten_senses
from src.deck_builder.source_sense_identity import source_sense_id


def _registry(guid="g1"):
    return [{
        "guid": guid,
        "word": "curriculum",
        "cefr": "B2",
        "list": "Oxford_5000",
        "variant": "",
        "pos": "noun",
        "status": "active",
        "deck_override": None,
    }]


def _note(guid="g1", collocations=None, idioms=""):
    return {
        "guid": guid,
        "word": "curriculum",
        "cefr": "B2",
        "pos": "noun",
        "tags": "Oxford_5000 CEFR::B2",
        "collocations": collocations
        if collocations is not None
        else "school/national curriculum|curriculum development|on the curriculum",
        "idioms": idioms,
    }


def _source(source, evidence):
    record = {
        "word": "curriculum",
        "homonym_index": None,
        "source": source,
        "source_files": [f"{source}_curriculum.html"],
        "pos_data": [{
            "pos": "noun",
            "definitions": [{
                "n": 1,
                "sensenum_local": None,
                "text": "the subjects taught in a school or course",
                "cefr": "B2",
                "register_tags": [],
                "topics": [],
                "collocations": {"collocations": []},
                "examples": [
                    {"text": "Spanish is on the curriculum.", "cf": "on the curriculum"},
                    {"text": "Spanish is in the curriculum.", "cf": "in the curriculum"},
                ],
                "collocation_evidence": evidence,
                "is_phrase": False,
                "is_idiom": False,
            }],
        }],
    }
    flat = _flatten_senses(record)[0]
    return record, source_sense_id(record, flat)


def _inputs():
    oxford, oxford_id = _source("oxford", [
        {
            "text": "on the curriculum",
            "source": "oxford",
            "origin": "oxford_example_cf",
            "evidence_kind": "example_linked",
            "example_index": 1,
            "example_text": "Spanish is on the curriculum.",
            "container_index": 1,
            "item_index": 1,
            "category": None,
            "truncated": False,
            "full_entry_url": None,
        },
        {
            "text": "in the curriculum",
            "source": "oxford",
            "origin": "oxford_example_cf",
            "evidence_kind": "example_linked",
            "example_index": 2,
            "example_text": "Spanish is in the curriculum.",
            "container_index": 1,
            "item_index": 2,
            "category": None,
            "truncated": False,
            "full_entry_url": None,
        },
        {
            "text": "across the curriculum",
            "source": "oxford",
            "origin": "oxford_collocations_snippet",
            "evidence_kind": "supporting",
            "example_index": None,
            "example_text": None,
            "container_index": 2,
            "item_index": 1,
            "category": "preposition",
            "truncated": True,
            "full_entry_url": "https://example.test/collocations",
        },
    ])
    cambridge, cambridge_id = _source("cambridge", [
        {
            "text": "on the curriculum",
            "source": "cambridge",
            "origin": "cambridge_example_lu",
            "evidence_kind": "example_linked",
            "example_index": 1,
            "example_text": "Spanish is on the curriculum.",
            "container_index": 1,
            "item_index": 1,
            "category": None,
            "truncated": False,
            "full_entry_url": None,
        },
        {
            "text": "curriculum for schools",
            "source": "cambridge",
            "origin": "cambridge_bare_lu",
            "evidence_kind": "supporting",
            "example_index": None,
            "example_text": None,
            "container_index": 2,
            "item_index": 1,
            "category": None,
            "truncated": False,
            "full_entry_url": None,
        },
    ])
    semantic = [{
        "schema_version": 4,
        "guid": "g1",
        "word": "curriculum",
        "cefr": "B2",
        "list": "Oxford_5000",
        "variant": "",
        "pos": "noun",
        "senses": [{
            "semantic_sense_id": "sem_curriculum",
            "order": 1,
            "source_sense_ids": [oxford_id, cambridge_id],
        }],
    }]
    return [oxford], [cambridge], semantic


def _complete(row):
    by_text = {item["text"]: item for item in row["current_items"]}
    candidates = {item["text"]: item for item in row["mandatory_candidates"]}

    final_specs = [
        ("on the curriculum", "oxford+cambridge", candidates["on the curriculum"]["evidence_ids"], [by_text["on the curriculum"]["current_item_id"]]),
        ("in the curriculum", "oxford", candidates["in the curriculum"]["evidence_ids"], []),
        ("school/national curriculum", "curated", [], [by_text["school/national curriculum"]["current_item_id"]]),
        ("curriculum development", "curated", [], [by_text["curriculum development"]["current_item_id"]]),
    ]
    row["final_items"] = [
        {
            "final_item_id": collocation_final_item_id(row["guid"], text),
            "text": text,
            "order": order,
            "source": source,
            "evidence_ids": evidence_ids,
            "current_item_ids": current_ids,
        }
        for order, (text, source, evidence_ids, current_ids) in enumerate(final_specs, 1)
    ]
    final_by_text = {item["text"]: item for item in row["final_items"]}
    for item in row["current_items"]:
        if item["text"] == "on the curriculum":
            item["decision"] = "keep_source"
            item["reason"] = (
                "Retain exact source-backed surface on the curriculum; evidence "
                + ", ".join(item["evidence_ids"])
            )
        else:
            item["decision"] = "keep_curated"
            item["reason"] = f"Retain reviewed learner pattern: {item['text']}"
        item["target_final_item_ids"] = [final_by_text[item["text"]]["final_item_id"]]
        item["reviewer"] = "reviewer"
        item["reviewed_at"] = "2026-07-18"
        item["approval"] = "approved"
    for candidate in row["mandatory_candidates"]:
        candidate["decision"] = (
            "covered" if candidate["text"] == "on the curriculum" else "included"
        )
        candidate["reason"] = (
            f"Review exact candidate {candidate['text']}; evidence "
            + ", ".join(candidate["evidence_ids"])
        )
        candidate["target_final_item_ids"] = [
            final_by_text[candidate["text"]]["final_item_id"]
        ]
        candidate["reviewer"] = "reviewer"
        candidate["reviewed_at"] = "2026-07-18"
        candidate["approval"] = "approved"
    return row


def test_scaffold_accounts_for_current_items_and_only_example_linked_candidates():
    oxford, cambridge, semantic = _inputs()

    rows = build_audit_rows(
        [_note()], _registry(), semantic, oxford, cambridge
    )

    assert len(rows) == 1
    row = rows[0]
    assert [item["text"] for item in row["current_items"]] == [
        "school/national curriculum",
        "curriculum development",
        "on the curriculum",
    ]
    assert [item["text"] for item in row["mandatory_candidates"]] == [
        "on the curriculum",
        "in the curriculum",
    ]
    on = row["mandatory_candidates"][0]
    assert on["sources"] == ["oxford", "cambridge"]
    assert len(on["evidence_ids"]) == 2
    assert "across the curriculum" not in {
        item["text"] for item in row["mandatory_candidates"]
    }
    assert {item["text"] for item in row["source_evidence"]} == {
        "on the curriculum",
        "in the curriculum",
        "across the curriculum",
        "curriculum for schools",
    }
    assert {item["source_headword"] for item in row["source_evidence"]} == {
        "curriculum"
    }
    assert row["idiom_phrases"] == []
    assert len(row["idiom_fingerprint"]) == 64
    assert validate_audit_rows(rows, _registry()) == []


def test_scaffold_reuses_reviews_only_when_both_input_fingerprints_match():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])

    reused = build_audit_rows(
        [_note()], _registry(), semantic, oxford, cambridge, existing_rows=rows
    )
    assert reused[0]["final_items"] == rows[0]["final_items"]
    assert reused[0]["current_items"][0]["decision"] == "keep_curated"

    changed = build_audit_rows(
        [_note(collocations="school curriculum|on the curriculum")],
        _registry(), semantic, oxford, cambridge, existing_rows=rows,
    )
    assert changed[0]["final_items"] == []
    assert {item["decision"] for item in changed[0]["current_items"]} == {"pending"}

    changed_idioms = build_audit_rows(
        [_note(idioms="on the curriculum :: a reviewed idiom")],
        _registry(), semantic, oxford, cambridge, existing_rows=rows,
    )
    assert changed_idioms[0]["final_items"] == []
    assert {item["decision"] for item in changed_idioms[0]["current_items"]} == {
        "pending"
    }


def test_live_idiom_change_makes_completed_audit_stale():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])

    errors = validate_current_audit(
        rows,
        [_note(idioms="on the curriculum :: a reviewed idiom")],
        _registry(),
        semantic,
        oxford,
        cambridge,
        require_complete=True,
    )

    assert "stale_collocation_audit_projection" in errors


def test_completed_audit_accepts_the_exact_post_promotion_projection():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])
    promoted = promote_audit_rows(rows, _registry())
    projected_card = apply_collocation_registry([_note()], promoted)[0]

    assert validate_current_audit(
        rows,
        [projected_card],
        _registry(),
        semantic,
        oxford,
        cambridge,
        require_complete=True,
    ) == []


def test_unreviewed_live_collocation_projection_remains_stale():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])
    promoted = promote_audit_rows(rows, _registry())
    projected_card = apply_collocation_registry([_note()], promoted)[0]
    projected_card["collocations"] = "unreviewed phrase"

    errors = validate_current_audit(
        rows,
        [projected_card],
        _registry(),
        semantic,
        oxford,
        cambridge,
        require_complete=True,
    )

    assert "stale_collocation_audit_projection" in errors


def test_require_complete_rejects_exact_idiom_collocation_overlap():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows(
        [_note(idioms="on the curriculum :: a reviewed idiom")],
        _registry(),
        semantic,
        oxford,
        cambridge,
    )
    _complete(rows[0])

    errors = validate_audit_rows(rows, _registry(), require_complete=True)

    assert any(error.startswith("collocation_duplicates_idiom:g1:") for error in errors)


def test_complete_validation_and_promotion_are_fail_closed_and_deterministic():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    assert any(
        error.startswith("unresolved_current_item:")
        for error in validate_audit_rows(rows, _registry(), require_complete=True)
    )

    _complete(rows[0])
    assert validate_audit_rows(rows, _registry(), require_complete=True) == []
    promoted = promote_audit_rows(rows, _registry())
    assert promoted[0]["idiom_fingerprint"] == rows[0]["idiom_fingerprint"]
    assert validate_registry_rows(promoted, _registry(), audit_rows=rows) == []
    assert serialize_registry_rows(promoted) == serialize_registry_rows(
        promote_audit_rows(copy.deepcopy(rows), list(reversed(_registry())))
    )
    payload = registry_payload_by_guid(promoted)["g1"]
    assert payload == {
        "collocations": "on the curriculum|in the curriculum|school/national curriculum|curriculum development",
        "collocation_sources": "oxford+cambridge|oxford|curated|curated",
    }

    oversized = copy.deepcopy(rows)
    extra = copy.deepcopy(oversized[0]["final_items"][-1])
    extra["text"] = "curriculum reform"
    extra["final_item_id"] = collocation_final_item_id("g1", extra["text"])
    extra["order"] = 5
    extra["current_item_ids"] = []
    oversized[0]["final_items"].append(extra)
    another = copy.deepcopy(extra)
    another["text"] = "curriculum policy"
    another["final_item_id"] = collocation_final_item_id("g1", another["text"])
    another["order"] = 6
    oversized[0]["final_items"].append(another)
    assert "too_many_final_items:g1:6" in validate_audit_rows(oversized, _registry())


def test_source_backed_final_must_be_exact_separate_and_evidence_bound():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])

    bad = copy.deepcopy(rows)
    final = bad[0]["final_items"][0]
    final["text"] = "on/in the curriculum"
    final["final_item_id"] = collocation_final_item_id("g1", final["text"])
    errors = validate_audit_rows(bad, _registry(), require_complete=True)
    assert any(error.startswith("source_item_contains_slash:") for error in errors)
    assert any(error.startswith("source_evidence_text_mismatch:") for error in errors)

    bad = copy.deepcopy(rows)
    bad[0]["final_items"][0]["evidence_ids"] = []
    assert any(
        error.startswith("source_item_without_evidence:")
        for error in validate_audit_rows(bad, _registry(), require_complete=True)
    )


def test_source_evidence_binding_allows_only_headword_number_inflection():
    assert collocation_text_matches_evidence(
        "generous portion",
        "generous portions",
        headword="portion",
    )
    assert not collocation_text_matches_evidence(
        "individual portion",
        "individual portions served",
        headword="portion",
    )
    assert not collocation_text_matches_evidence(
        "generous serving",
        "generous servings",
        headword="portion",
    )


def test_source_binding_prefers_exact_same_source_evidence_before_inflection():
    row = {
        "guid": "g1",
        "word": "loyalty",
        "source_evidence": [
            {
                "evidence_id": "cambridge-singular",
                "text": "loyalty to",
                "source": "cambridge",
            },
            {
                "evidence_id": "cambridge-plural",
                "text": "loyalties to",
                "source": "cambridge",
            },
            {
                "evidence_id": "oxford-singular",
                "text": "loyalty to",
                "source": "oxford",
            },
        ],
        "current_items": [],
        "mandatory_candidates": [],
        "final_items": [{
            "final_item_id": collocation_final_item_id("g1", "loyalty to"),
            "text": "loyalty to",
            "order": 1,
            "source": "cambridge",
            "evidence_ids": ["cambridge-singular"],
            "current_item_ids": [],
        }],
    }

    assert _validate_final_items(row, require_complete=False) == []


def test_source_binding_uses_headword_inflection_when_no_exact_surface_exists():
    row = {
        "guid": "g1",
        "word": "portion",
        "source_evidence": [{
            "evidence_id": "cambridge-plural",
            "text": "generous portions",
            "source": "cambridge",
        }],
        "current_items": [],
        "mandatory_candidates": [],
        "final_items": [{
            "final_item_id": collocation_final_item_id("g1", "generous portion"),
            "text": "generous portion",
            "order": 1,
            "source": "cambridge",
            "evidence_ids": ["cambridge-plural"],
            "current_item_ids": [],
        }],
    }

    assert _validate_final_items(row, require_complete=False) == []


def test_final_order_numbers_must_match_serialized_list_order():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])
    tampered = copy.deepcopy(rows)
    tampered[0]["final_items"][0]["order"], tampered[0]["final_items"][1]["order"] = 2, 1

    errors = validate_audit_rows(tampered, _registry(), require_complete=True)

    assert "non_contiguous_final_order:g1" in errors

    promoted = promote_audit_rows(rows, _registry())
    registry_tampered = copy.deepcopy(promoted)
    registry_tampered[0]["items"][0]["order"], registry_tampered[0]["items"][1]["order"] = 2, 1
    assert "non_contiguous_collocation_registry_order:g1" in validate_registry_rows(
        registry_tampered, _registry()
    )
    with pytest.raises(ValueError, match="Non-contiguous Collocation Registry order"):
        registry_payload_by_guid(registry_tampered)


def test_workbook_round_trip_rejects_immutable_changes_and_imports_item_decisions(tmp_path):
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    path = tmp_path / "collocations.xlsx"
    export_workbook(rows, path)

    workbook = load_workbook(path)
    sheet = workbook["Current Items"]
    decision_col = CURRENT_REVIEW_COLUMNS.index("decision") + 1
    sheet.cell(2, decision_col).value = "remove"
    reason_col = CURRENT_REVIEW_COLUMNS.index("reason") + 1
    sheet.cell(2, reason_col).value = "This exact current item is not useful."
    workbook.save(path)
    imported = import_workbook(rows, path)
    assert imported[0]["current_items"][0]["decision"] == "remove"
    assert rows[0]["current_items"][0]["decision"] == "pending"

    workbook = load_workbook(path)
    text_col = CURRENT_REVIEW_COLUMNS.index("text") + 1
    workbook["Current Items"].cell(2, text_col).value = "tampered"
    workbook.save(path)
    with pytest.raises(ValueError, match="Immutable Current Items column 'text'"):
        import_workbook(rows, path)

    export_workbook(rows, path)
    workbook = load_workbook(path)
    idioms_col = CARD_COLUMNS.index("idiom_phrases") + 1
    workbook["Cards"].cell(2, idioms_col).value = '["on the curriculum"]'
    workbook.save(path)
    with pytest.raises(ValueError, match="Immutable Cards column 'idiom_phrases'"):
        import_workbook(rows, path)

    export_workbook(rows, path)
    workbook = load_workbook(path)
    headword_col = EVIDENCE_COLUMNS.index("source_headword") + 1
    workbook["Evidence"].cell(2, headword_col).value = "tampered"
    workbook.save(path)
    with pytest.raises(ValueError, match="Immutable Evidence column 'source_headword'"):
        import_workbook(rows, path)


def test_source_headword_is_fingerprint_bound_evidence_provenance():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    tampered = copy.deepcopy(rows)
    tampered[0]["source_evidence"][0]["source_headword"] = "time"

    errors = validate_audit_rows(tampered, _registry())

    assert any(error.startswith("evidence_contract_mismatch:g1:") for error in errors)
    assert "source_fingerprint_mismatch:g1" in errors


def test_serialized_audit_is_canonical_and_sorted():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    payload = serialize_audit_rows(rows)
    parsed = [json.loads(line) for line in payload.splitlines()]
    assert parsed == rows
    assert payload.endswith("\n")


def test_registry_validation_rejects_missing_guid_and_misaligned_source_claim():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])
    promoted = promote_audit_rows(rows, _registry())

    assert any(
        error == "missing_collocation_registry_guid:g1"
        for error in validate_registry_rows([], _registry())
    )
    malformed = copy.deepcopy(promoted)
    malformed[0]["items"][0]["source"] = "curated"
    errors = validate_registry_rows(malformed, _registry())
    assert any(error.startswith("curated_collocation_registry_evidence:") for error in errors)

    stale = validate_registry_rows(
        promoted,
        _registry(),
        current_source_fingerprints={
            "g1": {
                "source_fingerprint": rows[0]["source_fingerprint"],
                "idiom_fingerprint": "0" * 64,
            }
        },
    )
    assert "stale_collocation_idiom_fingerprint:g1" in stale


def test_scaffold_rejects_malformed_example_linked_evidence_contract():
    oxford, cambridge, semantic = _inputs()
    malformed = copy.deepcopy(oxford)
    malformed[0]["pos_data"][0]["definitions"][0]["collocation_evidence"][0]["example_index"] = 99
    with pytest.raises(ValueError, match="example index is out of range"):
        build_audit_rows([_note()], _registry(), semantic, malformed, cambridge)


def test_scaffold_ignores_empty_source_placeholders_without_headwords():
    oxford, cambridge, semantic = _inputs()
    cambridge.append({
        "word": None,
        "source": "cambridge",
        "source_files": ["cambridge_empty.html"],
        "pos_data": [],
    })

    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)

    assert validate_audit_rows(rows, _registry()) == []


def test_apply_registry_updates_dicts_and_built_cards_with_exact_guid_coverage():
    oxford, cambridge, semantic = _inputs()
    rows = build_audit_rows([_note()], _registry(), semantic, oxford, cambridge)
    _complete(rows[0])
    promoted = promote_audit_rows(rows, _registry())
    card = BuiltCard(
        guid="g1",
        notetype="English Academic Vocabulary Model",
        deck="deck",
        word="curriculum",
        pos="noun",
        ipa="",
        definition="definition",
        example="example",
        collocations="legacy",
        wordfamily="",
        uk_audio="",
        us_audio="",
        source1="Oxford",
        source2="Oxford",
        cefr="B2",
        idioms="",
        tags="Oxford_5000",
        synonyms="",
        antonyms="",
    )

    updated_card = apply_collocation_registry([card], promoted)[0]
    assert updated_card.collocations.startswith("on the curriculum|in the curriculum")
    assert updated_card.collocation_sources == "oxford+cambridge|oxford|curated|curated"
    updated_dict = apply_collocation_registry([{"guid": "g1"}], promoted)[0]
    assert updated_dict["collocation_sources"] == updated_card.collocation_sources

    with pytest.raises(ValueError, match="unknown built-card GUIDs"):
        apply_collocation_registry([], promoted)
